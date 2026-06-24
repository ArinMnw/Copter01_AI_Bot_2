import sys, os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
sys.stdout.reconfigure(encoding="utf-8")

import argparse
import csv
import json
import re
import time
from collections import Counter, defaultdict
from datetime import datetime, timezone, timedelta

import MetaTrader5 as mt5

import config
import sim_s1_backtest
import sim_s2_backtest
import sim_s3_backtest
import sim_s4_backtest
import sim_s5_backtest
import sim_s458_backtest
import sim_s8_backtest
import sim_s9_backtest
import sim_s10_backtest
import sim_s11_backtest
import sim_s12_backtest
import sim_s13_backtest
import sim_s14_backtest
import sim_s15_backtest
import sim_s16_backtest
import sim_s17_backtest
import sim_s18_backtest
import sim_s19_backtest
import sim_lifecycle
from sim_s10_backtest import (
    TF_MAP as S10_TF_MAP,
    backtest_tf as backtest_s10_tf,
    s10_runtime_feature_coverage,
    s10_unreplayed_active_features,
    sync_strategy10_runtime_config,
    to_bkk,
)
from sim_s1_backtest import (
    TF_MAP as S1_TF_MAP,
    backtest_tf as backtest_s1_tf,
    s1_runtime_feature_coverage,
    s1_unreplayed_active_features,
)
from sim_s2_backtest import (
    TF_MAP as S2_TF_MAP,
    backtest_tf as backtest_s2_tf,
    s2_runtime_feature_coverage,
    s2_unreplayed_active_features,
)
from sim_s3_backtest import (
    TF_MAP as S3_TF_MAP,
    backtest_tf as backtest_s3_tf,
    s3_runtime_feature_coverage,
    s3_unreplayed_active_features,
)
from sim_s4_backtest import (
    TF_MAP as S4_TF_MAP,
    backtest_tf as backtest_s4_tf,
    s4_runtime_feature_coverage,
    s4_unreplayed_active_features,
)
from sim_s5_backtest import (
    TF_MAP as S5_TF_MAP,
    backtest_tf as backtest_s5_tf,
    s5_runtime_feature_coverage,
    s5_unreplayed_active_features,
)
from sim_s8_backtest import (
    TF_MAP as S8_TF_MAP,
    backtest_tf as backtest_s8_tf,
    s8_runtime_feature_coverage,
    s8_unreplayed_active_features,
)
from sim_s9_backtest import (
    TF_MAP as S9_TF_MAP,
    backtest_tf as backtest_s9_tf,
    s9_runtime_feature_coverage,
    s9_unreplayed_active_features,
)
from sim_s11_backtest import (
    TF_MAP as S11_TF_MAP,
    backtest_tf as backtest_s11_tf,
    s11_runtime_feature_coverage,
    s11_unreplayed_active_features,
)
from sim_s12_backtest import (
    TF_MAP as S12_TF_MAP,
    backtest_tf as backtest_s12_tf,
    s12_runtime_feature_coverage,
    s12_unreplayed_active_features,
)
from sim_s14_backtest import (
    TF_MAP as S14_TF_MAP,
    backtest_tf as backtest_s14_tf,
    s14_runtime_feature_coverage,
    s14_unreplayed_active_features,
)
from sim_s13_backtest import (
    TF_MAP as S13_TF_MAP,
    backtest_tf as backtest_s13_tf,
    s13_runtime_feature_coverage,
    s13_unreplayed_active_features,
)
from sim_s15_backtest import (
    TF_MAP as S15_TF_MAP,
    backtest_tf as backtest_s15_tf,
    s15_runtime_feature_coverage,
    s15_unreplayed_active_features,
)
from sim_s16_backtest import (
    TF_MAP as S16_TF_MAP,
    backtest_tf as backtest_s16_tf,
    s16_runtime_feature_coverage,
    s16_unreplayed_active_features,
)
from sim_s17_backtest import (
    TF_MAP as S17_TF_MAP,
    backtest_tf as backtest_s17_tf,
    s17_runtime_feature_coverage,
    s17_unreplayed_active_features,
)
from sim_s18_backtest import (
    TF_MAP as S18_TF_MAP,
    backtest_tf as backtest_s18_tf,
    s18_runtime_feature_coverage,
    s18_unreplayed_active_features,
)
from sim_s19_backtest import (
    TF_MAP as S19_TF_MAP,
    backtest_tf as backtest_s19_tf,
    s19_runtime_feature_coverage,
    s19_unreplayed_active_features,
)


HTF_TO_LTF = {
    "D1": "M15",
    "H12": "M15",
    "H4": "M5",
    "H1": "M1",
    "M30": "M1",
    "M15": "M1",
}

SUPPORTED_STRATEGIES = {1, 2, 3, 4, 5, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19}
ALL_STRATEGIES = set(range(1, 20))
RUN_STARTED_AT = time.perf_counter()
COMPARE_REPORT_DIR = os.path.join("excel_reports", "backtest_compare")
SCALE_OUT_COLUMNS = 4
FILL_TREND_RECHECK_SKIP_SIDS = {1, 2, 3, 9, 10, 11, 14, 15, 16, 17, 18, 19}


def elapsed() -> str:
    seconds = int(time.perf_counter() - RUN_STARTED_AT)
    return f"{seconds // 60:02d}:{seconds % 60:02d}"


def progress(message: str) -> None:
    print(f"[{elapsed()}] {message}", flush=True)


def scale_out_column_volume(symbol: str) -> float:
    return 0.04 if (symbol or "").upper().startswith("BTCUSD") else 0.01


def split_scale_out_pnl(deals: list[dict], column_volume: float, columns: int = SCALE_OUT_COLUMNS) -> list[float]:
    pnl_cols = [0.0 for _ in range(columns)]
    filled = [0.0 for _ in range(columns)]
    if column_volume <= 0:
        return pnl_cols

    for deal in sorted(deals or [], key=lambda d: d.get("time") or datetime.min):
        volume_left = float(deal.get("volume", 0.0) or 0.0)
        profit = float(deal.get("profit", 0.0) or 0.0)
        original_volume = volume_left
        if volume_left <= 0:
            continue
        for idx in range(columns):
            room = max(0.0, column_volume - filled[idx])
            if room <= 0:
                continue
            take = min(room, volume_left)
            if take <= 0:
                continue
            pnl_cols[idx] += profit * (take / original_volume)
            filled[idx] += take
            volume_left -= take
            if volume_left <= 1e-9:
                break
    return [round(v, 2) for v in pnl_cols]


def parse_bkk_dt(dt_str: str) -> datetime:
    for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M", "%Y-%m-%d %H:%M"):
        try:
            return datetime.strptime(dt_str, fmt)
        except ValueError:
            continue
    raise ValueError(f"Invalid datetime format: {dt_str}")


def parse_strategy_list(raw: str) -> list[int]:
    raw = (raw or "").strip().lower()
    if raw in ("", "active"):
        return [sid for sid in sorted(ALL_STRATEGIES) if config.active_strategies.get(sid, False)]
    if raw == "all":
        return list(range(1, 17))

    result = set()
    for part in raw.replace(" ", "").split(","):
        if not part:
            continue
        if "-" in part:
            a, b = part.split("-", 1)
            result.update(range(int(a), int(b) + 1))
        else:
            result.add(int(part))
    invalid = sorted(s for s in result if s not in ALL_STRATEGIES)
    if invalid:
        raise ValueError(f"Invalid strategy ids: {invalid}")
    return sorted(result)


def field(line: str, key: str) -> str:
    m = re.search(rf"{re.escape(key)}=([^|\s]+)", line)
    return m.group(1) if m else ""


def parse_log_ts(raw: str) -> datetime | None:
    try:
        return datetime.strptime(raw, "%Y-%m-%d %H:%M:%S")
    except ValueError:
        return None


def _excel_dt(value):
    if isinstance(value, datetime) and value.tzinfo is not None:
        return value.replace(tzinfo=None)
    return value


def _trail_path(events: list[dict], time_key: str, source_key: str, sl_key: str, limit: int = 8) -> str:
    parts = []
    for event in events[-limit:]:
        ts = _excel_dt(event.get(time_key))
        ts_text = ts.strftime("%m-%d %H:%M") if isinstance(ts, datetime) else str(ts or "")
        source = str(event.get(source_key, "") or "")
        sl = event.get(sl_key, "")
        try:
            sl_text = f"{float(sl):.2f}"
        except (TypeError, ValueError):
            sl_text = str(sl or "")
        parts.append(f"{ts_text} {source} {sl_text}".strip())
    return " -> ".join(parts)


def load_state_symbol() -> str:
    try:
        if os.path.exists(config.STATE_FILE):
            with open(config.STATE_FILE, "r", encoding="utf-8") as f:
                return str((json.load(f) or {}).get("symbol", "") or "")
    except Exception:
        return ""
    return ""


def default_log_files() -> list[str]:
    from log_sources import backtest_log_files
    return backtest_log_files()


def validate_backtest_log_files(paths: list[str]) -> None:
    forbidden = []
    for path in paths or []:
        name = os.path.basename(str(path)).lower()
        if name in {"bot.log", "system.log", "error.log"}:
            forbidden.append(path)
            continue
        if (
            (name.startswith("bot-") or name.startswith("system-") or name.startswith("error-"))
            and not name.startswith("backtest_")
        ):
            forbidden.append(path)
    if forbidden:
        raise SystemExit(
            "Backtest must not use live bot logs. Use logs/backtest_bot.log, "
            "logs/backtest_system.log, logs/backtest_error.log instead. "
            f"Forbidden log file(s): {', '.join(map(str, forbidden))}"
        )


def default_compare_report_base(
    start_bkk: datetime,
    end_bkk: datetime,
    tf: str | None,
    strategies: list[int],
    variant: str | None = None,
) -> str:
    tf_part = (tf or "ALL").upper()
    sid_part = "-".join(str(s) for s in strategies) or "active"
    start_part = start_bkk.strftime("%Y%m%d_%H%M")
    end_part = end_bkk.strftime("%Y%m%d_%H%M")
    suffix = f"_{variant}" if variant else ""
    return f"compare_s{sid_part}_{tf_part}_{start_part}_{end_part}{suffix}"


def strategy_report_dir(strategies: list[int]) -> str:
    sid_part = "-".join(str(s) for s in strategies) or "active"
    return os.path.join(COMPARE_REPORT_DIR, f"s{sid_part}")


def resolve_compare_output_path(raw_path: str | None, default_name: str, ext: str, strategies: list[int]) -> str:
    base_dir = strategy_report_dir(strategies)
    if raw_path is None:
        return os.path.join(base_dir, default_name + ext)
    raw_path = raw_path.strip()
    if raw_path == "":
        return os.path.join(base_dir, default_name + ext)
    root, raw_ext = os.path.splitext(raw_path)
    if raw_ext == "":
        raw_path = raw_path + ext
    if os.path.dirname(raw_path):
        return raw_path
    return os.path.join(base_dir, raw_path)


def _csv_value(value):
    if isinstance(value, datetime):
        return _excel_dt(value).strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, (list, tuple, set)):
        return "|".join(str(v) for v in value)
    if isinstance(value, dict):
        return json.dumps(value, ensure_ascii=False, sort_keys=True)
    return value


def _capture_raw_replay_context(args, rows: list[tuple[str, dict]]) -> None:
    bucket = getattr(args, "_raw_replay_context_trades", None)
    if bucket is not None:
        bucket.extend(rows)


def write_trades_csv(path: str, trades: list[tuple[str, dict]]) -> str:
    headers = [
        "report_tf", "sid", "tf", "lifecycle_tf", "signal", "pattern",
        "entry_time", "close_time", "close_type", "cancel_reason",
        "entry", "sl", "tp", "close_price", "pnl", "profit",
        "parallel_tfs", "parallel_patterns", "gap_bot", "gap_top",
        "final_gap_bot", "final_gap_top", "detect_time_raw", "entry_time_raw",
        "cancel_age_bars", "cancel_bars", "cancel_bar_high", "cancel_bar_low", "cancel_bar_touched_entry",
        "pd_h", "pd_l", "pd_fib_382", "pd_fib_618", "pd_fallback_used", "pd_outside_range",
        "sweep_scan_state", "sweep_scan_tf",
        "sweep_scan_price", "sweep_scan_time", "sweep_scan_ts",
        "sweep_scan_age_min", "sweep_scan_expiry_min",
        "sl_guard_scope", "sl_guard_key", "sl_guard_count", "sl_guard_since", "sl_guard_swing_ref",
    ]
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    try:
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.DictWriter(f, fieldnames=headers, extrasaction="ignore")
            w.writeheader()
            for report_tf, trade in trades:
                row = dict(trade)
                row["report_tf"] = report_tf
                w.writerow({key: _csv_value(row.get(key, "")) for key in headers})
        return path
    except PermissionError:
        root, ext = os.path.splitext(path)
        fallback = f"{root}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{ext or '.csv'}"
        with open(fallback, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.DictWriter(f, fieldnames=headers, extrasaction="ignore")
            w.writeheader()
            for report_tf, trade in trades:
                row = dict(trade)
                row["report_tf"] = report_tf
                w.writerow({key: _csv_value(row.get(key, "")) for key in headers})
        return fallback


def resolve_run_tfs(tf_arg: str | None) -> list[str]:
    if not tf_arg:
        return list(S10_TF_MAP.keys())
    target = tf_arg.upper()
    run_tfs = []
    if target in S10_TF_MAP:
        run_tfs.append(target)
    if target in HTF_TO_LTF and HTF_TO_LTF[target] not in run_tfs:
        run_tfs.append(HTF_TO_LTF[target])
    if not run_tfs:
        raise ValueError(f"Unsupported timeframe for current S10 replay: {tf_arg}")
    return run_tfs


def resolve_run_tfs_for_strategy(strategy_id: int, tf_arg: str | None) -> list[str]:
    if strategy_id == 1:
        if not tf_arg:
            return list(S1_TF_MAP.keys())
        target = tf_arg.upper()
        if target in S1_TF_MAP:
            return [target]
        raise ValueError(f"Unsupported timeframe for S1 replay: {tf_arg}")
    if strategy_id == 2:
        if not tf_arg:
            return list(S2_TF_MAP.keys())
        target = tf_arg.upper()
        if target in S2_TF_MAP:
            return [target]
        raise ValueError(f"Unsupported timeframe for S2 replay: {tf_arg}")
    if strategy_id == 3:
        if not tf_arg:
            return list(S3_TF_MAP.keys())
        target = tf_arg.upper()
        if target in S3_TF_MAP:
            return [target]
        raise ValueError(f"Unsupported timeframe for S3 replay: {tf_arg}")
    if strategy_id == 4:
        if not tf_arg:
            return list(S4_TF_MAP.keys())
        target = tf_arg.upper()
        if target in S4_TF_MAP:
            return [target]
        raise ValueError(f"Unsupported timeframe for S4 replay: {tf_arg}")
    if strategy_id == 5:
        if not tf_arg:
            return list(S5_TF_MAP.keys())
        target = tf_arg.upper()
        if target in S5_TF_MAP:
            return [target]
        raise ValueError(f"Unsupported timeframe for S5 replay: {tf_arg}")
    if strategy_id == 8:
        if not tf_arg:
            return list(S8_TF_MAP.keys())
        target = tf_arg.upper()
        if target in S8_TF_MAP:
            return [target]
        raise ValueError(f"Unsupported timeframe for S8 replay: {tf_arg}")
    if strategy_id == 9:
        if not tf_arg:
            return list(S9_TF_MAP.keys())
        target = tf_arg.upper()
        if target in S9_TF_MAP:
            return [target]
        raise ValueError(f"Unsupported timeframe for S9 replay: {tf_arg}")
    if strategy_id == 10:
        return resolve_run_tfs(tf_arg)
    if strategy_id == 11:
        if not tf_arg:
            return list(S11_TF_MAP.keys())
        target = tf_arg.upper()
        if target in S11_TF_MAP:
            return [target]
        raise ValueError(f"Unsupported timeframe for S11 replay: {tf_arg}")
    if strategy_id == 12:
        if not tf_arg:
            return list(S12_TF_MAP.keys())
        target = tf_arg.upper()
        if target in S12_TF_MAP:
            return [target]
        raise ValueError(f"Unsupported timeframe for S12 replay: {tf_arg}")
    if strategy_id == 13:
        if not tf_arg:
            return list(S13_TF_MAP.keys())
        target = tf_arg.upper()
        if target in S13_TF_MAP:
            return [target]
        raise ValueError(f"Unsupported timeframe for S13 replay: {tf_arg}")
    if strategy_id == 14:
        if not tf_arg:
            return list(S14_TF_MAP.keys())
        target = tf_arg.upper()
        if target in S14_TF_MAP:
            return [target]
        raise ValueError(f"Unsupported timeframe for S14 replay: {tf_arg}")
    if strategy_id == 15:
        if not tf_arg:
            return list(S15_TF_MAP.keys())
        target = tf_arg.upper()
        if target in S15_TF_MAP:
            return [target]
        raise ValueError(f"Unsupported timeframe for S15 replay: {tf_arg}")
    if strategy_id == 16:
        if not tf_arg:
            return list(S16_TF_MAP.keys())
        target = tf_arg.upper()
        if target in S16_TF_MAP:
            return [target]
        raise ValueError(f"Unsupported timeframe for S16 replay: {tf_arg}")
    if strategy_id == 17:
        if not tf_arg:
            return list(S17_TF_MAP.keys())
        target = tf_arg.upper()
        if target in S17_TF_MAP:
            return [target]
        raise ValueError(f"Unsupported timeframe for S17 replay: {tf_arg}")
    if strategy_id == 18:
        if not tf_arg:
            return list(S18_TF_MAP.keys())
        target = tf_arg.upper()
        if target in S18_TF_MAP:
            return [target]
        raise ValueError(f"Unsupported timeframe for S18 replay: {tf_arg}")
    if strategy_id == 19:
        if not tf_arg:
            return list(S19_TF_MAP.keys())
        target = tf_arg.upper()
        if target in S19_TF_MAP:
            return [target]
        raise ValueError(f"Unsupported timeframe for S19 replay: {tf_arg}")
    raise ValueError(f"Strategy S{strategy_id} is not implemented in this replay engine yet")


def resolve_s14_context_tfs(tf_arg: str | None) -> list[str]:
    base = resolve_run_tfs_for_strategy(14, tf_arg)
    if not tf_arg or not getattr(config, "SL_GUARD_GROUP_ENABLED", False):
        return base

    wanted = tf_arg.upper()
    context = set(base)
    for group in getattr(config, "SL_GUARD_GROUP_GROUPS", []) or []:
        if wanted in group:
            context.update(tf for tf in group if tf in S14_TF_MAP)
    return [tf for tf in S14_TF_MAP.keys() if tf in context]


def resolve_s8_context_tfs(tf_arg: str | None) -> list[str]:
    base = resolve_run_tfs_for_strategy(8, tf_arg)
    if not tf_arg or not getattr(config, "SL_GUARD_GROUP_ENABLED", False):
        return base

    wanted = tf_arg.upper()
    context = set(base)
    for group in getattr(config, "SL_GUARD_GROUP_GROUPS", []) or []:
        if wanted in group:
            context.update(tf for tf in group if tf in S8_TF_MAP)
    return [tf for tf in S8_TF_MAP.keys() if tf in context]


def resolve_guard_context_tfs(strategy_id: int, tf_arg: str | None, tf_map: dict[str, int]) -> list[str]:
    base = resolve_run_tfs_for_strategy(strategy_id, tf_arg)
    if not tf_arg or not getattr(config, "SL_GUARD_GROUP_ENABLED", False):
        return base

    wanted = tf_arg.upper()
    context = set(base)
    for group in getattr(config, "SL_GUARD_GROUP_GROUPS", []) or []:
        if wanted in group:
            context.update(tf for tf in group if tf in tf_map)
    return [tf for tf in tf_map.keys() if tf in context]


def resolve_s458_context_tfs(
    tf_arg: str | None,
    strategies: set[int] | None = None,
    *,
    include_connected_s2_context: bool = False,
) -> tuple[list[str], list[str]]:
    common_tfs = set(S4_TF_MAP) & set(S5_TF_MAP) & set(S8_TF_MAP)
    if tf_arg:
        target = tf_arg.upper()
        if target not in common_tfs:
            raise ValueError(f"Unsupported S1-S5/S8 unified timeframe: {tf_arg}")
        requested = [target]
    else:
        requested = [tf for tf in S8_TF_MAP.keys() if tf in common_tfs]

    context = set(requested)
    include_guard_context = bool(strategies and 8 in strategies)
    include_s2_parallel_context = bool(strategies and 2 in strategies and getattr(config, "FVG_PARALLEL", False))
    if tf_arg and include_guard_context and getattr(config, "SL_GUARD_GROUP_ENABLED", False):
        wanted = tf_arg.upper()
        for group in getattr(config, "SL_GUARD_GROUP_GROUPS", []) or []:
            if wanted in group:
                context.update(tf for tf in group if tf in common_tfs)
    if tf_arg and include_s2_parallel_context:
        wanted = tf_arg.upper()
        for group in getattr(config, "FVG_PARALLEL_GROUPS", []) or []:
            group_tfs = {tf for tf in group if tf in common_tfs}
            if wanted in group_tfs:
                context.update(group_tfs)

        if include_connected_s2_context:
            direct_context = set(context)
            max_direct_secs = max(sim_s458_backtest.TF_SECONDS.get(tf, 0) for tf in direct_context)
            for group in getattr(config, "FVG_PARALLEL_GROUPS", []) or []:
                group_tfs = {tf for tf in group if tf in common_tfs}
                if not group_tfs or not (group_tfs & direct_context):
                    continue
                if max(sim_s458_backtest.TF_SECONDS.get(tf, 0) for tf in group_tfs) <= max_direct_secs:
                    context.update(group_tfs)

    run_tfs = [tf for tf in S8_TF_MAP.keys() if tf in context]
    return requested, run_tfs


def _s458_tf_history_overlaps_window(tf_name: str, window_start_utc: datetime, window_end_utc: datetime) -> bool:
    tf_val = S8_TF_MAP.get(tf_name)
    if tf_val is None:
        return False
    real_start_utc = window_start_utc - timedelta(hours=getattr(config, "TZ_OFFSET", 7))
    real_end_utc = window_end_utc - timedelta(hours=getattr(config, "TZ_OFFSET", 7))
    rates = mt5.copy_rates_range(config.SYMBOL, tf_val, real_start_utc, real_end_utc)
    if rates is None or len(rates) == 0:
        return False
    oldest = int(min(r["time"] for r in rates))
    newest = int(max(r["time"] for r in rates))
    window_start_ts = int(real_start_utc.timestamp())
    window_end_ts = int(real_end_utc.timestamp())
    return oldest <= window_end_ts and newest >= window_start_ts


def _trade_tf_members(trade: dict, default_tf: str = "") -> set[str]:
    values = trade.get("tf_members") or trade.get("parallel_tfs") or []
    members = {str(tf).upper() for tf in values if tf}
    if default_tf:
        members.add(str(default_tf).upper())
    trade_tf = str(trade.get("tf", "") or "").upper()
    if trade_tf:
        members.add(trade_tf)
    return members


def _s2_multi_trade_matches_requested_tf(trade: dict, default_tf: str, requested_set: set[str]) -> bool:
    if not requested_set:
        return True
    if int(trade.get("sid", 0) or 0) != 2:
        return str(default_tf).upper() in requested_set
    return bool(_trade_tf_members(trade, default_tf) & requested_set)


def _s2_multi_report_tf(trade: dict, default_tf: str, requested_set: set[str]) -> str:
    default_tf = str(default_tf or trade.get("tf") or "").upper()
    if default_tf in requested_set:
        return default_tf
    for tf_name in trade.get("parallel_tfs") or trade.get("tf_members") or []:
        tf_name = str(tf_name).upper()
        if tf_name in requested_set:
            return tf_name
    return default_tf


def _replay_range_end_utc(window_end_utc: datetime, extra_days: int) -> datetime:
    real_end_utc = window_end_utc - timedelta(hours=getattr(config, "TZ_OFFSET", 7))
    return real_end_utc + timedelta(days=max(0, int(extra_days or 0)))


def load_live_filled_orders(log_files: list[str], start_bkk: datetime, end_bkk: datetime, symbol: str, strategies: set[int]) -> list[dict]:
    progress(f"Loading live orders from logs ({len(log_files)} files)...")
    fills = {}
    closes = {}
    seen_close = set()
    wanted_symbol = (symbol or "").split(".", 1)[0]

    for path in log_files:
        if not os.path.exists(path):
            continue
        with open(path, encoding="utf-8", errors="replace") as f:
            for line in f:
                m = re.match(r"^\[(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2})\]\s+(\S+)", line)
                if not m:
                    continue
                ts_raw, kind = m.group(1), m.group(2)
                ts = parse_log_ts(ts_raw)
                if ts is None:
                    continue
                ticket = field(line, "ticket")
                if not ticket:
                    continue

                if kind == "ENTRY_FILL":
                    sid_raw = field(line, "sid")
                    try:
                        sid = int(sid_raw)
                    except ValueError:
                        continue
                    if sid not in strategies:
                        continue
                    line_symbol = field(line, "symbol")
                    if line_symbol and wanted_symbol and wanted_symbol not in line_symbol:
                        continue
                    if ticket not in fills:
                        fills[ticket] = {
                            "ticket": ticket,
                            "fill_ts": ts,
                            "side": field(line, "side"),
                            "tf": field(line, "tf"),
                            "sid": sid,
                            "entry": _to_float(field(line, "price")),
                            "sl": _to_float(field(line, "sl")),
                            "tp": _to_float(field(line, "tp")),
                            "source": path,
                        }
                elif kind == "POSITION_CLOSED" and ticket not in seen_close:
                    line_symbol = field(line, "symbol")
                    if line_symbol and wanted_symbol and wanted_symbol not in line_symbol:
                        continue
                    seen_close.add(ticket)
                    closes[ticket] = {
                        "close_ts": ts,
                        "close_price": _to_float(field(line, "close_price")),
                        "profit": _to_float(field(line, "profit")),
                        "reason": field(line, "reason") or ("SL" if "SL Hit" in line else ("TP" if "TP Hit" in line else "Bot")),
                    }

    rows = []
    for ticket, fill in fills.items():
        if not (start_bkk <= fill["fill_ts"] <= end_bkk):
            continue
        close = closes.get(ticket, {})
        row = {**fill, **close}
        row["status"] = "CLOSED" if close else "OPEN"
        rows.append(row)
    rows.sort(key=lambda r: (r["fill_ts"], r["ticket"]))
    progress(f"Loaded {len(rows)} live filled order(s) from logs.")
    return rows


def enrich_rows_with_trail_logs(rows: list[dict], log_files: list[str]) -> None:
    if not rows:
        return
    by_ticket = {str(row.get("ticket", "") or ""): row for row in rows if row.get("ticket")}
    if not by_ticket:
        return

    trail_by_ticket: dict[str, list[dict]] = defaultdict(list)
    for path in log_files:
        if not os.path.exists(path):
            continue
        with open(path, encoding="utf-8", errors="replace") as f:
            for line in f:
                if "SL_CHANGED" not in line:
                    continue
                ticket = field(line, "ticket")
                if ticket not in by_ticket:
                    continue
                m = re.match(r"^\[(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2})\]", line)
                ts = parse_log_ts(m.group(1)) if m else None
                trail_by_ticket[ticket].append({
                    "ts": ts,
                    "old_sl": _to_float(field(line, "old_sl")),
                    "new_sl": _to_float(field(line, "new_sl")),
                    "source": field(line, "source"),
                    "reason": field(line, "reason"),
                })

    for ticket, events in trail_by_ticket.items():
        events.sort(key=lambda e: e.get("ts") or datetime.min)
        last = events[-1]
        row = by_ticket[ticket]
        row["live_trail_count"] = len(events)
        row["live_last_trail_ts"] = _excel_dt(last.get("ts"))
        row["live_last_trail_sl"] = last.get("new_sl", "")
        row["live_last_trail_source"] = last.get("source", "")
        row["live_trail_path"] = _trail_path(events, "ts", "source", "new_sl")
        close_price = float(row.get("close_price", 0.0) or 0.0)
        if close_price:
            row["live_close_vs_trail_sl_diff"] = round(close_price - float(last.get("new_sl", 0.0) or 0.0), 2)


def enrich_rows_with_sl_guard_logs(rows: list[dict], log_files: list[str]) -> None:
    if not rows:
        return
    by_ticket = {str(row.get("ticket", "") or ""): row for row in rows if row.get("ticket")}
    if not by_ticket:
        return

    activations: list[dict] = []
    close_requests: dict[str, dict] = {}
    close_events: list[dict] = []
    sl_hits: list[dict] = []
    for path in log_files:
        if not os.path.exists(path):
            continue
        with open(path, encoding="utf-8", errors="replace") as f:
            for line in f:
                if "SL_GUARD" not in line and "POSITION_CLOSE_REQUEST" not in line and "POSITION_CLOSED" not in line:
                    continue
                m = re.match(r"^\[(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2})\]\s+(\S+)", line)
                if not m:
                    continue
                ts = parse_log_ts(m.group(1))
                if ts is None:
                    continue
                kind = m.group(2)
                if kind == "SL_GUARD_GROUP_ACTIVATE":
                    activations.append({
                        "ts": ts,
                        "side": field(line, "side"),
                        "group": field(line, "group"),
                        "count": field(line, "count"),
                        "trigger_tf": field(line, "trigger_tf"),
                    })
                elif kind == "POSITION_CLOSE_REQUEST" and "SL Guard Group activate" in line:
                    ticket = field(line, "ticket")
                    if ticket in by_ticket:
                        close_requests[ticket] = {
                            "ts": ts,
                            "close_price": _to_float(field(line, "close_price")),
                            "bid": _to_float(field(line, "bid")),
                            "ask": _to_float(field(line, "ask")),
                            "spread": _to_float(field(line, "spread")),
                        }
                elif kind == "SL_GUARD_CLOSE":
                    ticket = field(line, "ticket")
                    if ticket in by_ticket:
                        close_events.append({
                            "ts": ts,
                            "ticket": ticket,
                            "side": field(line, "side"),
                        })
                elif kind == "POSITION_CLOSED" and "SL Hit" in line:
                    try:
                        sid = int(field(line, "sid") or 0)
                    except ValueError:
                        sid = 0
                    profit = _to_float(field(line, "profit"))
                    if profit < 0:
                        sl_hits.append({
                            "ts": ts,
                            "ticket": field(line, "ticket"),
                            "side": field(line, "side"),
                            "tf": field(line, "tf"),
                            "sid": sid,
                            "close_price": _to_float(field(line, "close_price")),
                            "profit": profit,
                            "reason": field(line, "reason"),
                        })

    activations.sort(key=lambda e: e["ts"])
    sl_hits.sort(key=lambda e: e["ts"])
    for activation in activations:
        group_tfs = set(str(activation.get("group", "") or "").replace("[", "").replace("]", "").split("+"))
        trigger_side = activation.get("side", "")
        candidates = [
            hit for hit in sl_hits
            if (not trigger_side or hit.get("side") == trigger_side)
            and (not group_tfs or hit.get("tf") in group_tfs)
            and timedelta(seconds=-10) <= hit["ts"] - activation["ts"] <= timedelta(seconds=10)
        ]
        activation["trigger_candidates"] = candidates
        activation["trigger_candidates_text"] = " | ".join(
            f"#{hit.get('ticket')} S{hit.get('sid')} {hit.get('tf')} close={float(hit.get('close_price', 0.0) or 0.0):.2f} pnl={float(hit.get('profit', 0.0) or 0.0):+.2f}"
            for hit in candidates[:6]
        )

    for event in close_events:
        row = by_ticket.get(event["ticket"])
        if not row:
            continue
        side = event.get("side") or row.get("side")
        row_tf = str(row.get("tf", "") or "")
        candidates = [
            act for act in activations
            if (not side or act.get("side") == side)
            and timedelta(seconds=0) <= event["ts"] - act["ts"] <= timedelta(seconds=10)
            and (not row_tf or row_tf in str(act.get("group", "")))
        ]
        activation = candidates[0] if candidates else {}
        request = close_requests.get(event["ticket"], {})
        row["live_sl_guard_close_ts"] = _excel_dt(event["ts"])
        row["live_sl_guard_activate_ts"] = _excel_dt(activation.get("ts"))
        row["live_sl_guard_group"] = activation.get("group", "")
        row["live_sl_guard_trigger_tf"] = activation.get("trigger_tf", "")
        row["live_sl_guard_count"] = activation.get("count", "")
        row["live_sl_guard_request_price"] = request.get("close_price", "")
        row["live_sl_guard_spread"] = request.get("spread", "")
        row["live_sl_guard_trigger_candidates"] = activation.get("trigger_candidates_text", "")


def load_live_sl_guard_activations(log_files: list[str]) -> list[dict]:
    activations: list[dict] = []
    for path in log_files:
        if not os.path.exists(path):
            continue
        with open(path, encoding="utf-8", errors="replace") as f:
            for line in f:
                if "SL_GUARD_GROUP_ACTIVATE" not in line and "POSITION_CLOSE_REQUEST" not in line:
                    continue
                m = re.match(r"^\[(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2})\]\s+(\S+)", line)
                if not m:
                    continue
                ts = parse_log_ts(m.group(1))
                if ts is None:
                    continue
                kind = m.group(2)
                if kind == "SL_GUARD_GROUP_ACTIVATE":
                    activations.append({
                        "ts": ts,
                        "side": field(line, "side"),
                        "group": field(line, "group"),
                        "count": field(line, "count"),
                        "trigger_tf": field(line, "trigger_tf"),
                        "request_ts": None,
                        "request_price": 0.0,
                    })
                elif kind == "POSITION_CLOSE_REQUEST" and "SL Guard Group activate" in line:
                    side = field(line, "side")
                    close_price = _to_float(field(line, "close_price"))
                    for act in reversed(activations):
                        if act.get("side") != side:
                            continue
                        delta = (ts - act["ts"]).total_seconds()
                        if 0 <= delta <= 10:
                            if not act.get("request_ts"):
                                act["request_ts"] = ts
                            if close_price:
                                act["request_price"] = close_price
                            break
    return [act for act in activations if act.get("request_price")]


def _parse_strategy_comment(comment: str) -> dict:
    comment = comment or ""
    m = re.search(r"\[([A-Z0-9]+(?:_[A-Z0-9]+)+)\]_S(\d+)(?:_#(\d+))?", comment)
    if m:
        tf_members = [part for part in m.group(1).split("_") if part]
        return {
            "htf": tf_members[0] if tf_members else "",
            "tf": tf_members[-1] if tf_members else "",
            "tf_members": tf_members,
            "sid": int(m.group(2)),
            "model": int(m.group(3) or 0),
        }

    m = re.search(r"(?:^|[^A-Z0-9])([A-Z0-9]+)_S(\d+)(?:[_#-](\d+))?", comment)
    if m:
        return {
            "htf": "",
            "tf": m.group(1),
            "tf_members": [m.group(1)],
            "sid": int(m.group(2)),
            "model": int(m.group(3) or 0),
        }

    m = re.search(r"(?:^|[^A-Z0-9])S(\d+)(?:[^0-9]|$)", comment)
    if m:
        return {"htf": "", "tf": "", "tf_members": [], "sid": int(m.group(1)), "model": 0}
    return {}


def _pattern_from_strategy_comment(comment: str) -> str:
    text = str(comment or "").upper()
    m = re.search(r"(?:^|[^A-Z0-9])(?:\[?[A-Z0-9_+-]+\]?)_S\d+_([^#\s]+)", text)
    if not m:
        return ""
    return m.group(1).strip("_")


def _s3_pattern_code(pattern: str, marubozu_source: str = "") -> str:
    text = str(pattern or "").upper()
    m = re.search(r"\[C1:([A-Z_]+)\]", text)
    if m:
        return m.group(1)
    if text in {"G", "R", "G_DOJI", "R_DOJI"}:
        return text
    source = str(marubozu_source or "").upper()
    if "NOENGULF" in source:
        return "NOENGULF"
    if "MARUBOZU" in source or "MARUBOZU" in text:
        return "MARUBOZU"
    return text


def _s14_family_from_comment(comment: str) -> str:
    text = str(comment or "").upper()
    if "_S14_" not in text:
        return ""
    code = text.split("_S14_", 1)[1]
    code = re.split(r"[^A-Z0-9]", code, maxsplit=1)[0]
    if not code:
        return ""
    if code.startswith("BSSH") or code.startswith("BSS"):
        return "BUY_ENGULF"
    if code.startswith("BRSH") or code.startswith("BRS"):
        return "BUY_SWEEP"
    if code.startswith("SSSH") or code.startswith("SSS"):
        return "SELL_ENGULF"
    if code.startswith("SRSH") or code.startswith("SRS"):
        return "SELL_SWEEP"
    legacy = {
        "BE": "BUY_ENGULF",
        "BS": "BUY_SWEEP",
        "SE": "SELL_ENGULF",
        "SS": "SELL_SWEEP",
    }
    for key, family in legacy.items():
        if code.startswith(key):
            return family
    return code


def _s14_family_from_pattern(side: str, pattern: str, sub_pattern: str = "") -> str:
    side = str(side or "").upper()
    sub = str(sub_pattern or "").lower()
    text = str(pattern or "").lower()
    if side not in ("BUY", "SELL"):
        return ""
    if "engulf" in sub or "swing" in text:
        return f"{side}_ENGULF"
    if "sweep" in sub or "sweep" in text:
        return f"{side}_SWEEP"
    return ""


def load_mt5_history_orders(start_bkk: datetime, end_bkk: datetime, symbol: str, strategies: set[int], close_search_days: int = 14) -> list[dict]:
    progress(f"Loading MT5 history for {symbol} (+{close_search_days} close-search day(s))...")
    bkk_tz = timezone(timedelta(hours=getattr(config, "TZ_OFFSET", 7)))
    from_dt = start_bkk.replace(tzinfo=bkk_tz)
    to_dt = (end_bkk + timedelta(days=close_search_days)).replace(tzinfo=bkk_tz)
    wanted_symbol = symbol or config.SYMBOL

    def _fetch_history_once() -> tuple[list, list, object]:
        fetched_deals = mt5.history_deals_get(from_dt, to_dt) or []
        fetched_orders = mt5.history_orders_get(from_dt, to_dt) or []
        try:
            err = mt5.last_error()
        except Exception:
            err = None
        return list(fetched_deals), list(fetched_orders), err

    deals, orders, first_err = _fetch_history_once()
    if not deals and not orders:
        progress(f"MT5 history returned empty (last_error={first_err}); reinitializing MT5 and retrying once...")
        try:
            mt5.shutdown()
        except Exception:
            pass
        initialized = mt5.initialize()
        try:
            retry_err = mt5.last_error()
        except Exception:
            retry_err = None
        if initialized:
            try:
                mt5.symbol_select(wanted_symbol, True)
            except Exception:
                pass
            deals, orders, retry_fetch_err = _fetch_history_once()
            progress(f"MT5 history retry loaded: deals={len(deals)} orders={len(orders)} last_error={retry_fetch_err}")
        else:
            progress(f"MT5 history retry skipped: initialize failed last_error={retry_err}")
    progress(f"MT5 history loaded: deals={len(deals)} orders={len(orders)}")

    order_by_ticket = {int(o.ticket): o for o in orders if getattr(o, "symbol", "") == wanted_symbol}
    positions = {}

    for d in sorted(deals, key=lambda deal: int(getattr(deal, "time", 0) or 0)):
        if getattr(d, "symbol", "") != wanted_symbol:
            continue
        comment = str(getattr(d, "comment", "") or "")
        meta = _parse_strategy_comment(comment)

        pos_id = int(getattr(d, "position_id", 0) or 0)
        if pos_id <= 0:
            continue
        if (not meta or meta.get("sid") not in strategies) and pos_id not in positions:
            continue
        if not meta or meta.get("sid") not in strategies:
            meta = {
                "tf": positions[pos_id].get("tf", ""),
                "htf": positions[pos_id].get("htf", ""),
                "tf_members": list(positions[pos_id].get("tf_members") or []),
                "sid": positions[pos_id].get("sid", 0),
                "model": positions[pos_id].get("model", 0),
            }
        row = positions.setdefault(pos_id, {
            "ticket": str(pos_id),
            "fill_ts": None,
            "close_ts": None,
            "side": "",
            "tf": meta.get("tf", ""),
            "htf": meta.get("htf", ""),
            "tf_members": list(meta.get("tf_members") or ([meta.get("tf")] if meta.get("tf") else [])),
            "sid": meta.get("sid", 0),
            "model": meta.get("model", 0),
            "entry": 0.0,
            "sl": 0.0,
            "tp": 0.0,
            "close_price": 0.0,
            "profit": 0.0,
            "reason": "",
            "status": "OPEN",
            "source": "mt5_history",
            "comment": comment,
            "entry_comment": "",
            "pattern": _pattern_from_strategy_comment(comment),
            "s14_family": "",
            "_out_volume": 0.0,
            "_out_value": 0.0,
            "_in_volume": 0.0,
            "_last_out_comment": "",
            "_scale_out_deals": [],
        })

        _deal_ts = int(getattr(d, "time", 0) or 0)
        if not _deal_ts:
            continue
        deal_time = datetime.fromtimestamp(_deal_ts, tz=bkk_tz).replace(tzinfo=None)
        deal_type = int(getattr(d, "type", -1))
        deal_entry = int(getattr(d, "entry", -1))
        volume = float(getattr(d, "volume", 0.0) or 0.0)
        price = float(getattr(d, "price", 0.0) or 0.0)

        if deal_entry == 0:
            row["fill_ts"] = deal_time if row["fill_ts"] is None else min(row["fill_ts"], deal_time)
            row["side"] = "BUY" if deal_type == 0 else "SELL" if deal_type == 1 else row["side"]
            row["entry"] = price if not row["entry"] else row["entry"]
            row["sl"] = float(getattr(d, "sl", 0.0) or row["sl"] or 0.0)
            row["tp"] = float(getattr(d, "tp", 0.0) or row["tp"] or 0.0)
            row["_in_volume"] += volume
            if not row["entry_comment"]:
                row["entry_comment"] = comment
                row["pattern"] = _pattern_from_strategy_comment(comment)
            if int(row.get("sid", 0) or 0) == 14 and not row.get("s14_family"):
                row["s14_family"] = _s14_family_from_comment(comment)
            ord_ = order_by_ticket.get(int(getattr(d, "order", 0) or 0)) or order_by_ticket.get(pos_id)
            if ord_:
                row["entry"] = float(getattr(ord_, "price_open", 0.0) or row["entry"])
                row["sl"] = float(getattr(ord_, "sl", 0.0) or row["sl"])
                row["tp"] = float(getattr(ord_, "tp", 0.0) or row["tp"])
        elif deal_entry in (1, 2, 3):
            deal_profit = float(getattr(d, "profit", 0.0) or 0.0)
            row["close_ts"] = deal_time if row["close_ts"] is None else max(row["close_ts"], deal_time)
            row["profit"] += deal_profit
            row["_out_volume"] += volume
            row["_out_value"] += volume * price
            row["_scale_out_deals"].append({
                "time": deal_time,
                "volume": volume,
                "profit": deal_profit,
                "price": price,
            })
            if row["_out_volume"] > 0:
                row["close_price"] = row["_out_value"] / row["_out_volume"]
            row["_last_out_comment"] = str(getattr(d, "comment", "") or "")

    rows = []
    for row in positions.values():
        if row["fill_ts"] is None:
            continue
        if not (start_bkk <= row["fill_ts"] <= end_bkk):
            continue
        if getattr(config, "SCALE_OUT_ENABLED", False):
            col_volume = scale_out_column_volume(symbol)
            for idx, value in enumerate(split_scale_out_pnl(row.get("_scale_out_deals", []), col_volume), 1):
                row[f"scale_out_{idx}_pnl"] = value
        in_volume = float(row.get("_in_volume", 0.0) or 0.0)
        out_volume = float(row.get("_out_volume", 0.0) or 0.0)
        if in_volume > 0 and out_volume >= in_volume - 1e-9:
            row["status"] = "CLOSED"
        elif out_volume > 0:
            row["status"] = "OPEN_PARTIAL"
        else:
            row["status"] = "OPEN"
        row["reason"] = row.get("_last_out_comment", "") or ("MT5_CLOSE" if row["status"] == "CLOSED" else row["status"])
        row.pop("_out_volume", None)
        row.pop("_out_value", None)
        row.pop("_in_volume", None)
        row.pop("_last_out_comment", None)
        row.pop("_scale_out_deals", None)
        rows.append(row)

    rows.sort(key=lambda r: (r["fill_ts"], r["ticket"]))
    progress(f"Loaded {len(rows)} live filled order(s) from MT5 history.")
    return rows


def _to_float(raw: str, default: float = 0.0) -> float:
    try:
        return float(raw)
    except (TypeError, ValueError):
        return default


def _sim_live_rows(sim_trades: list[tuple[str, dict]]) -> list[dict]:
    rows = []
    for htf, t in sim_trades:
        if t["close_type"] in ("CANCEL", "PD_FAIL", "OPEN_PENDING", "BLOCK"):
            continue
        sid = int(t.get("sid", 10) or 10)
        tf = t.get("tf") or (HTF_TO_LTF.get(htf, htf) if sid == 10 else htf)
        tf_members = list(t.get("tf_members") or t.get("parallel_tfs") or ([tf] if tf else []))
        fill_ts = t["entry_time"].replace(tzinfo=None)
        close_ts = t["close_time"].replace(tzinfo=None) if t.get("close_time") else None
        row = {
            "ticket": str(t.get("ticket", "")),
            "fill_ts": fill_ts,
            "close_ts": close_ts,
            "side": t.get("signal", ""),
            "tf": tf,
            "htf": t.get("htf_tf") or htf,
            "tf_members": tf_members,
            "sid": sid,
            "entry": float(t.get("entry", 0.0) or 0.0),
            "sl": float(t.get("sl", 0.0) or 0.0),
            "tp": float(t.get("tp", 0.0) or 0.0),
            "close_price": float(t.get("close_price", 0.0) or 0.0),
            "profit": float(t.get("pnl", 0.0) or 0.0),
            "reason": t.get("close_type", ""),
            "status": "CLOSED" if t.get("close_type") not in ("OPEN",) else "OPEN",
            "pattern": t.get("pattern", ""),
            "detect_ts": _naive_dt(t.get("detect_time")) or "",
            "source_candle_ts": to_bkk(t.get("source_candle_time")).replace(tzinfo=None) if t.get("source_candle_time") else "",
            "marubozu_source": t.get("marubozu_source", ""),
            "s14_family": _s14_family_from_pattern(t.get("signal", ""), t.get("pattern", ""), t.get("sub_pattern", "")) if sid == 14 else "",
            "bt_sl_guard_group": t.get("sl_guard_group", ""),
            "bt_sl_guard_trigger_tf": t.get("sl_guard_trigger_tf", ""),
        }
        for key in (
            "pd_h",
            "pd_l",
            "pd_fib_382",
            "pd_fib_618",
            "pd_fallback_used",
            "pd_outside_range",
            "pd_fill_h",
            "pd_fill_l",
            "pd_round2_h",
            "pd_round2_l",
            "pd_round2_changed",
            "pd_pending_h",
            "pd_pending_l",
            "pd_pending_round2_h",
            "pd_pending_round2_l",
            "pd_pending_round2_changed",
        ):
            row[key] = t.get(key, "")
        for key in (
            "s3_prev_sid_time_raw",
            "s3_prev_sid_gap_sec",
            "s3_prev_sid_adjacent",
            "s3_last_traded_time_raw",
            "s3_last_traded_matches_source",
            "s3_pending_same_sid_tf",
            "s3_open_same_sid_tf",
            "s3_active_same_sid_tf",
        ):
            row[key] = t.get(key, "")
        trail_events = list(t.get("trail_events") or [])
        if trail_events:
            last_trail = trail_events[-1]
            row["bt_trail_count"] = len(trail_events)
            row["bt_last_trail_ts"] = _excel_dt(last_trail.get("time"))
            row["bt_last_trail_sl"] = last_trail.get("new_sl", "")
            row["bt_last_trail_source"] = last_trail.get("tf", "")
            row["bt_trail_path"] = _trail_path(trail_events, "time", "tf", "new_sl")
        for idx in range(1, SCALE_OUT_COLUMNS + 1):
            row[f"scale_out_{idx}_pnl"] = t.get(f"scale_out_{idx}_pnl", "")
        rows.append(row)
    rows.sort(key=lambda r: (r["fill_ts"], r["side"], r["entry"]))
    return rows


def _naive_dt(value):
    if isinstance(value, datetime) and value.tzinfo is not None:
        return value.replace(tzinfo=None)
    return value


def _trade_pnl_at_price(trade: dict, close_price: float) -> float:
    realized = float(trade.get("realized_pnl", 0.0) or 0.0)
    remaining_units = int(trade.get("remaining_units", 1) or 0)
    if remaining_units <= 0:
        return round(realized, 2)
    if str(trade.get("signal", "")).upper() == "BUY":
        close_pnl = sim_s14_backtest.profit(float(close_price) - float(trade.get("entry", 0.0) or 0.0)) * remaining_units
    else:
        close_pnl = sim_s14_backtest.profit(float(trade.get("entry", 0.0) or 0.0) - float(close_price)) * remaining_units
    return round(realized + close_pnl, 2)


def apply_live_sl_guard_context_overlay(tf_trades: list[tuple[str, dict]], activations: list[dict]) -> list[tuple[str, dict]]:
    if not tf_trades or not activations:
        return tf_trades

    rows = []
    for idx, (tf_name, trade) in enumerate(tf_trades):
        rows.append({"idx": idx, "tf": tf_name, "trade": trade})

    for activation in sorted(activations, key=lambda a: a["ts"]):
        when = _naive_dt(activation.get("request_ts") or activation.get("ts"))
        close_price = float(activation.get("request_price", 0.0) or 0.0)
        side = str(activation.get("side", "") or "").upper()
        group_tfs = {
            tf.strip() for tf in str(activation.get("group", "") or "").replace("[", "").replace("]", "").split("+")
            if tf.strip()
        }
        if not when or not close_price or not side or not group_tfs:
            continue
        for row in rows:
            trade = row["trade"]
            trade_tf = str(trade.get("tf", row["tf"]) or row["tf"]).upper()
            if trade_tf not in group_tfs:
                continue
            if str(trade.get("signal", "") or "").upper() != side:
                continue
            entry_time = _naive_dt(trade.get("entry_time"))
            close_time = _naive_dt(trade.get("close_time"))
            if not entry_time or not close_time:
                continue
            if not (entry_time <= when < close_time):
                continue
            trade["close_type"] = "SL_GUARD_GROUP_LIVECTX"
            trade["close_price"] = close_price
            trade["close_time"] = when
            trade["pnl"] = _trade_pnl_at_price(trade, close_price)
            trade["sl_guard_group"] = activation.get("group", "")
            trade["sl_guard_trigger_tf"] = activation.get("trigger_tf", "")
            trade["sl_guard_context_source"] = "live_log"

    return [(row["tf"], row["trade"]) for row in rows]


def _minutes_abs(a: datetime, b: datetime) -> float:
    return abs((a - b).total_seconds()) / 60.0


def _nearest_candidate(row: dict, candidates: list[dict], prefix: str) -> dict:
    best = None
    best_score = None
    for candidate in candidates:
        if row.get("side") != candidate.get("side"):
            continue
        time_diff = _minutes_abs(row["fill_ts"], candidate["fill_ts"])
        entry_diff = abs(float(row.get("entry", 0.0) or 0.0) - float(candidate.get("entry", 0.0) or 0.0))
        score = time_diff * 10.0 + entry_diff
        if best_score is None or score < best_score:
            best = candidate
            best_score = score
    if not best:
        return {}
    try:
        sid = int(best.get("sid", 0) or 0)
    except (TypeError, ValueError):
        sid = 0
    return {
        f"nearest_{prefix}_fill_ts": best.get("fill_ts"),
        f"nearest_{prefix}_close_ts": best.get("close_ts", ""),
        f"nearest_{prefix}_close_price": best.get("close_price", ""),
        f"nearest_{prefix}_side": best.get("side", ""),
        f"nearest_{prefix}_tf": best.get("tf", ""),
        f"nearest_{prefix}_pattern": best.get("pattern", ""),
        f"nearest_{prefix}_s3_pattern_code": _s3_pattern_code(best.get("pattern", ""), best.get("marubozu_source", "")) if sid == 3 else "",
        f"nearest_{prefix}_marubozu_source": best.get("marubozu_source", ""),
        f"nearest_{prefix}_entry": best.get("entry"),
        f"nearest_{prefix}_pnl": best.get("profit", ""),
        f"nearest_{prefix}_reason": best.get("reason", ""),
        f"nearest_{prefix}_time_diff_min": round(_minutes_abs(row["fill_ts"], best["fill_ts"]), 2),
        f"nearest_{prefix}_entry_diff": round(abs(float(row.get("entry", 0.0) or 0.0) - float(best.get("entry", 0.0) or 0.0)), 2),
    }


def _nearest_compare_keys(prefix: str) -> tuple[str, ...]:
    return (
        f"nearest_{prefix}_fill_ts",
        f"nearest_{prefix}_close_ts",
        f"nearest_{prefix}_close_price",
        f"nearest_{prefix}_side",
        f"nearest_{prefix}_tf",
        f"nearest_{prefix}_pattern",
        f"nearest_{prefix}_s3_pattern_code",
        f"nearest_{prefix}_marubozu_source",
        f"nearest_{prefix}_entry",
        f"nearest_{prefix}_pnl",
        f"nearest_{prefix}_reason",
        f"nearest_{prefix}_time_diff_min",
        f"nearest_{prefix}_entry_diff",
    )


def _nearest_gap_reason(nearest: dict, prefix: str, time_tolerance_min: float, entry_tolerance: float) -> str:
    time_key = f"nearest_{prefix}_time_diff_min"
    entry_key = f"nearest_{prefix}_entry_diff"
    if time_key not in nearest or entry_key not in nearest:
        return "NO_SAME_SIDE_CANDIDATE"
    time_diff = float(nearest.get(time_key, 0.0) or 0.0)
    entry_diff = float(nearest.get(entry_key, 0.0) or 0.0)
    reasons = []
    if time_diff > time_tolerance_min:
        reasons.append("TIME_TOO_FAR")
    if entry_diff > entry_tolerance:
        reasons.append("ENTRY_TOO_FAR")
    if reasons:
        return "+".join(reasons)
    return "NEAREST_ALREADY_MATCHED_OR_GREEDY"


def _live_runtime_drift_label(live: dict) -> str:
    """Classify old live closes that current runtime would now skip."""
    try:
        sid = int(live.get("sid", 0) or 0)
    except (TypeError, ValueError):
        sid = 0
    reason = str(live.get("reason", "") or "").upper()
    if "PD ZONE" in reason and sid in set(getattr(config, "PDFIBOPLUS_SKIP_SIDS", ())):
        return "LIVE_HISTORICAL_PD_SKIP_DRIFT"
    if "FILL TREND" in reason and sid in FILL_TREND_RECHECK_SKIP_SIDS:
        return "LIVE_HISTORICAL_TREND_SKIP_DRIFT"
    return ""


def _live_only_gap_reason(live: dict, nearest: dict, time_tolerance_min: float, entry_tolerance: float) -> str:
    base = _nearest_gap_reason(nearest, "bt", time_tolerance_min, entry_tolerance)
    drift = _live_runtime_drift_label(live)
    if drift:
        return f"{drift}:{base}"
    reason = str(live.get("reason", "") or "").upper()
    entry_comment = str(live.get("entry_comment", live.get("comment", "")) or "").upper()
    if "PD ZONE" in reason:
        return f"LIVE_SIGNAL_PD_CLOSED_NO_REPLAY:{base}"
    if "_S14" in entry_comment:
        return f"LIVE_SIGNAL_NOT_REPLAYED:{base}"
    return base


def _backtest_only_gap_reason(sim: dict, nearest: dict, time_tolerance_min: float, entry_tolerance: float) -> str:
    base = _nearest_gap_reason(nearest, "live", time_tolerance_min, entry_tolerance)
    if int(sim.get("sid", 0) or 0) == 14:
        return f"BT_SIGNAL_NO_LIVE_FILL:{base}"
    return base


def _annotate_live_window_gap(sim: dict, live_rows: list[dict], time_tolerance_min: float) -> str:
    live_times = [row.get("fill_ts") for row in live_rows if isinstance(row.get("fill_ts"), datetime)]
    sim_ts = sim.get("fill_ts")
    if not isinstance(sim_ts, datetime):
        return ""
    if not live_times:
        return "BT_NO_LIVE_ROWS"
    first_live = min(live_times)
    last_live = max(live_times)
    sim["live_window_first_fill_ts"] = first_live
    sim["live_window_last_fill_ts"] = last_live
    tolerance = timedelta(minutes=float(time_tolerance_min or 0.0))
    if sim_ts < first_live - tolerance:
        return "BT_BEFORE_FIRST_FILTERED_LIVE_FILL"
    if sim_ts > last_live + tolerance:
        return "BT_AFTER_LAST_FILTERED_LIVE_FILL"
    return ""


def _compare_note(live: dict, sim: dict, match_quality: str = "") -> str:
    live_bucket = _close_bucket(live)
    sim_bucket = _close_bucket(sim)
    live_reason = str(live.get("reason", "") or "").upper()
    sim_reason = str(sim.get("reason", "") or "").upper()
    live_family = str(live.get("s14_family", "") or "")
    sim_family = str(sim.get("s14_family", "") or "")
    drift = _live_runtime_drift_label(live)
    if live_family and sim_family and live_family != sim_family:
        return "S14_FAMILY_DIFF"
    if live_bucket != sim_bucket:
        if drift:
            return drift
        if "SL GUARD" in live_reason or "SL_GUARD" in sim_reason:
            if str(match_quality or "").upper() == "LOOSE":
                return "LOOSE_MATCH_SL_GUARD"
            if live.get("live_sl_guard_group") and not sim.get("bt_sl_guard_group"):
                return "SL_GUARD_CONTEXT_MISSING_BT"
            return "CLOSE_LIFECYCLE_SL_GUARD"
        if "FILL TREND" in live_reason:
            return "LIVE_CLOSE_TREND_RECHECK"
        live_pd_close = "PD ZONE" in live_reason
        bt_pd_close = "PD" in sim_reason
        if live_pd_close and bt_pd_close:
            return "CLOSE_LIFECYCLE_PD"
        if live_pd_close:
            return "LIVE_CLOSE_PD_BT_NON_PD"
        if bt_pd_close:
            return "BT_CLOSE_PD_LIVE_NON_PD"
        return "CLOSE_BUCKET_DIFF"
    live_trail = live.get("live_last_trail_sl")
    bt_trail = sim.get("bt_last_trail_sl")
    if live_trail not in ("", None):
        if bt_trail in ("", None):
            return "TRAIL_SL_DIFF_MISSING_BT"
        try:
            if abs(float(live_trail or 0.0) - float(bt_trail or 0.0)) > 0.01:
                live_source = str(live.get("live_last_trail_source", "") or "")
                bt_source = str(sim.get("bt_last_trail_source", "") or "")
                if live_source and bt_source and live_source != bt_source:
                    return "TRAIL_SL_DIFF_SOURCE"
                return "TRAIL_SL_DIFF_PRICE"
        except (TypeError, ValueError):
            return "TRAIL_SL_DIFF_PRICE"
    if abs(float(live.get("profit", 0.0) or 0.0) - float(sim.get("profit", 0.0) or 0.0)) > 1.0:
        if drift:
            return drift
        if str(match_quality or "").upper() == "LOOSE":
            return "LOOSE_MATCH_PNL_DIFF"
        try:
            close_diff = abs(float(live.get("close_price", 0.0) or 0.0) - float(sim.get("close_price", 0.0) or 0.0))
            if close_diff > 1.0:
                return "CLOSE_PRICE_DIFF_SAME_BUCKET"
        except (TypeError, ValueError):
            pass
        return "PNL_DIFF_SAME_BUCKET"
    return ""


def compare_live_vs_backtest(
    live_rows: list[dict],
    sim_rows: list[dict],
    time_tolerance_min: float,
    entry_tolerance: float,
    pnl_tolerance: float,
    max_match_quality: str = "loose",
    prefer_same_s14_family: bool = False,
) -> dict:
    def _match_quality(time_diff: float, entry_diff: float) -> str:
        if time_diff <= 5 and entry_diff <= 1:
            return "EXACT"
        if time_diff <= 30 and entry_diff <= 3:
            return "NEAR"
        return "LOOSE"

    quality_rank = {"exact": 0, "near": 1, "loose": 2}
    max_quality_rank = quality_rank.get(str(max_match_quality or "loose").lower(), 2)
    candidates = []
    for live_idx, live in enumerate(live_rows):
        for sim_idx, sim in enumerate(sim_rows):
            if live["side"] != sim["side"]:
                continue
            entry_diff = abs(float(live["entry"]) - float(sim["entry"]))
            time_diff = _minutes_abs(live["fill_ts"], sim["fill_ts"])
            if entry_diff > entry_tolerance or time_diff > time_tolerance_min:
                continue
            quality = _match_quality(time_diff, entry_diff)
            if quality_rank[quality.lower()] > max_quality_rank:
                continue
            live_family = str(live.get("s14_family", "") or "")
            sim_family = str(sim.get("s14_family", "") or "")
            family_mismatch = bool(live_family and sim_family and live_family != sim_family)
            score = time_diff * 10.0 + entry_diff
            if prefer_same_s14_family and family_mismatch:
                score += 1_000_000.0
            candidates.append((score, time_diff, entry_diff, quality, family_mismatch, live_idx, sim_idx))

    matched_live = set()
    matched_sim = set()
    matches = []
    for _score, time_diff, entry_diff, quality, family_mismatch, live_idx, sim_idx in sorted(candidates):
        if live_idx in matched_live or sim_idx in matched_sim:
            continue
        matched_live.add(live_idx)
        matched_sim.add(sim_idx)
        live = live_rows[live_idx]
        sim = sim_rows[sim_idx]
        matches.append({
            "live": live,
            "sim": sim,
            "time_diff_min": time_diff,
            "entry_diff": entry_diff,
            "match_score": _score,
            "match_quality": quality,
            "s14_family_mismatch": family_mismatch,
            "pnl_diff": float(live.get("profit", 0.0) or 0.0) - float(sim.get("profit", 0.0) or 0.0),
        })

    live_only = []
    for live_idx, live in enumerate(live_rows):
        if live_idx in matched_live:
            continue
        enriched_live = dict(live)
        nearest = _nearest_candidate(live, sim_rows, "bt")
        enriched_live.update(nearest)
        enriched_live["gap_reason"] = _live_only_gap_reason(enriched_live, nearest, time_tolerance_min, entry_tolerance)
        live_only.append(enriched_live)

    backtest_only = []
    for sim_idx, sim_row in enumerate(sim_rows):
        if sim_idx in matched_sim:
            continue
        sim = dict(sim_row)
        nearest = _nearest_candidate(sim, live_rows, "live")
        sim.update(nearest)
        gap_reason = _backtest_only_gap_reason(sim, nearest, time_tolerance_min, entry_tolerance)
        live_window_gap = _annotate_live_window_gap(sim, live_rows, time_tolerance_min)
        if live_window_gap:
            gap_reason = f"{live_window_gap}:{gap_reason}"
        sim["gap_reason"] = gap_reason
        backtest_only.append(sim)
    mismatches = [
        m for m in matches
        if abs(m["pnl_diff"]) > pnl_tolerance
        or (m["live"].get("status") == "CLOSED" and m["sim"].get("status") == "CLOSED" and _close_bucket(m["live"]) != _close_bucket(m["sim"]))
    ]
    return {
        "matches": matches,
        "mismatches": mismatches,
        "live_only": live_only,
        "backtest_only": backtest_only,
    }


def enrich_compare_with_raw_replay_context(result: dict, all_trades: list[tuple[str, dict]]) -> None:
    raw_rows = []
    for report_tf, trade in all_trades:
        close_type = str(trade.get("close_type", "") or "")
        if close_type in ("SL", "TP", "SL_GUARD_CLOSE", "OPPOSITE_CLOSE", "LIMIT_SWEEP", "PD_FILL_FAIL", "TREND_FAIL", "RSI_FAIL"):
            continue
        entry_time = trade.get("entry_time")
        if not isinstance(entry_time, datetime):
            continue
        raw_rows.append((report_tf, trade, _excel_dt(entry_time)))

    for live in result.get("live_only", []):
        side = str(live.get("side", "") or "")
        fill_ts = live.get("fill_ts")
        if not isinstance(fill_ts, datetime):
            continue
        try:
            live_entry = float(live.get("entry", 0.0) or 0.0)
        except (TypeError, ValueError):
            live_entry = 0.0

        best = None
        for report_tf, trade, entry_time in raw_rows:
            if str(trade.get("signal", trade.get("side", "")) or "") != side:
                continue
            try:
                entry = float(trade.get("entry", 0.0) or 0.0)
            except (TypeError, ValueError):
                entry = 0.0
            time_diff = _minutes_abs(fill_ts, entry_time)
            entry_diff = abs(live_entry - entry)
            score = time_diff * 10.0 + entry_diff
            if best is None or score < best[0]:
                best = (score, report_tf, trade, time_diff, entry_diff)

        if best is None:
            continue
        _, report_tf, trade, time_diff, entry_diff = best
        live["nearest_raw_replay_tf"] = report_tf
        live["nearest_raw_replay_sid"] = trade.get("sid", "")
        live["nearest_raw_replay_side"] = trade.get("signal", trade.get("side", ""))
        live["nearest_raw_replay_entry_ts"] = _excel_dt(trade.get("entry_time"))
        live["nearest_raw_replay_entry"] = trade.get("entry", "")
        live["nearest_raw_replay_close_type"] = trade.get("close_type", "")
        live["nearest_raw_replay_cancel_reason"] = trade.get("cancel_reason", trade.get("reason", ""))
        live["nearest_raw_replay_pattern"] = trade.get("pattern", "")
        live["nearest_raw_replay_s3_pattern_code"] = _s3_pattern_code(trade.get("pattern", ""), trade.get("marubozu_source", "")) if int(trade.get("sid", 0) or 0) == 3 else ""
        live["nearest_raw_replay_marubozu_source"] = trade.get("marubozu_source", "")
        live["nearest_raw_replay_source_candle_ts"] = to_bkk(trade.get("source_candle_time")).replace(tzinfo=None) if trade.get("source_candle_time") else ""
        live["nearest_raw_replay_parallel_tfs"] = "|".join(str(tf) for tf in (trade.get("parallel_tfs") or []) if tf)
        live["nearest_raw_replay_gap_bot"] = trade.get("gap_bot", "")
        live["nearest_raw_replay_gap_top"] = trade.get("gap_top", "")
        live["nearest_raw_replay_final_gap_bot"] = trade.get("final_gap_bot", "")
        live["nearest_raw_replay_final_gap_top"] = trade.get("final_gap_top", "")
        live["nearest_raw_replay_cancel_age_bars"] = trade.get("cancel_age_bars", "")
        live["nearest_raw_replay_cancel_bars"] = trade.get("cancel_bars", "")
        live["nearest_raw_replay_cancel_bar_high"] = trade.get("cancel_bar_high", "")
        live["nearest_raw_replay_cancel_bar_low"] = trade.get("cancel_bar_low", "")
        live["nearest_raw_replay_cancel_bar_touched_entry"] = trade.get("cancel_bar_touched_entry", "")
        live["nearest_raw_replay_pd_h"] = trade.get("pd_h", "")
        live["nearest_raw_replay_pd_l"] = trade.get("pd_l", "")
        live["nearest_raw_replay_pd_fib_382"] = trade.get("pd_fib_382", "")
        live["nearest_raw_replay_pd_fib_618"] = trade.get("pd_fib_618", "")
        live["nearest_raw_replay_pd_fallback_used"] = trade.get("pd_fallback_used", "")
        live["nearest_raw_replay_pd_outside_range"] = trade.get("pd_outside_range", "")
        live["nearest_raw_replay_detect_time_raw"] = trade.get("detect_time_raw", "")
        live["nearest_raw_replay_sweep_scan_state"] = trade.get("sweep_scan_state", "")
        live["nearest_raw_replay_sweep_scan_tf"] = trade.get("sweep_scan_tf", "")
        live["nearest_raw_replay_sweep_scan_price"] = trade.get("sweep_scan_price", "")
        live["nearest_raw_replay_sweep_scan_time"] = trade.get("sweep_scan_time", "")
        live["nearest_raw_replay_sweep_scan_ts"] = trade.get("sweep_scan_ts", "")
        live["nearest_raw_replay_sweep_scan_age_min"] = trade.get("sweep_scan_age_min", "")
        live["nearest_raw_replay_sweep_scan_expiry_min"] = trade.get("sweep_scan_expiry_min", "")
        live["nearest_raw_replay_sl_guard_scope"] = trade.get("sl_guard_scope", "")
        live["nearest_raw_replay_sl_guard_key"] = trade.get("sl_guard_key", "")
        live["nearest_raw_replay_sl_guard_count"] = trade.get("sl_guard_count", "")
        live["nearest_raw_replay_sl_guard_since"] = trade.get("sl_guard_since", "")
        live["nearest_raw_replay_sl_guard_swing_ref"] = trade.get("sl_guard_swing_ref", "")
        live["nearest_raw_replay_time_diff_min"] = round(time_diff, 2)
        live["nearest_raw_replay_entry_diff"] = round(entry_diff, 2)
        raw_reason = str(trade.get("cancel_reason", trade.get("reason", "")) or "").upper()
        if "PD FIBO PLUS" in raw_reason:
            current_gap = str(live.get("gap_reason", "") or "")
            if not current_gap.startswith("REPLAY_PD_REJECTED:"):
                live["gap_reason"] = f"REPLAY_PD_REJECTED:{current_gap or 'NEAREST_RAW_PD_FAIL'}"


def _dedupe_live_rows(rows: list[dict]) -> list[dict]:
    by_ticket = {}
    no_ticket = []
    for row in rows:
        ticket = str(row.get("ticket", "") or "")
        if not ticket:
            no_ticket.append(row)
            continue
        existing = by_ticket.get(ticket)
        if existing is None or row.get("source") == "mt5_history":
            by_ticket[ticket] = row
    result = list(by_ticket.values()) + no_ticket
    result.sort(key=lambda r: (r.get("fill_ts") or datetime.min, str(r.get("ticket", ""))))
    return result


def _row_matches_requested_tf(row: dict, tf_arg: str | None) -> bool:
    if not tf_arg:
        return True
    target = tf_arg.upper()
    sid = int(row.get("sid", 0) or 0)
    row_tf = str(row.get("tf", "") or "").upper()
    row_htf = str(row.get("htf", "") or "").upper()
    tf_members = {str(tf).upper() for tf in (row.get("tf_members") or []) if tf}

    if sid == 10 and target in HTF_TO_LTF:
        return row_htf == target
    if target in tf_members:
        return True
    if row_tf:
        return row_tf == target
    if row_htf:
        return row_htf == target
    return True


def filter_live_rows_for_request(rows: list[dict], tf_arg: str | None) -> list[dict]:
    filtered = [row for row in rows if _row_matches_requested_tf(row, tf_arg)]
    if tf_arg and len(filtered) != len(rows):
        progress(f"TF filter {tf_arg.upper()}: live rows {len(rows)} -> {len(filtered)}")
    return filtered


def _close_bucket(row: dict) -> str:
    text = str(row.get("reason", "")).upper()
    profit = float(row.get("profit", 0.0) or 0.0)
    if "SL_GUARD" in text or "SL GUARD" in text:
        return "BOT"
    if "TP" in text:
        return "TP"
    if "SL" in text:
        return "SL"
    if profit > 0:
        return "TP?"
    if profit < 0:
        return "SL?"
    return "BOT"


def _reason_summary_rows(rows: list[dict], side: str, reason_key: str, pnl_key: str) -> list[dict]:
    buckets = {}
    for row in rows:
        reason = str(row.get(reason_key, "") or "UNKNOWN")
        side_value = str(row.get(f"{side}_side", row.get("live_side", row.get("bt_side", row.get("side", "")))) or "")
        key = (reason, side_value)
        item = buckets.setdefault(key, {
            "group": side,
            "reason": reason,
            "side": side_value,
            "count": 0,
            "pnl": 0.0,
            "first_ts": None,
            "last_ts": None,
        })
        item["count"] += 1
        item["pnl"] += float(row.get(pnl_key, 0.0) or 0.0)
        ts = row.get(f"{side}_fill_ts") or row.get("fill_ts")
        if ts:
            item["first_ts"] = ts if item["first_ts"] is None else min(item["first_ts"], ts)
            item["last_ts"] = ts if item["last_ts"] is None else max(item["last_ts"], ts)
    result = list(buckets.values())
    for item in result:
        item["pnl"] = round(item["pnl"], 2)
    result.sort(key=lambda r: (-int(r["count"]), str(r["reason"]), str(r["side"])))
    return result


def _gap_reason_summary_rows(rows: list[dict], side: str, pnl_key: str, group: str | None = None) -> list[dict]:
    buckets = {}
    for row in rows:
        reason = str(row.get("gap_reason", "") or "UNKNOWN")
        item = buckets.setdefault(reason, {
            "group": group or f"{side}_gap",
            "reason": reason,
            "side": str(row.get(f"{side}_side", row.get("live_side", row.get("bt_side", ""))) or ""),
            "count": 0,
            "pnl": 0.0,
            "first_ts": None,
            "last_ts": None,
        })
        item["count"] += 1
        item["pnl"] += float(row.get(pnl_key, 0.0) or 0.0)
        ts = row.get(f"{side}_fill_ts") or row.get("fill_ts")
        if ts:
            item["first_ts"] = ts if item["first_ts"] is None else min(item["first_ts"], ts)
            item["last_ts"] = ts if item["last_ts"] is None else max(item["last_ts"], ts)
    result = list(buckets.values())
    for item in result:
        item["pnl"] = round(item["pnl"], 2)
    result.sort(key=lambda r: (-int(r["count"]), str(r["reason"])))
    return result


def _trail_source_summary_rows(rows: list[dict]) -> list[dict]:
    buckets = {}
    for row in rows:
        live_source = str(row.get("live_last_trail_source", "") or "")
        bt_source = str(row.get("bt_last_trail_source", "") or "")
        if not live_source and not bt_source:
            continue
        reason = f"{live_source or '-'} -> {bt_source or '-'}"
        side_value = str(row.get("live_side", row.get("bt_side", "")) or "")
        key = (reason, side_value)
        item = buckets.setdefault(key, {
            "group": "trail_source_gap",
            "reason": reason,
            "side": side_value,
            "count": 0,
            "pnl": 0.0,
            "first_ts": None,
            "last_ts": None,
        })
        item["count"] += 1
        item["pnl"] += float(row.get("pnl_diff", 0.0) or 0.0)
        ts = row.get("live_fill_ts") or row.get("bt_fill_ts") or row.get("fill_ts")
        if ts:
            item["first_ts"] = ts if item["first_ts"] is None else min(item["first_ts"], ts)
            item["last_ts"] = ts if item["last_ts"] is None else max(item["last_ts"], ts)
    result = list(buckets.values())
    for item in result:
        item["pnl"] = round(item["pnl"], 2)
    result.sort(key=lambda r: (-int(r["count"]), str(r["reason"]), str(r["side"])))
    return result


def _field_summary_rows(
    rows: list[dict],
    group: str,
    field_key: str,
    side_key: str,
    pnl_key: str,
    ts_keys: tuple[str, ...],
) -> list[dict]:
    buckets = {}
    for row in rows:
        raw_value = row.get(field_key, "")
        value = "UNKNOWN" if raw_value in ("", None) else str(raw_value)
        side_value = str(row.get(side_key, row.get("live_side", row.get("bt_side", row.get("side", "")))) or "")
        key = (value, side_value)
        item = buckets.setdefault(key, {
            "group": group,
            "reason": value,
            "side": side_value,
            "count": 0,
            "pnl": 0.0,
            "first_ts": None,
            "last_ts": None,
        })
        item["count"] += 1
        item["pnl"] += float(row.get(pnl_key, 0.0) or 0.0)
        ts = None
        for ts_key in ts_keys:
            ts = row.get(ts_key)
            if ts:
                break
        if ts:
            item["first_ts"] = ts if item["first_ts"] is None else min(item["first_ts"], ts)
            item["last_ts"] = ts if item["last_ts"] is None else max(item["last_ts"], ts)
    result = list(buckets.values())
    for item in result:
        item["pnl"] = round(item["pnl"], 2)
    result.sort(key=lambda r: (-int(r["count"]), str(r["reason"]), str(r["side"])))
    return result


def compare_gap_summary_rows(result: dict) -> list[dict]:
    all_rows = compare_result_rows(result)
    live_only_rows = [r for r in all_rows if r.get("status") == "LIVE_ONLY"]
    bt_only_rows = [r for r in all_rows if r.get("status") == "BACKTEST_ONLY"]
    mismatch_rows = [r for r in all_rows if r.get("status") == "MISMATCH"]
    live_s14_rows = [r for r in live_only_rows if r.get("live_s14_family")]
    bt_s14_rows = [r for r in bt_only_rows if r.get("bt_s14_family")]
    mismatch_s14_rows = [r for r in mismatch_rows if r.get("live_s14_family") or r.get("bt_s14_family")]
    rows = []
    rows.extend(_reason_summary_rows(live_only_rows, "live", "live_reason", "live_pnl"))
    rows.extend(_reason_summary_rows(bt_only_rows, "bt", "bt_reason", "bt_pnl"))
    rows.extend(_reason_summary_rows(mismatch_rows, "live_mismatch", "live_reason", "pnl_diff"))
    rows.extend(_gap_reason_summary_rows(live_only_rows, "live", "live_pnl"))
    rows.extend(_gap_reason_summary_rows(bt_only_rows, "bt", "bt_pnl"))
    rows.extend(_gap_reason_summary_rows(mismatch_rows, "live", "pnl_diff", group="mismatch_gap"))
    rows.extend(_trail_source_summary_rows(mismatch_rows))
    rows.extend(_field_summary_rows(
        live_only_rows,
        "live_entry_comment",
        "live_entry_comment",
        "live_side",
        "live_pnl",
        ("live_fill_ts", "live_close_ts"),
    ))
    rows.extend(_field_summary_rows(
        mismatch_rows,
        "mismatch_entry_comment",
        "live_entry_comment",
        "live_side",
        "pnl_diff",
        ("live_fill_ts", "bt_fill_ts", "live_close_ts"),
    ))
    rows.extend(_field_summary_rows(
        live_s14_rows,
        "live_s14_family",
        "live_s14_family",
        "live_side",
        "live_pnl",
        ("live_fill_ts", "live_close_ts"),
    ))
    rows.extend(_field_summary_rows(
        bt_s14_rows,
        "bt_s14_family",
        "bt_s14_family",
        "bt_side",
        "bt_pnl",
        ("bt_fill_ts", "bt_close_ts"),
    ))
    rows.extend(_field_summary_rows(
        mismatch_s14_rows,
        "mismatch_s14_family",
        "live_s14_family",
        "live_side",
        "pnl_diff",
        ("live_fill_ts", "bt_fill_ts", "live_close_ts"),
    ))
    rows.extend(_field_summary_rows(
        mismatch_s14_rows,
        "s14_family_mismatch",
        "s14_family_mismatch",
        "live_side",
        "pnl_diff",
        ("live_fill_ts", "bt_fill_ts", "live_close_ts"),
    ))
    return rows


def print_compare_report(result: dict) -> None:
    matches = result["matches"]
    mismatches = result["mismatches"]
    live_only = result["live_only"]
    backtest_only = result["backtest_only"]
    live_total = sum(float(m["live"].get("profit", 0.0) or 0.0) for m in matches)
    sim_total = sum(float(m["sim"].get("profit", 0.0) or 0.0) for m in matches)
    print("\nLive vs Backtest Compare:")
    print(f"  Matched        : {len(matches)}")
    print(f"  Mismatches     : {len(mismatches)}")
    print(f"  Live only      : {len(live_only)}")
    print(f"  Backtest only  : {len(backtest_only)}")
    print(f"  Matched P&L    : live={live_total:+.2f} | backtest={sim_total:+.2f} | diff={live_total - sim_total:+.2f}")
    if matches:
        quality_counts = Counter(str(m.get("match_quality", "UNKNOWN")) for m in matches)
        quality_text = " | ".join(f"{k}={quality_counts.get(k, 0)}" for k in ("EXACT", "NEAR", "LOOSE"))
        print(f"  Match Quality  : {quality_text}")

    if mismatches:
        print("\n  Mismatch detail:")
        for m in mismatches[:20]:
            live = m["live"]
            sim = m["sim"]
            print(
                f"    LIVE #{live['ticket']} {live['fill_ts']} {live['side']} entry={live['entry']:.2f} "
                f"pnl={float(live.get('profit', 0.0) or 0.0):+.2f} reason={live.get('reason', '')} | "
                f"BT {sim['fill_ts']} entry={sim['entry']:.2f} pnl={sim['profit']:+.2f} reason={sim.get('reason', '')} "
                f"diff={m['pnl_diff']:+.2f} match={m.get('match_quality', '')}"
            )

    if live_only:
        print("\n  Live-only orders:")
        for row in live_only[:20]:
            print(f"    LIVE #{row['ticket']} {row['fill_ts']} {row['side']} entry={row['entry']:.2f} pnl={float(row.get('profit', 0.0) or 0.0):+.2f} reason={row.get('reason', '')}")

    if backtest_only:
        print("\n  Backtest-only orders:")
        for row in backtest_only[:20]:
            print(f"    BT {row['fill_ts']} {row['side']} entry={row['entry']:.2f} pnl={row['profit']:+.2f} reason={row.get('reason', '')}")

    gap_rows = compare_gap_summary_rows(result)
    if gap_rows:
        print("\n  Gap summary:")
        preferred_groups = (
            "mismatch_gap",
            "trail_source_gap",
            "live_s14_family",
            "bt_s14_family",
            "mismatch_s14_family",
            "s14_family_mismatch",
            "live_gap",
            "bt_gap",
            "live",
            "bt",
            "live_mismatch",
        )
        printed = 0
        for group in preferred_groups:
            group_rows = [row for row in gap_rows if row.get("group") == group]
            if not group_rows:
                continue
            print(f"    [{group}]")
            for row in group_rows[:8]:
                print(
                    f"      {row['reason']} {row['side']} "
                    f"count={row['count']} pnl={float(row.get('pnl', 0.0) or 0.0):+.2f}"
                )
                printed += 1
            if printed >= 24:
                break


def compare_result_rows(result: dict) -> list[dict]:
    def _scale_cols(prefix: str, row: dict) -> dict:
        return {
            f"{prefix}_scale_out_{idx}_pnl": row.get(f"scale_out_{idx}_pnl", "")
            for idx in range(1, SCALE_OUT_COLUMNS + 1)
        }

    bt_pd_diag_keys = (
        "pd_h",
        "pd_l",
        "pd_fib_382",
        "pd_fib_618",
        "pd_fallback_used",
        "pd_outside_range",
        "pd_fill_h",
        "pd_fill_l",
        "pd_round2_h",
        "pd_round2_l",
        "pd_round2_changed",
        "pd_pending_h",
        "pd_pending_l",
        "pd_pending_round2_h",
        "pd_pending_round2_l",
        "pd_pending_round2_changed",
    )
    bt_s3_diag_keys = (
        "s3_prev_sid_time_raw",
        "s3_prev_sid_gap_sec",
        "s3_prev_sid_adjacent",
        "s3_last_traded_time_raw",
        "s3_last_traded_matches_source",
        "s3_pending_same_sid_tf",
        "s3_open_same_sid_tf",
        "s3_active_same_sid_tf",
    )

    def _add_bt_pd_diag(prefix_row: dict, sim_row: dict) -> None:
        for key in bt_pd_diag_keys:
            prefix_row[f"bt_{key}"] = sim_row.get(key, "")

    def _add_bt_s3_diag(prefix_row: dict, sim_row: dict) -> None:
        for key in bt_s3_diag_keys:
            prefix_row[f"bt_{key}"] = sim_row.get(key, "")

    rows = []
    for m in result["matches"]:
        live = m["live"]
        sim = m["sim"]
        status = "MISMATCH" if m in result["mismatches"] else "MATCH"
        row = {
            "status": status,
            "gap_reason": _compare_note(live, sim, m.get("match_quality", "")) if status == "MISMATCH" else "",
            "live_ticket": live["ticket"],
            "live_fill_ts": live["fill_ts"],
            "live_close_ts": live.get("close_ts", ""),
            "live_close_price": live.get("close_price", ""),
            "live_side": live["side"],
            "live_tf": live.get("tf", ""),
            "live_pattern": live.get("pattern", ""),
            "live_s3_pattern_code": _s3_pattern_code(live.get("pattern", "")) if int(live.get("sid", 0) or 0) == 3 else "",
            "live_entry": live["entry"],
            "live_pnl": live.get("profit", 0.0),
            "live_reason": live.get("reason", ""),
            "live_entry_comment": live.get("entry_comment", live.get("comment", "")),
            "live_s14_family": live.get("s14_family", ""),
            "live_trail_count": live.get("live_trail_count", ""),
            "live_last_trail_ts": live.get("live_last_trail_ts", ""),
            "live_last_trail_sl": live.get("live_last_trail_sl", ""),
            "live_last_trail_source": live.get("live_last_trail_source", ""),
            "live_trail_path": live.get("live_trail_path", ""),
            "live_close_vs_trail_sl_diff": live.get("live_close_vs_trail_sl_diff", ""),
            "live_sl_guard_close_ts": live.get("live_sl_guard_close_ts", ""),
            "live_sl_guard_activate_ts": live.get("live_sl_guard_activate_ts", ""),
            "live_sl_guard_group": live.get("live_sl_guard_group", ""),
            "live_sl_guard_trigger_tf": live.get("live_sl_guard_trigger_tf", ""),
            "live_sl_guard_count": live.get("live_sl_guard_count", ""),
            "live_sl_guard_trigger_candidates": live.get("live_sl_guard_trigger_candidates", ""),
            "live_sl_guard_request_price": live.get("live_sl_guard_request_price", ""),
            "live_sl_guard_spread": live.get("live_sl_guard_spread", ""),
            "bt_fill_ts": sim["fill_ts"],
            "bt_close_ts": sim.get("close_ts", ""),
            "bt_close_price": sim.get("close_price", ""),
            "bt_side": sim.get("side", ""),
            "bt_tf": sim.get("tf", ""),
            "bt_pattern": sim.get("pattern", ""),
            "bt_s3_pattern_code": _s3_pattern_code(sim.get("pattern", ""), sim.get("marubozu_source", "")) if int(sim.get("sid", 0) or 0) == 3 else "",
            "bt_detect_ts": sim.get("detect_ts", ""),
            "bt_source_candle_ts": sim.get("source_candle_ts", ""),
            "bt_marubozu_source": sim.get("marubozu_source", ""),
            "bt_entry": sim["entry"],
            "bt_pnl": sim["profit"],
            "bt_reason": sim.get("reason", ""),
            "bt_s14_family": sim.get("s14_family", ""),
            "bt_trail_count": sim.get("bt_trail_count", ""),
            "bt_last_trail_ts": sim.get("bt_last_trail_ts", ""),
            "bt_last_trail_sl": sim.get("bt_last_trail_sl", ""),
            "bt_last_trail_source": sim.get("bt_last_trail_source", ""),
            "bt_trail_path": sim.get("bt_trail_path", ""),
            "bt_sl_guard_group": sim.get("bt_sl_guard_group", ""),
            "bt_sl_guard_trigger_tf": sim.get("bt_sl_guard_trigger_tf", ""),
            "match_quality": m.get("match_quality", ""),
            "s14_family_mismatch": m.get("s14_family_mismatch", ""),
            "match_score": round(float(m.get("match_score", 0.0) or 0.0), 2),
            "time_diff_min": round(m["time_diff_min"], 2),
            "entry_diff": round(m["entry_diff"], 2),
            "close_price_diff": round(float(live.get("close_price", 0.0) or 0.0) - float(sim.get("close_price", 0.0) or 0.0), 2),
            "pnl_diff": round(m["pnl_diff"], 2),
        }
        row.update(_scale_cols("live", live))
        row.update(_scale_cols("bt", sim))
        _add_bt_pd_diag(row, sim)
        _add_bt_s3_diag(row, sim)
        rows.append(row)
    for live in result["live_only"]:
        row = {"status": "LIVE_ONLY", "gap_reason": live.get("gap_reason", ""), "live_ticket": live["ticket"], "live_fill_ts": live["fill_ts"], "live_close_ts": live.get("close_ts", ""), "live_close_price": live.get("close_price", ""), "live_side": live["side"], "live_tf": live.get("tf", ""), "live_pattern": live.get("pattern", ""), "live_s3_pattern_code": _s3_pattern_code(live.get("pattern", "")) if int(live.get("sid", 0) or 0) == 3 else "", "live_entry": live["entry"], "live_pnl": live.get("profit", 0.0), "live_reason": live.get("reason", ""), "live_entry_comment": live.get("entry_comment", live.get("comment", "")), "live_s14_family": live.get("s14_family", ""), "live_trail_count": live.get("live_trail_count", ""), "live_last_trail_ts": live.get("live_last_trail_ts", ""), "live_last_trail_sl": live.get("live_last_trail_sl", ""), "live_last_trail_source": live.get("live_last_trail_source", ""), "live_trail_path": live.get("live_trail_path", ""), "live_close_vs_trail_sl_diff": live.get("live_close_vs_trail_sl_diff", ""), "live_sl_guard_close_ts": live.get("live_sl_guard_close_ts", ""), "live_sl_guard_activate_ts": live.get("live_sl_guard_activate_ts", ""), "live_sl_guard_group": live.get("live_sl_guard_group", ""), "live_sl_guard_trigger_tf": live.get("live_sl_guard_trigger_tf", ""), "live_sl_guard_count": live.get("live_sl_guard_count", ""), "live_sl_guard_trigger_candidates": live.get("live_sl_guard_trigger_candidates", ""), "live_sl_guard_request_price": live.get("live_sl_guard_request_price", ""), "live_sl_guard_spread": live.get("live_sl_guard_spread", "")}
        row.update(_scale_cols("live", live))
        for key in _nearest_compare_keys("bt"):
            row[key] = live.get(key, "")
        for key in (
            "nearest_raw_replay_tf",
            "nearest_raw_replay_sid",
            "nearest_raw_replay_side",
            "nearest_raw_replay_entry_ts",
            "nearest_raw_replay_entry",
            "nearest_raw_replay_close_type",
            "nearest_raw_replay_cancel_reason",
            "nearest_raw_replay_pattern",
            "nearest_raw_replay_s3_pattern_code",
            "nearest_raw_replay_marubozu_source",
            "nearest_raw_replay_source_candle_ts",
            "nearest_raw_replay_parallel_tfs",
            "nearest_raw_replay_gap_bot",
            "nearest_raw_replay_gap_top",
            "nearest_raw_replay_final_gap_bot",
            "nearest_raw_replay_final_gap_top",
            "nearest_raw_replay_cancel_age_bars",
            "nearest_raw_replay_cancel_bars",
            "nearest_raw_replay_cancel_bar_high",
            "nearest_raw_replay_cancel_bar_low",
            "nearest_raw_replay_cancel_bar_touched_entry",
            "nearest_raw_replay_pd_h",
            "nearest_raw_replay_pd_l",
            "nearest_raw_replay_pd_fib_382",
            "nearest_raw_replay_pd_fib_618",
            "nearest_raw_replay_pd_fallback_used",
            "nearest_raw_replay_pd_outside_range",
            "nearest_raw_replay_detect_time_raw",
            "nearest_raw_replay_sweep_scan_state",
            "nearest_raw_replay_sweep_scan_tf",
            "nearest_raw_replay_sweep_scan_price",
            "nearest_raw_replay_sweep_scan_time",
            "nearest_raw_replay_sweep_scan_ts",
            "nearest_raw_replay_sweep_scan_age_min",
            "nearest_raw_replay_sweep_scan_expiry_min",
            "nearest_raw_replay_sl_guard_scope",
            "nearest_raw_replay_sl_guard_key",
            "nearest_raw_replay_sl_guard_count",
            "nearest_raw_replay_sl_guard_since",
            "nearest_raw_replay_sl_guard_swing_ref",
            "nearest_raw_replay_time_diff_min",
            "nearest_raw_replay_entry_diff",
        ):
            row[key] = live.get(key, "")
        rows.append(row)
    for sim in result["backtest_only"]:
        row = {"status": "BACKTEST_ONLY", "gap_reason": sim.get("gap_reason", ""), "bt_fill_ts": sim["fill_ts"], "bt_close_ts": sim.get("close_ts", ""), "bt_close_price": sim.get("close_price", ""), "bt_side": sim.get("side", ""), "bt_tf": sim.get("tf", ""), "bt_pattern": sim.get("pattern", ""), "bt_s3_pattern_code": _s3_pattern_code(sim.get("pattern", ""), sim.get("marubozu_source", "")) if int(sim.get("sid", 0) or 0) == 3 else "", "bt_detect_ts": sim.get("detect_ts", ""), "bt_source_candle_ts": sim.get("source_candle_ts", ""), "bt_marubozu_source": sim.get("marubozu_source", ""), "bt_entry": sim["entry"], "bt_pnl": sim["profit"], "bt_reason": sim.get("reason", ""), "bt_s14_family": sim.get("s14_family", ""), "bt_trail_count": sim.get("bt_trail_count", ""), "bt_last_trail_ts": sim.get("bt_last_trail_ts", ""), "bt_last_trail_sl": sim.get("bt_last_trail_sl", ""), "bt_last_trail_source": sim.get("bt_last_trail_source", ""), "bt_trail_path": sim.get("bt_trail_path", ""), "bt_sl_guard_group": sim.get("bt_sl_guard_group", ""), "bt_sl_guard_trigger_tf": sim.get("bt_sl_guard_trigger_tf", "")}
        row.update(_scale_cols("bt", sim))
        _add_bt_pd_diag(row, sim)
        _add_bt_s3_diag(row, sim)
        for key in _nearest_compare_keys("live"):
            row[key] = sim.get(key, "")
        rows.append(row)
    return rows


def write_compare_csv(path: str, result: dict) -> str:
    rows = compare_result_rows(result)

    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    headers = sorted({k for row in rows for k in row.keys()})
    try:
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            writer.writerows(rows)
        return path
    except PermissionError:
        root, ext = os.path.splitext(path)
        fallback = f"{root}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{ext or '.csv'}"
        with open(fallback, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            writer.writerows(rows)
        return fallback


def write_compare_summary_csv(path: str, result: dict) -> str:
    root, ext = os.path.splitext(path)
    summary_path = f"{root}_summary{ext or '.csv'}"
    rows = compare_gap_summary_rows(result)
    headers = ["group", "reason", "side", "count", "pnl", "first_ts", "last_ts"]

    os.makedirs(os.path.dirname(summary_path) or ".", exist_ok=True)
    try:
        with open(summary_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            writer.writerows(rows)
        return summary_path
    except PermissionError:
        fallback = f"{root}_summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}{ext or '.csv'}"
        with open(fallback, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.DictWriter(f, fieldnames=headers)
            writer.writeheader()
            writer.writerows(rows)
        return fallback


def _save_workbook_with_fallback(wb, path: str) -> str:
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    try:
        wb.save(path)
        return path
    except PermissionError:
        root, ext = os.path.splitext(path)
        fallback = f"{root}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{ext or '.xlsx'}"
        wb.save(fallback)
        return fallback


def write_compare_xlsx(path: str, result: dict, meta: dict | None = None) -> str:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise RuntimeError("openpyxl is required for --compare-xlsx") from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF")
    ok_fill = PatternFill("solid", fgColor="E2F0D9")
    warn_fill = PatternFill("solid", fgColor="FFF2CC")
    bad_fill = PatternFill("solid", fgColor="F8CBAD")
    neutral_fill = PatternFill("solid", fgColor="D9EAF7")
    bold = Font(bold=True)

    matches = result["matches"]
    mismatches = result["mismatches"]
    live_only = result["live_only"]
    backtest_only = result["backtest_only"]
    live_total = sum(float(m["live"].get("profit", 0.0) or 0.0) for m in matches)
    sim_total = sum(float(m["sim"].get("profit", 0.0) or 0.0) for m in matches)

    summary_rows = [
        ("Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Symbol", (meta or {}).get("symbol", "")),
        ("Window", (meta or {}).get("window", "")),
        ("Strategy", (meta or {}).get("strategies", "")),
        ("TF Filter", (meta or {}).get("tf_filter", "")),
        ("Match Minutes", (meta or {}).get("match_minutes", "")),
        ("Match Entry Points", (meta or {}).get("match_entry_points", "")),
        ("Max Match Quality", (meta or {}).get("max_match_quality", "")),
        ("Prefer Same S14 Family", (meta or {}).get("prefer_same_s14_family", "")),
        ("Hybrid Live Guard Context", (meta or {}).get("hybrid_live_guard_context", "")),
        ("Scale Out Enabled", str(getattr(config, "SCALE_OUT_ENABLED", False))),
        ("Scale Out Column Lot", (meta or {}).get("scale_out_column_volume", "")),
        ("Matched", len(matches)),
        ("Mismatches", len(mismatches)),
        ("Live Only", len(live_only)),
        ("Backtest Only", len(backtest_only)),
        ("Matched Live P&L", live_total),
        ("Matched Backtest P&L", sim_total),
        ("Matched Diff", live_total - sim_total),
    ]
    ws.append(["Metric", "Value"])
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    for row in summary_rows:
        ws.append(list(row))
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 42
    ws.freeze_panes = "A2"

    def add_rows_sheet(title: str, rows: list[dict], preferred_headers: list[str] | None = None) -> None:
        sheet = wb.create_sheet(title)
        if rows:
            headers = preferred_headers or sorted({k for row in rows for k in row.keys()})
        else:
            headers = preferred_headers or ["status", "note"]
            rows = [{"status": "", "note": "No rows"}]
        sheet.append(headers)
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        for row in rows:
            sheet.append([row.get(h, "") for h in headers])
            status = str(row.get("status", ""))
            fill = None
            if status == "MATCH":
                fill = ok_fill
            elif status == "MISMATCH":
                fill = bad_fill
            elif status in ("LIVE_ONLY", "BACKTEST_ONLY"):
                fill = warn_fill
            if fill:
                for cell in sheet[sheet.max_row]:
                    cell.fill = fill
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for idx, header in enumerate(headers, 1):
            max_len = len(str(header))
            for row_idx in range(2, min(sheet.max_row, 80) + 1):
                max_len = max(max_len, len(str(sheet.cell(row_idx, idx).value or "")))
            sheet.column_dimensions[get_column_letter(idx)].width = min(max(max_len + 2, 10), 34)
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top")

    all_rows = compare_result_rows(result)
    preferred = [
        "status", "gap_reason", "live_ticket", "live_fill_ts", "live_close_ts", "live_close_price", "live_side", "live_tf", "live_pattern", "live_s3_pattern_code", "live_entry", "live_pnl",
        "live_scale_out_1_pnl", "live_scale_out_2_pnl", "live_scale_out_3_pnl", "live_scale_out_4_pnl",
        "live_reason", "live_entry_comment",
        "live_trail_count", "live_last_trail_ts", "live_last_trail_sl", "live_last_trail_source", "live_trail_path", "live_close_vs_trail_sl_diff",
        "live_sl_guard_close_ts", "live_sl_guard_activate_ts", "live_sl_guard_group", "live_sl_guard_trigger_tf", "live_sl_guard_count",
        "live_sl_guard_trigger_candidates", "live_sl_guard_request_price", "live_sl_guard_spread",
        "bt_fill_ts", "bt_close_ts", "bt_close_price", "bt_side", "bt_tf", "bt_pattern", "bt_s3_pattern_code", "bt_detect_ts", "bt_source_candle_ts", "bt_marubozu_source", "bt_entry", "bt_pnl",
        "bt_scale_out_1_pnl", "bt_scale_out_2_pnl", "bt_scale_out_3_pnl", "bt_scale_out_4_pnl",
        "bt_reason", "match_quality", "match_score", "time_diff_min", "entry_diff", "close_price_diff", "pnl_diff",
        "bt_pd_h", "bt_pd_l", "bt_pd_fib_382", "bt_pd_fib_618", "bt_pd_fallback_used", "bt_pd_outside_range",
        "bt_pd_fill_h", "bt_pd_fill_l", "bt_pd_round2_h", "bt_pd_round2_l", "bt_pd_round2_changed",
        "bt_pd_pending_h", "bt_pd_pending_l", "bt_pd_pending_round2_h", "bt_pd_pending_round2_l", "bt_pd_pending_round2_changed",
        "bt_s3_prev_sid_time_raw", "bt_s3_prev_sid_gap_sec", "bt_s3_prev_sid_adjacent",
        "bt_s3_last_traded_time_raw", "bt_s3_last_traded_matches_source",
        "bt_s3_pending_same_sid_tf", "bt_s3_open_same_sid_tf", "bt_s3_active_same_sid_tf",
        "bt_trail_count", "bt_last_trail_ts", "bt_last_trail_sl", "bt_last_trail_source", "bt_trail_path",
        "bt_sl_guard_group", "bt_sl_guard_trigger_tf",
        "live_window_first_fill_ts", "live_window_last_fill_ts",
        *_nearest_compare_keys("bt"),
        "nearest_raw_replay_tf", "nearest_raw_replay_sid", "nearest_raw_replay_side",
        "nearest_raw_replay_entry_ts", "nearest_raw_replay_entry",
        "nearest_raw_replay_close_type", "nearest_raw_replay_cancel_reason",
        "nearest_raw_replay_pattern", "nearest_raw_replay_s3_pattern_code",
        "nearest_raw_replay_marubozu_source", "nearest_raw_replay_source_candle_ts",
        "nearest_raw_replay_parallel_tfs",
        "nearest_raw_replay_gap_bot", "nearest_raw_replay_gap_top",
        "nearest_raw_replay_final_gap_bot", "nearest_raw_replay_final_gap_top",
        "nearest_raw_replay_cancel_age_bars", "nearest_raw_replay_cancel_bars",
        "nearest_raw_replay_cancel_bar_high", "nearest_raw_replay_cancel_bar_low",
        "nearest_raw_replay_cancel_bar_touched_entry",
        "nearest_raw_replay_pd_h", "nearest_raw_replay_pd_l",
        "nearest_raw_replay_pd_fib_382", "nearest_raw_replay_pd_fib_618",
        "nearest_raw_replay_pd_fallback_used", "nearest_raw_replay_pd_outside_range",
        "nearest_raw_replay_detect_time_raw",
        "nearest_raw_replay_sweep_scan_state", "nearest_raw_replay_sweep_scan_tf",
        "nearest_raw_replay_sweep_scan_price", "nearest_raw_replay_sweep_scan_time",
        "nearest_raw_replay_sweep_scan_ts", "nearest_raw_replay_sweep_scan_age_min",
        "nearest_raw_replay_sweep_scan_expiry_min",
        "nearest_raw_replay_sl_guard_scope", "nearest_raw_replay_sl_guard_key",
        "nearest_raw_replay_sl_guard_count", "nearest_raw_replay_sl_guard_since",
        "nearest_raw_replay_sl_guard_swing_ref",
        "nearest_raw_replay_time_diff_min", "nearest_raw_replay_entry_diff",
        *_nearest_compare_keys("live"),
    ]
    add_rows_sheet("All Compare", all_rows, preferred)
    add_rows_sheet("Mismatches", [r for r in all_rows if r.get("status") == "MISMATCH"], preferred)
    add_rows_sheet("Live Only", [r for r in all_rows if r.get("status") == "LIVE_ONLY"], preferred)
    add_rows_sheet("Backtest Only", [r for r in all_rows if r.get("status") == "BACKTEST_ONLY"], preferred)
    add_rows_sheet("Matches", [r for r in all_rows if r.get("status") == "MATCH"], preferred)
    add_rows_sheet(
        "Gap Summary",
        compare_gap_summary_rows(result),
        ["group", "reason", "side", "count", "pnl", "first_ts", "last_ts"],
    )

    legend_start = ws.max_row + 2
    ws.cell(legend_start, 1).value = "Status Legend"
    ws.cell(legend_start, 1).font = bold
    legend = [("MATCH", "Matched within tolerance"), ("MISMATCH", "Matched but outcome/P&L differs"), ("LIVE_ONLY", "Real order exists but replay did not create it"), ("BACKTEST_ONLY", "Replay created order but no matching real fill")]
    for label, desc in legend:
        ws.append([label, desc])
        fill = ok_fill if label == "MATCH" else bad_fill if label == "MISMATCH" else warn_fill
        ws.cell(ws.max_row, 1).fill = fill
        ws.cell(ws.max_row, 2).fill = neutral_fill

    return _save_workbook_with_fallback(wb, path)


def print_feature_snapshot() -> None:
    print("Runtime Feature Snapshot:")
    print(f"  Symbol                 : {config.SYMBOL}")
    print(f"  Volume                 : {getattr(config, 'AUTO_VOLUME', 0.01)}")
    print(f"  SL Guard per-TF        : {getattr(config, 'SL_GUARD_ENABLED', False)}")
    print(f"  SL Guard combined      : {getattr(config, 'SL_GUARD_COMBINED_ENABLED', False)}")
    print(f"  SL Guard group         : {getattr(config, 'SL_GUARD_GROUP_ENABLED', False)}")


def _coverage_status(item: dict) -> str:
    replay = str(item.get("replay", ""))
    if replay == "apply":
        return "replayed"
    if replay == "partial":
        return "partial"
    if replay == "ready":
        return "ready" if item.get("config_on") else "off ready"
    if replay.startswith("skip"):
        return "skipped"
    return "ACTIVE GAP" if item.get("config_on") else "off gap"


def print_s1_coverage() -> None:
    print("\nS1 Runtime Coverage:")
    for item in s1_runtime_feature_coverage():
        status = _coverage_status(item)
        print(f"  {item['name']:<34} config={str(item['config_on']):<5} {status:<12} {item['note']}")

    gaps = s1_unreplayed_active_features()
    if gaps:
        print("\nWARNING: Active S1 runtime features not replayed yet:")
        for item in gaps:
            print(f"  - {item['name']}: {item['note']}")


def print_s2_coverage() -> None:
    print("\nS2 Runtime Coverage:")
    for item in s2_runtime_feature_coverage():
        status = _coverage_status(item)
        print(f"  {item['name']:<34} config={str(item['config_on']):<5} {status:<12} {item['note']}")

    gaps = s2_unreplayed_active_features()
    if gaps:
        print("\nWARNING: Active S2 runtime features not replayed yet:")
        for item in gaps:
            print(f"  - {item['name']}: {item['note']}")


def print_s3_coverage() -> None:
    print("\nS3 Runtime Coverage:")
    for item in s3_runtime_feature_coverage():
        status = _coverage_status(item)
        print(f"  {item['name']:<34} config={str(item['config_on']):<5} {status:<12} {item['note']}")

    gaps = s3_unreplayed_active_features()
    if gaps:
        print("\nWARNING: Active S3 runtime features not replayed yet:")
        for item in gaps:
            print(f"  - {item['name']}: {item['note']}")


def print_s4_coverage() -> None:
    print("\nS4 Runtime Coverage:")
    for item in s4_runtime_feature_coverage():
        status = _coverage_status(item)
        print(f"  {item['name']:<34} config={str(item['config_on']):<5} {status:<12} {item['note']}")

    gaps = s4_unreplayed_active_features()
    if gaps:
        print("\nWARNING: Active S4 runtime features not replayed yet:")
        for item in gaps:
            print(f"  - {item['name']}: {item['note']}")


def print_s5_coverage() -> None:
    print("\nS5 Runtime Coverage:")
    for item in s5_runtime_feature_coverage():
        status = _coverage_status(item)
        print(f"  {item['name']:<34} config={str(item['config_on']):<5} {status:<12} {item['note']}")

    gaps = s5_unreplayed_active_features()
    if gaps:
        print("\nWARNING: Active S5 runtime features not replayed yet:")
        for item in gaps:
            print(f"  - {item['name']}: {item['note']}")


def print_s8_coverage() -> None:
    print("\nS8 Runtime Coverage:")
    for item in s8_runtime_feature_coverage():
        status = _coverage_status(item)
        print(f"  {item['name']:<34} config={str(item['config_on']):<5} {status:<12} {item['note']}")

    gaps = s8_unreplayed_active_features()
    if gaps:
        print("\nWARNING: Active S8 runtime features not replayed yet:")
        for item in gaps:
            print(f"  - {item['name']}: {item['note']}")


def print_s9_coverage() -> None:
    print("\nS9 Runtime Coverage:")
    for item in s9_runtime_feature_coverage():
        runtime = item["runtime"]
        replay = item["replay"]
        if runtime == "skip_s9":
            status = "runtime skip"
        elif replay == "apply":
            status = "replayed"
        elif replay == "partial":
            status = "partial"
        elif replay == "ready":
            status = "ready"
        elif item["config_on"]:
            status = "ACTIVE GAP"
        else:
            status = "off gap"
        print(f"  {item['name']:<34} config={str(item['config_on']):<5} {status:<12} {item['note']}")

    gaps = s9_unreplayed_active_features()
    if gaps:
        print("\nWARNING: Active S9 runtime features not replayed yet:")
        for item in gaps:
            print(f"  - {item['name']}: {item['note']}")


def print_s10_coverage() -> None:
    print("\nS10 Runtime Coverage:")
    for item in s10_runtime_feature_coverage():
        if item["runtime"] == "skip_s10":
            status = "runtime skip"
        elif item["replay"] == "apply":
            status = "replayed"
        elif item["replay"] == "ready":
            status = "ready"
        elif item["config_on"]:
            status = "ACTIVE GAP"
        else:
            status = "off gap"
        print(f"  {item['name']:<34} config={str(item['config_on']):<5} {status:<12} {item['note']}")

    gaps = s10_unreplayed_active_features()
    if gaps:
        print("\nWARNING: Active S10 runtime features not replayed yet:")
        for item in gaps:
            print(f"  - {item['name']}: {item['note']}")


def print_s11_coverage() -> None:
    print("\nS11 Runtime Coverage:")
    for item in s11_runtime_feature_coverage():
        runtime = item["runtime"]
        replay = item["replay"]
        if runtime == "skip_s11":
            status = "runtime skip"
        elif replay == "apply":
            status = "replayed"
        elif replay == "partial":
            status = "partial"
        elif replay == "ready":
            status = "ready"
        elif item["config_on"]:
            status = "ACTIVE GAP"
        else:
            status = "off gap"
        print(f"  {item['name']:<34} config={str(item['config_on']):<5} {status:<12} {item['note']}")

    gaps = s11_unreplayed_active_features()
    if gaps:
        print("\nWARNING: Active S11 runtime features not replayed yet:")
        for item in gaps:
            print(f"  - {item['name']}: {item['note']}")


def print_s12_coverage() -> None:
    print("\nS12 Runtime Coverage:")
    for item in s12_runtime_feature_coverage():
        runtime = item["runtime"]
        replay = item["replay"]
        if runtime == "skip_s12_or_market":
            status = "runtime skip"
        elif replay == "apply":
            status = "replayed"
        elif replay == "partial":
            status = "partial"
        elif replay == "ready":
            status = "ready"
        elif item["config_on"]:
            status = "ACTIVE GAP"
        else:
            status = "off gap"
        print(f"  {item['name']:<34} config={str(item['config_on']):<5} {status:<12} {item['note']}")

    gaps = s12_unreplayed_active_features()
    if gaps:
        print("\nWARNING: Active S12 runtime features not replayed yet:")
        for item in gaps:
            print(f"  - {item['name']}: {item['note']}")


def print_s14_coverage() -> None:
    print("\nS14 Runtime Coverage:")
    for item in s14_runtime_feature_coverage():
        runtime = item["runtime"]
        replay = item["replay"]
        if runtime == "skip_s14":
            status = "runtime skip"
        elif replay == "apply":
            status = "replayed"
        elif replay == "partial":
            status = "partial"
        elif replay == "ready":
            status = "ready"
        elif item["config_on"]:
            status = "ACTIVE GAP"
        else:
            status = "off gap"
        print(f"  {item['name']:<34} config={str(item['config_on']):<5} {status:<12} {item['note']}")

    gaps = s14_unreplayed_active_features()
    if gaps:
        print("\nWARNING: Active S14 runtime features not replayed yet:")
        for item in gaps:
            print(f"  - {item['name']}: {item['note']}")


def print_s13_coverage() -> None:
    print("\nS13 Runtime Coverage:")
    for item in s13_runtime_feature_coverage():
        runtime = item["runtime"]
        replay = item["replay"]
        if runtime == "skip_s13":
            status = "runtime skip"
        elif replay == "apply":
            status = "replayed"
        elif replay == "partial":
            status = "partial"
        elif replay == "ready":
            status = "ready"
        elif item["config_on"]:
            status = "ACTIVE GAP"
        else:
            status = "off gap"
        print(f"  {item['name']:<34} config={str(item['config_on']):<5} {status:<12} {item['note']}")

    gaps = s13_unreplayed_active_features()
    if gaps:
        print("\nWARNING: Active S13 runtime features not replayed yet:")
        for item in gaps:
            print(f"  - {item['name']}: {item['note']}")


def print_s15_coverage() -> None:
    print("\nS15 Runtime Coverage:")
    for item in s15_runtime_feature_coverage():
        runtime = item["runtime"]
        replay = item["replay"]
        if runtime == "skip_s15":
            status = "runtime skip"
        elif replay == "apply":
            status = "replayed"
        elif replay == "partial":
            status = "partial"
        elif replay == "ready":
            status = "ready"
        elif item["config_on"]:
            status = "ACTIVE GAP"
        else:
            status = "off gap"
        print(f"  {item['name']:<34} config={str(item['config_on']):<5} {status:<12} {item['note']}")

    gaps = s15_unreplayed_active_features()
    if gaps:
        print("\nWARNING: Active S15 runtime features not replayed yet:")
        for item in gaps:
            print(f"  - {item['name']}: {item['note']}")


def print_s16_coverage() -> None:
    print("\nS16 Runtime Coverage:")
    for item in s16_runtime_feature_coverage():
        runtime = item["runtime"]
        replay = item["replay"]
        if runtime == "skip_s16":
            status = "runtime skip"
        elif replay == "apply":
            status = "replayed"
        elif replay == "partial":
            status = "partial"
        elif replay == "ready":
            status = "ready"
        elif item["config_on"]:
            status = "ACTIVE GAP"
        else:
            status = "off gap"
        print(f"  {item['name']:<34} config={str(item['config_on']):<5} {status:<12} {item['note']}")

    gaps = s16_unreplayed_active_features()
    if gaps:
        print("\nWARNING: Active S16 runtime features not replayed yet:")
        for item in gaps:
            print(f"  - {item['name']}: {item['note']}")


def print_s17_coverage() -> None:
    print("\nS17 Runtime Coverage:")
    for item in s17_runtime_feature_coverage():
        runtime = item["runtime"]
        replay = item["replay"]
        if runtime == "skip_s17":
            status = "runtime skip"
        elif replay == "apply":
            status = "replayed"
        elif replay == "partial":
            status = "partial"
        elif replay == "ready":
            status = "ready"
        elif item["config_on"]:
            status = "ACTIVE GAP"
        else:
            status = "off gap"
        print(f"  {item['name']:<34} config={str(item['config_on']):<5} {status:<12} {item['note']}")

    gaps = s17_unreplayed_active_features()
    if gaps:
        print("\nWARNING: Active S17 runtime features not replayed yet:")
        for item in gaps:
            print(f"  - {item['name']}: {item['note']}")


def print_s18_coverage() -> None:
    print("\nS18 Runtime Coverage:")
    for item in s18_runtime_feature_coverage():
        runtime = item["runtime"]
        replay = item["replay"]
        if runtime == "skip_s18":
            status = "runtime skip"
        elif replay == "apply":
            status = "replayed"
        elif replay == "partial":
            status = "partial"
        elif replay == "ready":
            status = "ready"
        elif item["config_on"]:
            status = "ACTIVE GAP"
        else:
            status = "off gap"
        print(f"  {item['name']:<34} config={str(item['config_on']):<5} {status:<12} {item['note']}")

    gaps = s18_unreplayed_active_features()
    if gaps:
        print("\nWARNING: Active S18 runtime features not replayed yet:")
        for item in gaps:
            print(f"  - {item['name']}: {item['note']}")


def print_s19_coverage() -> None:
    print("\nS19 Runtime Coverage:")
    for item in s19_runtime_feature_coverage():
        runtime = item["runtime"]
        replay = item["replay"]
        if runtime == "skip_s19":
            status = "runtime skip"
        elif replay == "apply":
            status = "replayed"
        elif replay == "partial":
            status = "partial"
        elif replay == "ready":
            status = "ready"
        elif item["config_on"]:
            status = "ACTIVE GAP"
        else:
            status = "off gap"
        print(f"  {item['name']:<34} config={str(item['config_on']):<5} {status:<12} {item['note']}")

    gaps = s19_unreplayed_active_features()
    if gaps:
        print("\nWARNING: Active S19 runtime features not replayed yet:")
        for item in gaps:
            print(f"  - {item['name']}: {item['note']}")


def format_trade(tf_display: str, idx: int, t: dict) -> None:
    et = t["entry_time"].strftime("%Y-%m-%d %H:%M")
    ct = t["close_time"].strftime("%Y-%m-%d %H:%M") if t["close_type"] not in ("OPEN", "OPEN_PENDING") else "OPEN"
    pnl_s = f"{'+' if t['pnl'] >= 0 else ''}{t['pnl']:.2f}"
    print(f"\n--- Trade #{idx} ---")
    print(f"  [{tf_display}] {et} {t['signal']} [{t.get('pattern', 'S10')}]")
    print(f"  Entry  = {t['entry']:.2f} | SL = {t['sl']:.2f} | TP = {t['tp']:.2f}")
    cp = t.get('close_price')
    cp_str = f"{cp:.2f}" if cp is not None else "?"
    print(f"  Result -> {t['close_type']} @ {cp_str} [{ct}]  PnL={pnl_s} USD")
    if t.get("cancel_reason"):
        print(f"  Reason : {t['cancel_reason']}")

    htf_tf = t.get("s10_htf_tf")
    if htf_tf:
        parent_h = t.get("s10_parent_high", 0.0)
        parent_l = t.get("s10_parent_low", 0.0)
        parent_t = to_bkk(t.get("s10_parent_time", 0)).strftime("%Y-%m-%d %H:%M") if t.get("s10_parent_time") else "?"
        sweep_t = to_bkk(t.get("s10_sweep_time", 0)).strftime("%d-%m %H:%M") if t.get("s10_sweep_time") else "?"
        print(f"  HTF    : {htf_tf} | Parent H={parent_h:.2f} L={parent_l:.2f} Time={parent_t} | Sweep={sweep_t}")

    if int(t.get("sid", 10) or 10) == 14:
        def _fmt_time(value) -> str:
            if not value:
                return "?"
            if isinstance(value, datetime):
                return value.strftime("%Y-%m-%d %H:%M")
            try:
                return to_bkk(int(value)).strftime("%Y-%m-%d %H:%M")
            except (TypeError, ValueError, OSError):
                return str(value)

        ref_s = _fmt_time(t.get("ref_time"))
        sweep_s = _fmt_time(t.get("sweep_bar_time"))
        engulf_s = _fmt_time(t.get("engulf_bar_time"))
        sub_pattern = t.get("sub_pattern", "-")
        print(f"  S14    : {sub_pattern} | Ref={t.get('ref_source', '?')} @ {ref_s} | Sweep={sweep_s} | Engulf={engulf_s}")
        if t.get("rsi_at_ref") is not None or t.get("rsi_at_rej") is not None:
            print(f"  RSI    : ref={t.get('rsi_at_ref', '?')} rejection={t.get('rsi_at_rej', '?')}")


def run_s10(args, window_start_utc: datetime, window_end_utc: datetime) -> list[tuple[str, dict]]:
    all_trades = []
    for tf_name in resolve_run_tfs_for_strategy(10, args.tf):
        progress(f"Running S10 replay on {tf_name}...")
        trades = backtest_s10_tf(tf_name, S10_TF_MAP[tf_name])
        progress(f"S10 replay on {tf_name} produced {len(trades)} raw event(s).")
        filtered = [t for t in trades if window_start_utc <= t["entry_time"] <= window_end_utc]
        if args.tf and args.tf.upper() in HTF_TO_LTF:
            filtered = [t for t in filtered if t.get("s10_htf_tf") == args.tf.upper()]
        if args.exclude_cancelled:
            filtered = [
                t for t in filtered
                if t["close_type"] not in ("CANCEL", "PD_FAIL", "OPEN_PENDING", "BLOCK")
            ]
        for t in filtered:
            t.setdefault("sid", 10)
            t.setdefault("tf", HTF_TO_LTF.get(t.get("s10_htf_tf", tf_name), tf_name))
        all_trades.extend((t.get("s10_htf_tf", tf_name), t) for t in filtered)
        progress(f"S10 replay on {tf_name} kept {len(filtered)} event(s) in window.")
    return all_trades


def run_s1(args, window_start_utc: datetime, window_end_utc: datetime) -> list[tuple[str, dict]]:
    all_trades = []
    range_end_utc = _replay_range_end_utc(window_end_utc, getattr(args, "mt5_close_search_days", 14))
    for tf_name in resolve_run_tfs_for_strategy(1, args.tf):
        progress(f"Running S1 replay on {tf_name}...")
        trades = backtest_s1_tf(tf_name, S1_TF_MAP[tf_name], range_end_utc=range_end_utc)
        progress(f"S1 replay on {tf_name} produced {len(trades)} raw event(s).")
        filtered = [t for t in trades if window_start_utc <= t["entry_time"] <= window_end_utc]
        if args.exclude_cancelled:
            filtered = [
                t for t in filtered
                if t["close_type"] not in ("CANCEL", "PD_FAIL", "OPEN_PENDING", "BLOCK", "OPEN")
            ]
        for t in filtered:
            t.setdefault("sid", 1)
            t.setdefault("tf", tf_name)
        all_trades.extend((tf_name, t) for t in filtered)
        progress(f"S1 replay on {tf_name} kept {len(filtered)} event(s) in window.")
    return all_trades


def run_s2(args, window_start_utc: datetime, window_end_utc: datetime) -> list[tuple[str, dict]]:
    all_trades = []
    range_end_utc = _replay_range_end_utc(window_end_utc, getattr(args, "mt5_close_search_days", 14))
    for tf_name in resolve_run_tfs_for_strategy(2, args.tf):
        progress(f"Running S2 replay on {tf_name}...")
        trades = backtest_s2_tf(tf_name, S2_TF_MAP[tf_name], range_end_utc=range_end_utc)
        progress(f"S2 replay on {tf_name} produced {len(trades)} raw event(s).")
        filtered = [t for t in trades if window_start_utc <= t["entry_time"] <= window_end_utc]
        if args.exclude_cancelled:
            filtered = [
                t for t in filtered
                if t["close_type"] not in ("CANCEL", "PD_FAIL", "OPEN_PENDING", "BLOCK", "OPEN")
            ]
        for t in filtered:
            t.setdefault("sid", 2)
            t.setdefault("tf", tf_name)
        all_trades.extend((tf_name, t) for t in filtered)
        progress(f"S2 replay on {tf_name} kept {len(filtered)} event(s) in window.")
    return all_trades


def run_s3(args, window_start_utc: datetime, window_end_utc: datetime) -> list[tuple[str, dict]]:
    all_trades = []
    range_end_utc = _replay_range_end_utc(window_end_utc, getattr(args, "mt5_close_search_days", 14))
    for tf_name in resolve_run_tfs_for_strategy(3, args.tf):
        progress(f"Running S3 replay on {tf_name}...")
        trades = backtest_s3_tf(tf_name, S3_TF_MAP[tf_name], range_end_utc=range_end_utc)
        progress(f"S3 replay on {tf_name} produced {len(trades)} raw event(s).")
        filtered = [t for t in trades if window_start_utc <= t["entry_time"] <= window_end_utc]
        if args.exclude_cancelled:
            filtered = [
                t for t in filtered
                if t["close_type"] not in ("CANCEL", "PD_FAIL", "OPEN_PENDING", "BLOCK", "OPEN")
            ]
        for t in filtered:
            t.setdefault("sid", 3)
            t.setdefault("tf", tf_name)
        all_trades.extend((tf_name, t) for t in filtered)
        progress(f"S3 replay on {tf_name} kept {len(filtered)} event(s) in window.")
    return all_trades


def run_s4(args, window_start_utc: datetime, window_end_utc: datetime) -> list[tuple[str, dict]]:
    all_trades = []
    range_end_utc = _replay_range_end_utc(window_end_utc, getattr(args, "mt5_close_search_days", 14))
    for tf_name in resolve_run_tfs_for_strategy(4, args.tf):
        progress(f"Running S4 replay on {tf_name}...")
        trades = backtest_s4_tf(tf_name, S4_TF_MAP[tf_name], range_end_utc=range_end_utc)
        progress(f"S4 replay on {tf_name} produced {len(trades)} raw event(s).")
        filtered = [t for t in trades if window_start_utc <= t["entry_time"] <= window_end_utc]
        if args.exclude_cancelled:
            filtered = [
                t for t in filtered
                if t["close_type"] not in ("CANCEL", "PD_FAIL", "OPEN_PENDING", "BLOCK", "OPEN")
            ]
        for t in filtered:
            t.setdefault("sid", 4)
            t.setdefault("tf", tf_name)
        all_trades.extend((tf_name, t) for t in filtered)
        progress(f"S4 replay on {tf_name} kept {len(filtered)} event(s) in window.")
    return all_trades


def run_s5(args, window_start_utc: datetime, window_end_utc: datetime) -> list[tuple[str, dict]]:
    all_trades = []
    range_end_utc = _replay_range_end_utc(window_end_utc, getattr(args, "mt5_close_search_days", 14))
    for tf_name in resolve_run_tfs_for_strategy(5, args.tf):
        progress(f"Running S5 replay on {tf_name}...")
        trades = backtest_s5_tf(tf_name, S5_TF_MAP[tf_name], range_end_utc=range_end_utc)
        progress(f"S5 replay on {tf_name} produced {len(trades)} raw event(s).")
        filtered = [t for t in trades if window_start_utc <= t["entry_time"] <= window_end_utc]
        if args.exclude_cancelled:
            filtered = [
                t for t in filtered
                if t["close_type"] not in ("CANCEL", "PD_FAIL", "OPEN_PENDING", "BLOCK", "OPEN")
            ]
        for t in filtered:
            t.setdefault("sid", 5)
            t.setdefault("tf", tf_name)
        all_trades.extend((tf_name, t) for t in filtered)
        progress(f"S5 replay on {tf_name} kept {len(filtered)} event(s) in window.")
    return all_trades


def run_s8(args, window_start_utc: datetime, window_end_utc: datetime) -> list[tuple[str, dict]]:
    all_trades = []
    requested_tfs = set(resolve_run_tfs_for_strategy(8, args.tf))
    run_tfs = resolve_s8_context_tfs(args.tf)
    if set(run_tfs) != requested_tfs:
        progress(f"S8 context replay includes SL Guard Group TFs: {', '.join(run_tfs)}")

    raw_tf_trades = []
    range_end_utc = _replay_range_end_utc(window_end_utc, getattr(args, "mt5_close_search_days", 14))
    for tf_name in run_tfs:
        progress(f"Running S8 replay on {tf_name}...")
        trades = backtest_s8_tf(tf_name, S8_TF_MAP[tf_name], range_end_utc=range_end_utc)
        progress(f"S8 replay on {tf_name} produced {len(trades)} raw event(s).")
        for t in trades:
            t.setdefault("sid", 8)
            t.setdefault("tf", tf_name)
            raw_tf_trades.append((tf_name, t))

    if len(run_tfs) > 1:
        progress("Applying S8 SL Guard Group context overlay...")
        raw_tf_trades = sim_s14_backtest.apply_sl_guard_group_overlay(raw_tf_trades)

    for tf_name in run_tfs:
        tf_rows = [(tf, t) for tf, t in raw_tf_trades if tf == tf_name]
        filtered = [
            t for tf, t in tf_rows
            if tf in requested_tfs and window_start_utc <= t["entry_time"] <= window_end_utc
        ]
        if args.exclude_cancelled:
            filtered = [
                t for t in filtered
                if t["close_type"] not in ("CANCEL", "PD_FAIL", "OPEN_PENDING", "BLOCK", "OPEN")
            ]
        all_trades.extend((tf_name, t) for t in filtered)
        if tf_name in requested_tfs:
            progress(f"S8 replay on {tf_name} kept {len(filtered)} event(s) in window.")
    return all_trades


def run_s458_unified(args, window_start_utc: datetime, window_end_utc: datetime, strategies: set[int]) -> list[tuple[str, dict]]:
    all_trades = []
    requested_tfs, run_tfs = resolve_s458_context_tfs(
        args.tf,
        strategies,
        include_connected_s2_context=getattr(args, "s2_include_connected_fvg_context", False),
    )
    requested_set = set(requested_tfs)
    if len(run_tfs) > len(requested_tfs):
        kept_tfs = []
        for tf_name in run_tfs:
            if tf_name in requested_set or _s458_tf_history_overlaps_window(tf_name, window_start_utc, window_end_utc):
                kept_tfs.append(tf_name)
            else:
                progress(f"Skipping unified S1-S5/S8 context TF {tf_name}: fetched history does not overlap the requested window.")
        run_tfs = kept_tfs
    if set(run_tfs) != requested_set:
        context_parts = []
        if 8 in strategies and getattr(config, "SL_GUARD_GROUP_ENABLED", False):
            context_parts.append("SL Guard Group")
        if 2 in strategies and getattr(config, "FVG_PARALLEL", False):
            context_parts.append("S2 FVG Parallel")
        label = " + ".join(context_parts) if context_parts else "context"
        progress(f"Unified S1-S5/S8 context replay includes {label} TFs: {', '.join(run_tfs)}")

    raw_tf_trades = []
    range_end_utc = _replay_range_end_utc(window_end_utc, getattr(args, "mt5_close_search_days", 14))
    if strategy_set := set(strategies):
        use_multi_s2 = strategy_set == {2} and getattr(config, "FVG_PARALLEL", False) and len(run_tfs) > 1
    else:
        use_multi_s2 = False
    if use_multi_s2:
        progress(f"Running unified S2 multi-TF parallel replay on {', '.join(run_tfs)}...")
        trades = sim_s458_backtest.backtest_multi_tf(
            {tf_name: S8_TF_MAP[tf_name] for tf_name in run_tfs},
            strategies,
            range_end_utc=range_end_utc,
            scan_until_utc=window_end_utc,
            progress_cb=progress,
        )
        progress(f"Unified S2 multi-TF parallel replay produced {len(trades)} raw event(s).")
        for t in trades:
            tf_name = str(t.get("tf") or "")
            raw_tf_trades.append((tf_name, t))
    else:
        for tf_name in run_tfs:
            progress(f"Running unified S1-S5/S8 replay on {tf_name}...")
            trades = sim_s458_backtest.backtest_tf(
                tf_name,
                S8_TF_MAP[tf_name],
                strategies,
                range_end_utc=range_end_utc,
                scan_until_utc=window_end_utc,
                progress_cb=progress,
            )
            progress(f"Unified S1-S5/S8 replay on {tf_name} produced {len(trades)} raw event(s).")
            for t in trades:
                t.setdefault("tf", tf_name)
                raw_tf_trades.append((tf_name, t))

    if len(run_tfs) > 1 and not use_multi_s2:
        progress("Applying unified S1-S5/S8 SL Guard Group context overlay...")
        raw_tf_trades = sim_s14_backtest.apply_sl_guard_group_overlay(raw_tf_trades)

    for tf_name in run_tfs:
        tf_rows = [(tf, t) for tf, t in raw_tf_trades if tf == tf_name]
        if use_multi_s2:
            filtered_rows = [
                (_s2_multi_report_tf(t, tf, requested_set), t)
                for tf, t in tf_rows
                if _s2_multi_trade_matches_requested_tf(t, tf, requested_set)
                and window_start_utc <= t["entry_time"] <= window_end_utc
            ]
        else:
            filtered_rows = [
                (tf, t) for tf, t in tf_rows
                if tf in requested_set and window_start_utc <= t["entry_time"] <= window_end_utc
            ]
        _capture_raw_replay_context(args, filtered_rows)
        filtered = [t for _, t in filtered_rows]
        if args.exclude_cancelled:
            filtered_rows = [
                (report_tf, t) for report_tf, t in filtered_rows
                if t["close_type"] not in ("CANCEL", "PD_FAIL", "OPEN_PENDING", "BLOCK", "OPEN")
            ]
            filtered = [t for _, t in filtered_rows]
        all_trades.extend((report_tf, t) for report_tf, t in filtered_rows)
        if tf_name in requested_set:
            progress(f"Unified S1-S5/S8 replay on {tf_name} kept {len(filtered)} event(s) in window.")
    return all_trades


def run_s9(args, window_start_utc: datetime, window_end_utc: datetime) -> list[tuple[str, dict]]:
    all_trades = []
    requested_tfs = set(resolve_run_tfs_for_strategy(9, args.tf))
    run_tfs = resolve_guard_context_tfs(9, args.tf, S9_TF_MAP)
    if set(run_tfs) != requested_tfs:
        progress(f"S9 context replay includes SL Guard Group TFs: {', '.join(run_tfs)}")
    raw_tf_trades = []
    for tf_name in run_tfs:
        progress(f"Running S9 replay on {tf_name}...")
        trades = backtest_s9_tf(tf_name, S9_TF_MAP[tf_name], range_end_utc=window_end_utc)
        progress(f"S9 replay on {tf_name} produced {len(trades)} raw event(s).")
        for t in trades:
            t.setdefault("sid", 9)
            t.setdefault("tf", tf_name)
            raw_tf_trades.append((tf_name, t))

    if getattr(config, "SL_GUARD_GROUP_ENABLED", False) and getattr(config, "SL_GUARD_CLOSE_ON_ACTIVATE", True):
        progress("Applying S9 SL Guard Group context overlay...")
        raw_tf_trades = sim_s14_backtest.apply_sl_guard_group_overlay(raw_tf_trades)

    for tf_name in run_tfs:
        tf_rows = [(tf, t) for tf, t in raw_tf_trades if tf == tf_name]
        filtered = [
            t for tf, t in tf_rows
            if tf in requested_tfs and window_start_utc <= t["entry_time"] <= window_end_utc
        ]
        if args.exclude_cancelled:
            filtered = [
                t for t in filtered
                if t["close_type"] not in ("CANCEL", "PD_FAIL", "OPEN_PENDING", "BLOCK", "OPEN")
            ]
        all_trades.extend((tf_name, t) for t in filtered)
        if tf_name in requested_tfs:
            progress(f"S9 replay on {tf_name} kept {len(filtered)} event(s) in window.")
    return all_trades


def run_s11(args, window_start_utc: datetime, window_end_utc: datetime) -> list[tuple[str, dict]]:
    all_trades = []
    requested_tfs = set(resolve_run_tfs_for_strategy(11, args.tf))
    run_tfs = resolve_guard_context_tfs(11, args.tf, S11_TF_MAP)
    range_end_utc = _replay_range_end_utc(window_end_utc, getattr(args, "mt5_close_search_days", 14))
    if set(run_tfs) != requested_tfs:
        progress(f"S11 context replay includes SL Guard Group TFs: {', '.join(run_tfs)}")
    raw_tf_trades = []
    for tf_name in run_tfs:
        progress(f"Running S11 replay on {tf_name}...")
        trades = backtest_s11_tf(tf_name, S11_TF_MAP[tf_name], range_end_utc=range_end_utc)
        progress(f"S11 replay on {tf_name} produced {len(trades)} raw event(s).")
        for t in trades:
            t.setdefault("sid", 11)
            t.setdefault("tf", tf_name)
            raw_tf_trades.append((tf_name, t))

    if getattr(config, "SL_GUARD_GROUP_ENABLED", False) and getattr(config, "SL_GUARD_CLOSE_ON_ACTIVATE", True):
        progress("Applying S11 SL Guard Group context overlay...")
        raw_tf_trades = sim_s14_backtest.apply_sl_guard_group_overlay(raw_tf_trades)

    for tf_name in run_tfs:
        tf_rows = [(tf, t) for tf, t in raw_tf_trades if tf == tf_name]
        filtered = [
            t for tf, t in tf_rows
            if tf in requested_tfs and window_start_utc <= t["entry_time"] <= window_end_utc
        ]
        if args.exclude_cancelled:
            filtered = [
                t for t in filtered
                if t["close_type"] not in ("CANCEL", "PD_FAIL", "OPEN_PENDING", "BLOCK", "OPEN")
            ]
        all_trades.extend((tf_name, t) for t in filtered)
        if tf_name in requested_tfs:
            progress(f"S11 replay on {tf_name} kept {len(filtered)} event(s) in window.")
    return all_trades


def run_s12(args, window_start_utc: datetime, window_end_utc: datetime) -> list[tuple[str, dict]]:
    all_trades = []
    for tf_name in resolve_run_tfs_for_strategy(12, args.tf):
        progress(f"Running S12 replay on {tf_name}...")
        trades = backtest_s12_tf(tf_name, S12_TF_MAP[tf_name])
        progress(f"S12 replay on {tf_name} produced {len(trades)} raw event(s).")
        filtered = [t for t in trades if window_start_utc <= t["entry_time"] <= window_end_utc]
        if args.exclude_cancelled:
            filtered = [
                t for t in filtered
                if t["close_type"] not in ("CANCEL", "PD_FAIL", "OPEN_PENDING", "BLOCK", "OPEN")
            ]
        for t in filtered:
            t.setdefault("sid", 12)
            t.setdefault("tf", tf_name)
        all_trades.extend((tf_name, t) for t in filtered)
        progress(f"S12 replay on {tf_name} kept {len(filtered)} event(s) in window.")
    return all_trades


def run_s13(args, window_start_utc: datetime, window_end_utc: datetime) -> list[tuple[str, dict]]:
    all_trades = []
    requested_tfs = set(resolve_run_tfs_for_strategy(13, args.tf))
    run_tfs = resolve_guard_context_tfs(13, args.tf, S13_TF_MAP)
    if set(run_tfs) != requested_tfs:
        progress(f"S13 context replay includes SL Guard Group TFs: {', '.join(run_tfs)}")
    raw_tf_trades = []
    for tf_name in run_tfs:
        progress(f"Running S13 replay on {tf_name}...")
        trades = backtest_s13_tf(tf_name, S13_TF_MAP[tf_name])
        progress(f"S13 replay on {tf_name} produced {len(trades)} raw event(s).")
        for t in trades:
            t.setdefault("sid", 13)
            t.setdefault("tf", tf_name)
            raw_tf_trades.append((tf_name, t))

    if getattr(config, "SL_GUARD_GROUP_ENABLED", False) and getattr(config, "SL_GUARD_CLOSE_ON_ACTIVATE", True):
        progress("Applying S13 SL Guard Group context overlay...")
        raw_tf_trades = sim_s14_backtest.apply_sl_guard_group_overlay(raw_tf_trades)

    for tf_name in run_tfs:
        tf_rows = [(tf, t) for tf, t in raw_tf_trades if tf == tf_name]
        filtered = [
            t for tf, t in tf_rows
            if tf in requested_tfs and window_start_utc <= t["entry_time"] <= window_end_utc
        ]
        if args.exclude_cancelled:
            filtered = [
                t for t in filtered
                if t["close_type"] not in ("CANCEL", "PD_FAIL", "OPEN_PENDING", "BLOCK", "OPEN")
            ]
        all_trades.extend((tf_name, t) for t in filtered)
        if tf_name in requested_tfs:
            progress(f"S13 replay on {tf_name} kept {len(filtered)} event(s) in window.")
    return all_trades


def run_s14(args, window_start_utc: datetime, window_end_utc: datetime) -> list[tuple[str, dict]]:
    all_trades = []
    requested_tfs = set(resolve_run_tfs_for_strategy(14, args.tf))
    run_tfs = resolve_s14_context_tfs(args.tf)
    range_end_utc = _replay_range_end_utc(window_end_utc, getattr(args, "mt5_close_search_days", 14))
    if set(run_tfs) != requested_tfs:
        progress(f"S14 context replay includes SL Guard Group TFs: {', '.join(run_tfs)}")
    raw_tf_trades = []
    for tf_name in run_tfs:
        progress(f"Running S14 replay on {tf_name}...")
        trades = backtest_s14_tf(
            tf_name,
            S14_TF_MAP[tf_name],
            range_end_utc=range_end_utc,
            fill_next_bar=getattr(args, "s14_fill_next_bar", False),
        )
        progress(f"S14 replay on {tf_name} produced {len(trades)} raw event(s).")
        for t in trades:
            t.setdefault("sid", 14)
            t.setdefault("tf", tf_name)
            raw_tf_trades.append((tf_name, t))

    if len(run_tfs) > 1:
        progress("Applying S14 SL Guard Group overlay...")
        raw_tf_trades = sim_s14_backtest.apply_sl_guard_group_overlay(raw_tf_trades)
    if getattr(args, "hybrid_live_guard_context", False):
        guard_log_files = args.log_files if args.log_files else default_log_files()
        live_activations = load_live_sl_guard_activations(guard_log_files)
        if live_activations:
            progress(f"Applying hybrid live SL Guard context ({len(live_activations)} activation(s))...")
            raw_tf_trades = apply_live_sl_guard_context_overlay(raw_tf_trades, live_activations)

    for tf_name in run_tfs:
        tf_rows = [(tf, t) for tf, t in raw_tf_trades if tf == tf_name]
        filtered = [
            t for tf, t in tf_rows
            if tf in requested_tfs and window_start_utc <= t["entry_time"] <= window_end_utc
        ]
        if args.exclude_cancelled:
            filtered = [
                t for t in filtered
                if t["close_type"] not in ("CANCEL", "PD_FAIL", "OPEN_PENDING", "BLOCK")
            ]
        all_trades.extend((tf_name, t) for t in filtered)
        if tf_name in requested_tfs:
            progress(f"S14 replay on {tf_name} kept {len(filtered)} event(s) in window.")
    return all_trades


def run_s15(args, window_start_utc: datetime, window_end_utc: datetime) -> list[tuple[str, dict]]:
    all_trades = []
    requested_tfs = set(resolve_run_tfs_for_strategy(15, args.tf))
    run_tfs = resolve_guard_context_tfs(15, args.tf, S15_TF_MAP)
    if set(run_tfs) != requested_tfs:
        progress(f"S15 context replay includes SL Guard Group TFs: {', '.join(run_tfs)}")
    raw_tf_trades = []
    for tf_name in run_tfs:
        progress(f"Running S15 replay on {tf_name}...")
        trades = backtest_s15_tf(tf_name, S15_TF_MAP[tf_name])
        progress(f"S15 replay on {tf_name} produced {len(trades)} raw event(s).")
        for t in trades:
            t.setdefault("sid", 15)
            t.setdefault("tf", tf_name)
            raw_tf_trades.append((tf_name, t))

    if getattr(config, "SL_GUARD_GROUP_ENABLED", False) and getattr(config, "SL_GUARD_CLOSE_ON_ACTIVATE", True):
        progress("Applying S15 SL Guard Group context overlay...")
        raw_tf_trades = sim_s14_backtest.apply_sl_guard_group_overlay(raw_tf_trades)

    for tf_name in run_tfs:
        tf_rows = [(tf, t) for tf, t in raw_tf_trades if tf == tf_name]
        filtered = [
            t for tf, t in tf_rows
            if tf in requested_tfs and window_start_utc <= t["entry_time"] <= window_end_utc
        ]
        if args.exclude_cancelled:
            filtered = [
                t for t in filtered
                if t["close_type"] not in ("CANCEL", "PD_FAIL", "OPEN_PENDING", "BLOCK", "OPEN")
            ]
        all_trades.extend((tf_name, t) for t in filtered)
        if tf_name in requested_tfs:
            progress(f"S15 replay on {tf_name} kept {len(filtered)} event(s) in window.")
    return all_trades


def run_s16(args, window_start_utc: datetime, window_end_utc: datetime) -> list[tuple[str, dict]]:
    all_trades = []
    requested_tfs = set(resolve_run_tfs_for_strategy(16, args.tf))
    run_tfs = resolve_guard_context_tfs(16, args.tf, S16_TF_MAP)
    if set(run_tfs) != requested_tfs:
        progress(f"S16 context replay includes SL Guard Group TFs: {', '.join(run_tfs)}")
    raw_tf_trades = []
    for tf_name in run_tfs:
        progress(f"Running S16 replay on {tf_name}...")
        trades = backtest_s16_tf(tf_name, S16_TF_MAP[tf_name])
        progress(f"S16 replay on {tf_name} produced {len(trades)} raw event(s).")
        for t in trades:
            t.setdefault("sid", 16)
            t.setdefault("tf", tf_name)
            raw_tf_trades.append((tf_name, t))

    if getattr(config, "SL_GUARD_GROUP_ENABLED", False) and getattr(config, "SL_GUARD_CLOSE_ON_ACTIVATE", True):
        progress("Applying S16 SL Guard Group context overlay...")
        raw_tf_trades = sim_s14_backtest.apply_sl_guard_group_overlay(raw_tf_trades)

    for tf_name in run_tfs:
        tf_rows = [(tf, t) for tf, t in raw_tf_trades if tf == tf_name]
        filtered = [
            t for tf, t in tf_rows
            if tf in requested_tfs and window_start_utc <= t["entry_time"] <= window_end_utc
        ]
        if args.exclude_cancelled:
            filtered = [
                t for t in filtered
                if t["close_type"] not in ("CANCEL", "PD_FAIL", "OPEN_PENDING", "BLOCK", "OPEN")
            ]
        all_trades.extend((tf_name, t) for t in filtered)
        if tf_name in requested_tfs:
            progress(f"S16 replay on {tf_name} kept {len(filtered)} event(s) in window.")
    return all_trades


def run_s17(args, window_start_utc: datetime, window_end_utc: datetime) -> list[tuple[str, dict]]:
    all_trades = []
    requested_tfs = set(resolve_run_tfs_for_strategy(17, args.tf))
    run_tfs = resolve_guard_context_tfs(17, args.tf, S17_TF_MAP)
    if set(run_tfs) != requested_tfs:
        progress(f"S17 context replay includes SL Guard Group TFs: {', '.join(run_tfs)}")
    raw_tf_trades = []
    for tf_name in run_tfs:
        progress(f"Running S17 replay on {tf_name}...")
        trades = backtest_s17_tf(tf_name, S17_TF_MAP[tf_name])
        progress(f"S17 replay on {tf_name} produced {len(trades)} raw event(s).")
        for t in trades:
            t.setdefault("sid", 17)
            t.setdefault("tf", tf_name)
            raw_tf_trades.append((tf_name, t))

    if getattr(config, "SL_GUARD_GROUP_ENABLED", False) and getattr(config, "SL_GUARD_CLOSE_ON_ACTIVATE", True):
        progress("Applying S17 SL Guard Group context overlay...")
        raw_tf_trades = sim_s14_backtest.apply_sl_guard_group_overlay(raw_tf_trades)

    for tf_name in run_tfs:
        tf_rows = [(tf, t) for tf, t in raw_tf_trades if tf == tf_name]
        filtered = [
            t for tf, t in tf_rows
            if tf in requested_tfs and window_start_utc <= t["entry_time"] <= window_end_utc
        ]
        if args.exclude_cancelled:
            filtered = [
                t for t in filtered
                if t["close_type"] not in ("CANCEL", "PD_FAIL", "OPEN_PENDING", "BLOCK", "OPEN")
            ]
        all_trades.extend((tf_name, t) for t in filtered)
        if tf_name in requested_tfs:
            progress(f"S17 replay on {tf_name} kept {len(filtered)} event(s) in window.")
    return all_trades


def run_s18(args, window_start_utc: datetime, window_end_utc: datetime) -> list[tuple[str, dict]]:
    all_trades = []
    requested_tfs = set(resolve_run_tfs_for_strategy(18, args.tf))
    run_tfs = resolve_guard_context_tfs(18, args.tf, S18_TF_MAP)
    if set(run_tfs) != requested_tfs:
        progress(f"S18 context replay includes SL Guard Group TFs: {', '.join(run_tfs)}")
    raw_tf_trades = []
    for tf_name in run_tfs:
        progress(f"Running S18 replay on {tf_name}...")
        trades = backtest_s18_tf(tf_name, S18_TF_MAP[tf_name])
        progress(f"S18 replay on {tf_name} produced {len(trades)} raw event(s).")
        for t in trades:
            t.setdefault("sid", 18)
            t.setdefault("tf", tf_name)
            raw_tf_trades.append((tf_name, t))

    if getattr(config, "SL_GUARD_GROUP_ENABLED", False) and getattr(config, "SL_GUARD_CLOSE_ON_ACTIVATE", True):
        progress("Applying S18 SL Guard Group context overlay...")
        raw_tf_trades = sim_s14_backtest.apply_sl_guard_group_overlay(raw_tf_trades)

    for tf_name in run_tfs:
        tf_rows = [(tf, t) for tf, t in raw_tf_trades if tf == tf_name]
        filtered = [
            t for tf, t in tf_rows
            if tf in requested_tfs and window_start_utc <= t["entry_time"] <= window_end_utc
        ]
        if args.exclude_cancelled:
            filtered = [
                t for t in filtered
                if t["close_type"] not in ("CANCEL", "PD_FAIL", "OPEN_PENDING", "BLOCK", "OPEN")
            ]
        all_trades.extend((tf_name, t) for t in filtered)
        if tf_name in requested_tfs:
            progress(f"S18 replay on {tf_name} kept {len(filtered)} event(s) in window.")
    return all_trades


def run_s19(args, window_start_utc: datetime, window_end_utc: datetime) -> list[tuple[str, dict]]:
    all_trades = []
    requested_tfs = set(resolve_run_tfs_for_strategy(19, args.tf))
    run_tfs = resolve_guard_context_tfs(19, args.tf, S19_TF_MAP)
    if set(run_tfs) != requested_tfs:
        progress(f"S19 context replay includes SL Guard Group TFs: {', '.join(run_tfs)}")
    raw_tf_trades = []
    for tf_name in run_tfs:
        progress(f"Running S19 replay on {tf_name}...")
        trades = backtest_s19_tf(tf_name, S19_TF_MAP[tf_name])
        progress(f"S19 replay on {tf_name} produced {len(trades)} raw event(s).")
        for t in trades:
            t.setdefault("sid", 19)
            t.setdefault("tf", tf_name)
            raw_tf_trades.append((tf_name, t))

    if getattr(config, "SL_GUARD_GROUP_ENABLED", False) and getattr(config, "SL_GUARD_CLOSE_ON_ACTIVATE", True):
        progress("Applying S19 SL Guard Group context overlay...")
        raw_tf_trades = sim_s14_backtest.apply_sl_guard_group_overlay(raw_tf_trades)

    for tf_name in run_tfs:
        tf_rows = [(tf, t) for tf, t in raw_tf_trades if tf == tf_name]
        filtered = [
            t for tf, t in tf_rows
            if tf in requested_tfs and window_start_utc <= t["entry_time"] <= window_end_utc
        ]
        if args.exclude_cancelled:
            filtered = [
                t for t in filtered
                if t["close_type"] not in ("CANCEL", "PD_FAIL", "OPEN_PENDING", "BLOCK", "OPEN")
            ]
        all_trades.extend((tf_name, t) for t in filtered)
        if tf_name in requested_tfs:
            progress(f"S19 replay on {tf_name} kept {len(filtered)} event(s) in window.")
    return all_trades


def apply_system_sl_guard_group_overlay(
    trades: list[tuple[str, dict]],
    strategies: set[int],
) -> list[tuple[str, dict]]:
    if len(strategies) <= 1:
        return trades
    if not getattr(config, "SL_GUARD_GROUP_ENABLED", False):
        return trades
    if not getattr(config, "SL_GUARD_CLOSE_ON_ACTIVATE", True):
        return trades
    progress("Applying system-level SL Guard Group overlay across selected strategies...")
    before = {
        id(trade)
        for _, trade in trades
        if str(trade.get("close_type", "")) == "SL_GUARD_GROUP"
        and trade.get("sl_guard_trigger_tf")
    }
    result = sim_s14_backtest.apply_sl_guard_group_overlay(trades)
    changed = sum(
        1 for _, trade in result
        if str(trade.get("close_type", "")) == "SL_GUARD_GROUP"
        and trade.get("sl_guard_trigger_tf")
        and id(trade) not in before
    )
    if changed:
        progress(f"System SL Guard Group overlay closed {changed} trade(s).")
    else:
        progress("System SL Guard Group overlay made no additional closes.")
    return result


def _system_overlay_pnl(trade: dict, close_price: float) -> float:
    volume = float(trade.get("volume", getattr(config, "VOLUME", 0.01)) or 0.01)
    multiplier = 100.0 * volume
    if str(trade.get("signal") or trade.get("side") or "").upper() == "BUY":
        pnl = (float(close_price) - float(trade.get("entry", 0.0) or 0.0)) * multiplier
    else:
        pnl = (float(trade.get("entry", 0.0) or 0.0) - float(close_price)) * multiplier
    return round(float(trade.get("realized_pnl", 0.0) or 0.0) + pnl, 2)


def apply_system_opposite_order_overlay(
    trades: list[tuple[str, dict]],
    strategies: set[int],
) -> list[tuple[str, dict]]:
    if len(strategies) <= 1:
        return trades
    if not getattr(config, "OPPOSITE_ORDER_ENABLED", True):
        return trades

    mode = str(getattr(config, "OPPOSITE_ORDER_MODE", "sl_protect") or "sl_protect")
    if mode not in ("sl_protect", "tp_close"):
        return trades

    rows = []
    skip_sids = set(getattr(sim_lifecycle, "OPPOSITE_ORDER_SKIP_SIDS", set()))
    for idx, (tf_name, trade) in enumerate(trades):
        side = str(trade.get("signal") or trade.get("side") or "").upper()
        entry_time = trade.get("entry_time")
        close_time = trade.get("close_time")
        if side not in ("BUY", "SELL") or not entry_time or not close_time:
            rows.append({"idx": idx, "tf": tf_name, "trade": trade, "eligible": False})
            continue
        if int(trade.get("sid", 0) or 0) in skip_sids:
            rows.append({"idx": idx, "tf": tf_name, "trade": trade, "eligible": False})
            continue
        trade.setdefault("tf", tf_name)
        rows.append({"idx": idx, "tf": tf_name, "trade": trade, "eligible": True})

    changed = 0
    progress("Applying system-level Opposite Order overlay across selected strategies...")
    for current in sorted([r for r in rows if r["eligible"]], key=lambda r: r["trade"]["entry_time"]):
        cur_trade = current["trade"]
        cur_side = str(cur_trade.get("signal") or cur_trade.get("side") or "").upper()
        cur_tf = str(cur_trade.get("tf") or current["tf"] or "")
        cur_entry_time = cur_trade.get("entry_time")
        cur_entry_price = float(cur_trade.get("entry", 0.0) or 0.0)
        if not cur_tf or cur_entry_price <= 0:
            continue

        for prior in rows:
            if not prior["eligible"] or prior is current:
                continue
            old = prior["trade"]
            old_side = str(old.get("signal") or old.get("side") or "").upper()
            if old_side == cur_side:
                continue
            if old.get("system_opposite_overlay"):
                continue
            if str(old.get("tf") or prior["tf"] or "") != cur_tf:
                continue
            if not (old.get("entry_time") < cur_entry_time < old.get("close_time")):
                continue

            if mode == "tp_close":
                old["close_type"] = "OPPOSITE_CLOSE"
                old["close_price"] = round(cur_entry_price, 2)
                old["close_time"] = cur_entry_time
                old["pnl"] = _system_overlay_pnl(old, cur_entry_price)
                old["profit"] = old["pnl"]
                old["reason"] = "System Opposite Order overlay"
                old["system_opposite_overlay"] = "tp_close"
                changed += 1
                continue

            close_type = str(old.get("close_type", ""))
            if close_type not in ("SL", "SL_GUARD_CLOSE", "SL_GUARD_GROUP"):
                continue
            spread = 0.0
            protected_sl = float(old.get("entry", 0.0) or 0.0)
            if old_side == "BUY":
                should_protect = float(old.get("close_price", 0.0) or 0.0) < protected_sl
            else:
                should_protect = float(old.get("close_price", 0.0) or 0.0) > protected_sl
            if not should_protect:
                continue
            old["close_type"] = "OPPOSITE_SL_PROTECT"
            old["close_price"] = round(protected_sl + spread if old_side == "BUY" else protected_sl - spread, 2)
            old["pnl"] = _system_overlay_pnl(old, old["close_price"])
            old["profit"] = old["pnl"]
            old["reason"] = "System Opposite Order SL protect overlay"
            old["system_opposite_overlay"] = "sl_protect"
            changed += 1

    if changed:
        progress(f"System Opposite Order overlay adjusted {changed} trade(s).")
    else:
        progress("System Opposite Order overlay made no additional adjustments.")
    return [(row["tf"], row["trade"]) for row in sorted(rows, key=lambda r: r["idx"])]


def apply_system_limit_guard_overlay(
    trades: list[tuple[str, dict]],
    strategies: set[int],
    *,
    exclude_cancelled: bool = False,
) -> list[tuple[str, dict]]:
    if len(strategies) <= 1:
        return trades
    if not getattr(config, "LIMIT_GUARD", False):
        return trades

    symbol_info = mt5.symbol_info(config.SYMBOL)
    point = float(getattr(symbol_info, "point", 0.01) or 0.01)
    guard_dist = float(getattr(config, "LIMIT_GUARD_POINTS", 200) or 200) * point * config.points_scale()
    tf_separate = str(getattr(config, "LIMIT_GUARD_TF_MODE", "separate")) == "separate"
    skip_sids = set(getattr(sim_lifecycle, "LIMIT_GUARD_SKIP_SIDS", set()))

    rows = []
    for idx, (tf_name, trade) in enumerate(trades):
        side = str(trade.get("signal") or trade.get("side") or "").upper()
        entry_time = trade.get("entry_time")
        close_time = trade.get("close_time")
        if side not in ("BUY", "SELL") or not entry_time:
            rows.append({"idx": idx, "tf": tf_name, "trade": trade, "eligible": False})
            continue
        if int(trade.get("sid", 0) or 0) in skip_sids:
            rows.append({"idx": idx, "tf": tf_name, "trade": trade, "eligible": False})
            continue
        trade.setdefault("tf", tf_name)
        rows.append({"idx": idx, "tf": tf_name, "trade": trade, "eligible": bool(close_time)})

    changed = 0
    progress("Applying system-level Limit Guard overlay across selected strategies...")
    active_rows: list[dict] = []
    for row in sorted([r for r in rows if r["eligible"]], key=lambda r: r["trade"]["entry_time"]):
        trade = row["trade"]
        side = str(trade.get("signal") or trade.get("side") or "").upper()
        tf_name = str(trade.get("tf") or row["tf"] or "")
        entry = float(trade.get("entry", 0.0) or 0.0)
        entry_time = trade.get("entry_time")
        blocked_reason = ""

        for prior in active_rows:
            old = prior["trade"]
            if str(old.get("signal") or old.get("side") or "").upper() != side:
                continue
            if tf_separate and str(old.get("tf") or prior["tf"] or "") != tf_name:
                continue
            if not (old.get("entry_time") < entry_time < old.get("close_time")):
                continue
            pos_entry = float(old.get("entry", 0.0) or 0.0)
            if side == "BUY" and entry > pos_entry and entry > pos_entry + guard_dist:
                blocked_reason = (
                    f"System Limit Guard [{tf_name}->{old.get('tf', '?')}]: BUY LIMIT {entry:.2f} > "
                    f"BUY pos {pos_entry:.2f} & price {entry:.2f} > {pos_entry + guard_dist:.2f}"
                )
                break
            if side == "SELL" and entry < pos_entry and entry < pos_entry - guard_dist:
                blocked_reason = (
                    f"System Limit Guard [{tf_name}->{old.get('tf', '?')}]: SELL LIMIT {entry:.2f} < "
                    f"SELL pos {pos_entry:.2f} & price {entry:.2f} < {pos_entry - guard_dist:.2f}"
                )
                break

        if blocked_reason:
            trade["close_type"] = "CANCEL"
            trade["close_price"] = None
            trade["close_time"] = entry_time
            trade["pnl"] = 0.0
            trade["profit"] = 0.0
            trade["reason"] = blocked_reason
            trade["cancel_reason"] = blocked_reason
            trade["system_limit_guard_overlay"] = True
            changed += 1
            continue

        active_rows.append(row)

    if changed:
        progress(f"System Limit Guard overlay cancelled {changed} trade(s).")
    else:
        progress("System Limit Guard overlay made no additional cancels.")

    result = [(row["tf"], row["trade"]) for row in sorted(rows, key=lambda r: r["idx"])]
    if exclude_cancelled:
        result = [
            (tf_name, trade) for tf_name, trade in result
            if not trade.get("system_limit_guard_overlay")
        ]
    return result


def apply_system_same_bar_duplicate_overlay(
    trades: list[tuple[str, dict]],
    strategies: set[int],
) -> list[tuple[str, dict]]:
    target_sids = {4, 5, 8}
    if not (set(strategies) & target_sids):
        return trades

    progress("Applying system-level same-bar duplicate overlay across selected strategies...")
    seen: set[tuple] = set()
    drop_ids: set[int] = set()
    ordered = sorted(
        enumerate(trades),
        key=lambda item: (
            item[1][1].get("entry_time") or item[1][1].get("detect_time") or datetime.max.replace(tzinfo=timezone.utc),
            int(item[1][1].get("sid", 999) or 999),
            item[0],
        ),
    )
    for idx, (tf_name, trade) in ordered:
        sid = int(trade.get("sid", 0) or 0)
        if sid not in target_sids:
            continue
        side = str(trade.get("signal") or trade.get("side") or "").upper()
        if side not in ("BUY", "SELL"):
            continue
        entry_time_raw = int(
            trade.get("entry_time_raw")
            or trade.get("detect_time_raw")
            or 0
        )
        entry = round(float(trade.get("entry", 0.0) or 0.0), 2)
        if not entry_time_raw or entry <= 0:
            continue
        key = (str(trade.get("tf") or tf_name or ""), sid, side, entry_time_raw, entry)
        if key in seen:
            trade["system_duplicate_overlay"] = True
            trade["cancel_reason"] = "System same-bar duplicate setup overlay"
            drop_ids.add(idx)
            continue
        seen.add(key)

    if drop_ids:
        progress(f"System same-bar duplicate overlay removed {len(drop_ids)} duplicate event(s).")
    else:
        progress("System same-bar duplicate overlay made no removals.")
    return [row for idx, row in enumerate(trades) if idx not in drop_ids]


def main() -> None:
    parser = argparse.ArgumentParser(description="Backtest auto-trade flow using bot_state/config where implemented.")
    parser.add_argument("--start", required=True, help="Bangkok start time, e.g. 2026-06-01 00:00")
    parser.add_argument("--end", required=True, help="Bangkok end time, e.g. 2026-06-06 05:00")
    parser.add_argument("--tf", help="HTF/LTF to test. For S10, H1 maps to M1.")
    parser.add_argument("--strategies", default="active", help="active, all, or list/range like 10 or 1,9-10,15")
    parser.add_argument("--context-strategies", help="Diagnostic only: run extra replay strategies for shared state while reporting/comparing only --strategies.")
    parser.add_argument("--symbol", help="Symbol. If omitted, uses bot_state.json then config.SYMBOL.")
    parser.add_argument("--since", help="Simulation start in Bangkok time. Defaults to sim module SINCE.")
    parser.add_argument("--exclude-cancelled", "--only-filled", action="store_true", dest="exclude_cancelled")
    parser.add_argument("--compare-live", action="store_true", help="Compare backtest filled trades with ENTRY_FILL/POSITION_CLOSED logs.")
    parser.add_argument("--compare-mt5-history", action="store_true", help="Compare backtest filled trades with MT5 history deals/orders.")
    parser.add_argument("--compare-source", choices=("log", "mt5", "both"), default="log", help="Source used when --compare-live is set.")
    parser.add_argument("--log-files", nargs="*", help="Log files to compare. Defaults to logs/backtest_bot.log and backtest archives; never reads live bot.log unless explicitly passed.")
    parser.add_argument("--match-minutes", type=float, default=180.0, help="Max fill-time difference for live/backtest matching.")
    parser.add_argument("--match-entry-points", type=float, default=1.0, help="Max entry price difference for live/backtest matching.")
    parser.add_argument("--max-match-quality", choices=("exact", "near", "loose"), default="loose", help="Highest match quality allowed. Use near/exact to avoid loose matches.")
    parser.add_argument("--pnl-tolerance", type=float, default=1.0, help="P&L difference threshold for mismatch reporting.")
    parser.add_argument("--prefer-same-s14-family", action="store_true", help="Prefer S14 matches with the same normalized family. Diagnostic mode; default keeps time/entry priority.")
    parser.add_argument("--s14-disable-rsi-div", action="store_true", help="Diagnostic only: temporarily set S14_RSI_DIV_ENABLED=False for this replay run.")
    parser.add_argument("--s14-enable-sweep-return", action="store_true", help="Diagnostic only: temporarily set S14_SWEEP_RETURN=True for this replay run.")
    parser.add_argument("--s14-fill-next-bar", action="store_true", help="Diagnostic only: fill S14 market entries on the next bar open time after the closed-bar signal, matching scanner timing more closely.")
    parser.add_argument("--s2-include-connected-fvg-context", action="store_true", help="Diagnostic only: include one-hop connected FVG parallel TFs such as M1 via M5 for an M15 S2 run. Can be slow.")
    parser.add_argument("--s2-parallel-lifecycle-tf", action="store_true", help="Diagnostic only: manage S2 parallel fills/SL/TP on the smallest TF in the parallel group instead of the signal TF.")
    parser.add_argument("--s2-disable-sweep-filter", action="store_true", help="Diagnostic only: temporarily set SWEEP_FILTER_ENABLED=False for this replay run.")
    parser.add_argument("--s2-fill-before-cancel-bars", action="store_true", help="Diagnostic only: when a cancel_bars expiry bar also touches entry, let S2 fill before cancelling.")
    parser.add_argument("--s3-disable-pd-fibo-plus", action="store_true", help="Diagnostic only: temporarily add S3 to PDFIBOPLUS_SKIP_SIDS for this replay run.")
    parser.add_argument("--mt5-close-search-days", type=int, default=14, help="Extra MT5 history days after --end to find exits for filled orders.")
    parser.add_argument("--compare-csv", nargs="?", const="", help="Optional CSV path/name for compare detail. Defaults under excel_reports/backtest_compare.")
    parser.add_argument("--compare-xlsx", nargs="?", const="", help="Optional Excel .xlsx path/name for compare detail. Defaults under excel_reports/backtest_compare.")
    parser.add_argument("--dump-trades-csv", nargs="?", const="", help="Optional raw replay events CSV, including cancelled/open events. Defaults under excel_reports/backtest_compare.")
    parser.add_argument("--hybrid-live-guard-context", action="store_true", help="Apply live SL Guard Group activations from logs as compare-time overlay for replay trades.")
    parser.add_argument("--allow-restore-fail", action="store_true", help="Diagnostic only: continue even when bot_state/config restore fails.")
    args = parser.parse_args()
    validate_backtest_log_files(args.log_files or [])

    start_bkk = parse_bkk_dt(args.start)
    end_bkk = parse_bkk_dt(args.end)
    if end_bkk < start_bkk:
        raise ValueError("End datetime must be after start datetime")
    window_start_utc = start_bkk.replace(tzinfo=timezone.utc)
    window_end_utc = end_bkk.replace(tzinfo=timezone.utc)

    if args.since:
        since_utc = parse_bkk_dt(args.since).replace(tzinfo=timezone.utc) - timedelta(hours=7)
        sim_s1_backtest.SINCE = since_utc
        sim_s2_backtest.SINCE = since_utc
        sim_s3_backtest.SINCE = since_utc
        sim_s4_backtest.SINCE = since_utc
        sim_s5_backtest.SINCE = since_utc
        sim_s458_backtest.SINCE = since_utc
        sim_s8_backtest.SINCE = since_utc
        sim_s9_backtest.SINCE = since_utc
        sim_s10_backtest.SINCE = since_utc
        sim_s11_backtest.SINCE = since_utc
        sim_s12_backtest.SINCE = since_utc
        sim_s13_backtest.SINCE = since_utc
        sim_s14_backtest.SINCE = since_utc
        sim_s15_backtest.SINCE = since_utc
        sim_s16_backtest.SINCE = since_utc
        sim_s17_backtest.SINCE = since_utc
        sim_s18_backtest.SINCE = since_utc
        sim_s19_backtest.SINCE = since_utc

    progress("Initializing MT5...")
    if not mt5.initialize():
        print("MT5 init failed:", mt5.last_error())
        return

    progress("Restoring bot_state/config...")
    restore_info = config.restore_runtime_state()
    if not restore_info.get("restored") and not args.allow_restore_fail:
        print(f"Restore failed: {restore_info.get('reason', restore_info)}")
        print("Refusing to run backtest because restored bot_state/config is required for real-bot parity.")
        mt5.shutdown()
        raise SystemExit(2)
    s14_diagnostic_overrides = []
    if args.s14_disable_rsi_div:
        config.S14_RSI_DIV_ENABLED = False
        s14_diagnostic_overrides.append("S14_RSI_DIV_ENABLED=False")
    if args.s14_enable_sweep_return:
        config.S14_SWEEP_RETURN = True
        s14_diagnostic_overrides.append("S14_SWEEP_RETURN=True")
    s2_diagnostic_overrides = []
    if args.s2_disable_sweep_filter:
        config.SWEEP_FILTER_ENABLED = False
        s2_diagnostic_overrides.append("SWEEP_FILTER_ENABLED=False")
    sim_s458_backtest.S2_PARALLEL_LIFECYCLE_TF = bool(args.s2_parallel_lifecycle_tf)
    sim_s458_backtest.S2_FILL_BEFORE_CANCEL_BARS = bool(args.s2_fill_before_cancel_bars)
    s3_diagnostic_overrides = []
    if args.s3_disable_pd_fibo_plus:
        pd_skip_sids = set(getattr(config, "PDFIBOPLUS_SKIP_SIDS", ()))
        pd_skip_sids.add(3)
        config.PDFIBOPLUS_SKIP_SIDS = tuple(sorted(pd_skip_sids))
        sim_lifecycle.PDFIBOPLUS_SKIP_SIDS = set(config.PDFIBOPLUS_SKIP_SIDS)
        s3_diagnostic_overrides.append("PDFIBOPLUS_SKIP_SIDS+=S3")
    selected_symbol = args.symbol or load_state_symbol() or config.SYMBOL
    config.set_runtime_symbol(selected_symbol)
    try:
        from log_sources import ensure_backtest_log_files
        ensure_backtest_log_files()
    except Exception:
        pass
    sim_s1_backtest.SYMBOL = selected_symbol
    sim_s2_backtest.SYMBOL = selected_symbol
    sim_s3_backtest.SYMBOL = selected_symbol
    sim_s4_backtest.SYMBOL = selected_symbol
    sim_s5_backtest.SYMBOL = selected_symbol
    sim_s458_backtest.SYMBOL = selected_symbol
    sim_s8_backtest.SYMBOL = selected_symbol
    sim_s9_backtest.SYMBOL = selected_symbol
    sim_s10_backtest.SYMBOL = selected_symbol
    sim_s11_backtest.SYMBOL = selected_symbol
    sim_s12_backtest.SYMBOL = selected_symbol
    sim_s13_backtest.SYMBOL = selected_symbol
    sim_s14_backtest.SYMBOL = selected_symbol
    sim_s15_backtest.SYMBOL = selected_symbol
    sim_s16_backtest.SYMBOL = selected_symbol
    sim_s17_backtest.SYMBOL = selected_symbol
    sim_s18_backtest.SYMBOL = selected_symbol
    sim_s19_backtest.SYMBOL = selected_symbol
    sync_strategy10_runtime_config()

    strategies = parse_strategy_list(args.strategies)
    requested_context_strategies = parse_strategy_list(args.context_strategies) if args.context_strategies else []
    context_strategies = [sid for sid in requested_context_strategies if config.active_strategies.get(sid, False)]
    skipped_context_strategies = [sid for sid in requested_context_strategies if sid not in context_strategies]
    replay_strategies = sorted(set(strategies) | set(context_strategies))
    unsupported = [sid for sid in replay_strategies if sid not in SUPPORTED_STRATEGIES]

    print(f"Symbol   : {config.SYMBOL}")
    print(f"Restore  : {restore_info}")
    if s14_diagnostic_overrides:
        print(f"S14 Diagnostic Overrides: {', '.join(s14_diagnostic_overrides)}")
    if s2_diagnostic_overrides:
        print(f"S2 Diagnostic Overrides: {', '.join(s2_diagnostic_overrides)}")
    if s3_diagnostic_overrides:
        print(f"S3 Diagnostic Overrides: {', '.join(s3_diagnostic_overrides)}")
    print(f"Window   : {start_bkk} -> {end_bkk} (Bangkok timezone)")
    print(f"Strategy : {strategies}")
    if requested_context_strategies:
        print(f"Context  : {context_strategies or '-'} (requested: {requested_context_strategies}; replay set: {replay_strategies}; report stays: {strategies})")
        if skipped_context_strategies:
            print(f"Context skipped because OFF in restored config: {skipped_context_strategies}")
    print_feature_snapshot()
    if 1 in replay_strategies:
        print_s1_coverage()
    if 2 in replay_strategies:
        print_s2_coverage()
    if 3 in replay_strategies:
        print_s3_coverage()
    if 4 in replay_strategies:
        print_s4_coverage()
    if 5 in replay_strategies:
        print_s5_coverage()
    if 8 in replay_strategies:
        print_s8_coverage()
    if 9 in replay_strategies:
        print_s9_coverage()
    if 10 in replay_strategies:
        print_s10_coverage()
    if 11 in replay_strategies:
        print_s11_coverage()
    if 12 in replay_strategies:
        print_s12_coverage()
    if 13 in replay_strategies:
        print_s13_coverage()
    if 14 in replay_strategies:
        print_s14_coverage()
    if 15 in replay_strategies:
        print_s15_coverage()
    if 16 in replay_strategies:
        print_s16_coverage()
    if 17 in replay_strategies:
        print_s17_coverage()
    if 18 in replay_strategies:
        print_s18_coverage()
    if 19 in replay_strategies:
        print_s19_coverage()
    if unsupported:
        print(f"\nNot implemented in this replay engine yet: {unsupported}")
    print("=" * 72)

    all_trades = []
    args._raw_replay_context_trades = []
    strategy_set = set(replay_strategies)
    unified_s458 = (
        strategy_set.issubset({1, 2, 3, 4, 5, 8})
        and (len(strategy_set) > 1 or bool(strategy_set & {1, 2, 3}))
    )

    if unified_s458:
        for sid in sorted(strategy_set):
            if not config.active_strategies.get(sid, False):
                progress(f"S{sid} selected but OFF in restored config; replay still runs for requested strategy audit.")
        all_trades.extend(run_s458_unified(args, window_start_utc, window_end_utc, strategy_set))

    if not unified_s458 and 1 in replay_strategies:
        if not config.active_strategies.get(1, False):
            progress("S1 selected but OFF in restored config; replay still runs for requested strategy audit.")
        all_trades.extend(run_s1(args, window_start_utc, window_end_utc))
    if not unified_s458 and 2 in replay_strategies:
        if not config.active_strategies.get(2, False):
            progress("S2 selected but OFF in restored config; replay still runs for requested strategy audit.")
        all_trades.extend(run_s2(args, window_start_utc, window_end_utc))
    if not unified_s458 and 3 in replay_strategies:
        if not config.active_strategies.get(3, False):
            progress("S3 selected but OFF in restored config; replay still runs for requested strategy audit.")
        all_trades.extend(run_s3(args, window_start_utc, window_end_utc))
    if not unified_s458 and 4 in replay_strategies:
        if not config.active_strategies.get(4, False):
            progress("S4 selected but OFF in restored config; replay still runs for requested strategy audit.")
        all_trades.extend(run_s4(args, window_start_utc, window_end_utc))
    if not unified_s458 and 5 in replay_strategies:
        if not config.active_strategies.get(5, False):
            progress("S5 selected but OFF in restored config; replay still runs for requested strategy audit.")
        all_trades.extend(run_s5(args, window_start_utc, window_end_utc))
    if not unified_s458 and 8 in replay_strategies:
        if not config.active_strategies.get(8, False):
            progress("S8 selected but OFF in restored config; replay still runs for requested strategy audit.")
        all_trades.extend(run_s8(args, window_start_utc, window_end_utc))
    if not unified_s458 and 9 in replay_strategies:
        if not config.active_strategies.get(9, False):
            progress("S9 selected but OFF in restored config; replay still runs for requested strategy audit.")
        all_trades.extend(run_s9(args, window_start_utc, window_end_utc))
    if 10 in replay_strategies:
        if not config.active_strategies.get(10, False):
            progress("S10 selected but OFF in restored config; replay still runs for requested strategy audit.")
        all_trades.extend(run_s10(args, window_start_utc, window_end_utc))
    if 11 in replay_strategies:
        if not config.active_strategies.get(11, False):
            progress("S11 selected but OFF in restored config; replay still runs for requested strategy audit.")
        all_trades.extend(run_s11(args, window_start_utc, window_end_utc))
    if 12 in replay_strategies:
        if not config.active_strategies.get(12, False):
            progress("S12 selected but OFF in restored config; replay still runs for requested strategy audit.")
        all_trades.extend(run_s12(args, window_start_utc, window_end_utc))
    if 13 in replay_strategies:
        if not config.active_strategies.get(13, False):
            progress("S13 selected but OFF in restored config; replay still runs for requested strategy audit.")
        all_trades.extend(run_s13(args, window_start_utc, window_end_utc))
    if 14 in replay_strategies:
        if not config.active_strategies.get(14, False):
            progress("S14 selected but OFF in restored config; replay still runs for requested strategy audit.")
        all_trades.extend(run_s14(args, window_start_utc, window_end_utc))
    if 15 in replay_strategies:
        if not config.active_strategies.get(15, False):
            progress("S15 selected but OFF in restored config; replay still runs for requested strategy audit.")
        all_trades.extend(run_s15(args, window_start_utc, window_end_utc))
    if 16 in replay_strategies:
        if not config.active_strategies.get(16, False):
            progress("S16 selected but OFF in restored config; replay still runs for requested strategy audit.")
        all_trades.extend(run_s16(args, window_start_utc, window_end_utc))
    if 17 in replay_strategies:
        if not config.active_strategies.get(17, False):
            progress("S17 selected but OFF in restored config; replay still runs for requested strategy audit.")
        all_trades.extend(run_s17(args, window_start_utc, window_end_utc))
    if 18 in replay_strategies:
        if not config.active_strategies.get(18, False):
            progress("S18 selected but OFF in restored config; replay still runs for requested strategy audit.")
        all_trades.extend(run_s18(args, window_start_utc, window_end_utc))
    if 19 in replay_strategies:
        if not config.active_strategies.get(19, False):
            progress("S19 selected but OFF in restored config; replay still runs for requested strategy audit.")
        all_trades.extend(run_s19(args, window_start_utc, window_end_utc))

    all_trades = apply_system_limit_guard_overlay(all_trades, set(replay_strategies), exclude_cancelled=args.exclude_cancelled)
    all_trades = apply_system_same_bar_duplicate_overlay(all_trades, set(replay_strategies))
    all_trades = apply_system_opposite_order_overlay(all_trades, set(replay_strategies))
    all_trades = apply_system_sl_guard_group_overlay(all_trades, set(replay_strategies))
    all_trades = sorted(
        all_trades,
        key=lambda row: (
            row[1].get("entry_time") or datetime.max.replace(tzinfo=timezone.utc),
            row[0],
            int(row[1].get("sid", 0) or 0),
            str(row[1].get("signal", "")),
            float(row[1].get("entry", 0.0) or 0.0),
        ),
    )
    report_strategy_set = set(strategies)
    report_trades = [(tf, trade) for tf, trade in all_trades if int(trade.get("sid", 0) or 0) in report_strategy_set]
    if context_strategies:
        progress(f"Context strategy filter: replay rows {len(all_trades)} -> report rows {len(report_trades)} for S{','.join(str(s) for s in strategies)}.")

    report_variants = []
    if args.hybrid_live_guard_context:
        report_variants.append("hybrid_guard")
    if args.prefer_same_s14_family:
        report_variants.append("s14_family")
    if args.s14_disable_rsi_div:
        report_variants.append("s14_no_rsi_div")
    if args.s14_enable_sweep_return:
        report_variants.append("s14_sweep_return")
    if args.s14_fill_next_bar:
        report_variants.append("s14_next_bar")
    if args.s2_include_connected_fvg_context:
        report_variants.append("s2_connected_fvg")
    if args.s2_parallel_lifecycle_tf:
        report_variants.append("s2_lifecycle_tf")
    if args.s2_disable_sweep_filter:
        report_variants.append("s2_no_sweep_filter")
    if args.s2_fill_before_cancel_bars:
        report_variants.append("s2_fill_before_cancel")
    if args.s3_disable_pd_fibo_plus:
        report_variants.append("s3_no_pd")
    if context_strategies:
        report_variants.append("ctx_" + "-".join(str(s) for s in context_strategies))
    report_variant = "_".join(report_variants) if report_variants else None
    default_report_name = default_compare_report_base(
        start_bkk,
        end_bkk,
        args.tf,
        strategies,
        variant=report_variant,
    )

    if args.dump_trades_csv is not None:
        trades_path = resolve_compare_output_path(
            args.dump_trades_csv,
            default_report_name + "_trades",
            ".csv",
            strategies,
        )
        progress(f"Writing raw replay trades CSV: {trades_path}")
        written_trades_csv = write_trades_csv(trades_path, report_trades)
        if written_trades_csv != trades_path:
            print(f"\nReplay Trades CSV: {written_trades_csv} (requested file was locked)")
        else:
            print(f"\nReplay Trades CSV: {written_trades_csv}")

    groups = defaultdict(list)
    for htf, trade in report_trades:
        groups[htf].append(trade)

    grand_total = 0.0
    for htf, trades in groups.items():
        total = sum(t["pnl"] for t in trades)
        grand_total += total
        ltf = HTF_TO_LTF.get(htf, "") if all(int(t.get("sid", 10) or 10) == 10 for t in trades) else ""
        tf_display = f"{htf} ({ltf})" if ltf else htf
        print(f"\n## {tf_display} ({len(trades)} events)  P&L={total:+.2f} USD")
        for idx, trade in enumerate(trades, 1):
            format_trade(tf_display, idx, trade)

    print("\n" + "=" * 72)
    print(f"GRAND TOTAL: {grand_total:+.2f} USD")

    if args.compare_live or args.compare_mt5_history:
        live_rows = []
        if args.compare_live and args.compare_source in ("log", "both"):
            log_files = args.log_files if args.log_files else default_log_files()
            live_rows.extend(load_live_filled_orders(log_files, start_bkk, end_bkk, config.SYMBOL, set(strategies)))
        if args.compare_mt5_history or (args.compare_live and args.compare_source in ("mt5", "both")):
            live_rows.extend(load_mt5_history_orders(
                start_bkk,
                end_bkk,
                config.SYMBOL,
                set(strategies),
                close_search_days=max(0, int(args.mt5_close_search_days)),
            ))
        live_rows = _dedupe_live_rows(live_rows)
        trail_log_files = args.log_files if args.log_files else default_log_files()
        enrich_rows_with_trail_logs(live_rows, trail_log_files)
        enrich_rows_with_sl_guard_logs(live_rows, trail_log_files)
        live_rows = filter_live_rows_for_request(live_rows, args.tf)
        sim_rows = _sim_live_rows(report_trades)
        progress(f"Comparing live={len(live_rows)} vs backtest={len(sim_rows)} filled order(s)...")
        compare = compare_live_vs_backtest(
            live_rows,
            sim_rows,
            time_tolerance_min=args.match_minutes,
            entry_tolerance=args.match_entry_points,
            pnl_tolerance=args.pnl_tolerance,
            max_match_quality=args.max_match_quality,
            prefer_same_s14_family=args.prefer_same_s14_family,
        )
        raw_replay_context = getattr(args, "_raw_replay_context_trades", None) or all_trades
        enrich_compare_with_raw_replay_context(compare, raw_replay_context)
        print_compare_report(compare)
        if args.compare_csv is not None:
            csv_path = resolve_compare_output_path(args.compare_csv, default_report_name, ".csv", strategies)
            progress(f"Writing compare CSV: {csv_path}")
            written_csv = write_compare_csv(csv_path, compare)
            if written_csv != csv_path:
                print(f"\nCompare CSV: {written_csv} (requested file was locked)")
            else:
                print(f"\nCompare CSV: {written_csv}")
            written_summary_csv = write_compare_summary_csv(written_csv, compare)
            print(f"Compare Summary CSV: {written_summary_csv}")
        if args.compare_xlsx is not None:
            xlsx_path = resolve_compare_output_path(args.compare_xlsx, default_report_name, ".xlsx", strategies)
            progress(f"Writing compare XLSX: {xlsx_path}")
            written_xlsx = write_compare_xlsx(
                xlsx_path,
                compare,
                meta={
                    "symbol": config.SYMBOL,
                    "window": f"{start_bkk} -> {end_bkk}",
                    "strategies": ",".join(str(s) for s in strategies),
                    "context_strategies": ",".join(str(s) for s in context_strategies),
                    "tf_filter": (args.tf or "ALL").upper(),
                    "match_minutes": args.match_minutes,
                    "match_entry_points": args.match_entry_points,
                    "max_match_quality": args.max_match_quality,
                    "prefer_same_s14_family": args.prefer_same_s14_family,
                    "hybrid_live_guard_context": args.hybrid_live_guard_context,
                    "s14_disable_rsi_div": args.s14_disable_rsi_div,
                    "s14_enable_sweep_return": args.s14_enable_sweep_return,
                    "s14_fill_next_bar": args.s14_fill_next_bar,
                    "s2_include_connected_fvg_context": args.s2_include_connected_fvg_context,
                    "s2_parallel_lifecycle_tf": args.s2_parallel_lifecycle_tf,
                    "s2_fill_before_cancel_bars": args.s2_fill_before_cancel_bars,
                    "s3_disable_pd_fibo_plus": args.s3_disable_pd_fibo_plus,
                    "scale_out_column_volume": scale_out_column_volume(config.SYMBOL),
                },
            )
            if written_xlsx != xlsx_path:
                print(f"\nCompare XLSX: {written_xlsx} (requested file was locked)")
            else:
                print(f"\nCompare XLSX: {written_xlsx}")

    mt5.shutdown()
    progress("Done.")


if __name__ == "__main__":
    main()
