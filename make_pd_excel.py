"""สร้าง Excel report สำหรับ sim_pd_multitf"""
import csv, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import sys
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

# อ่าน CSV
rows = []
with open('excel_reports/sim_pd_multitf.csv', encoding='utf-8-sig') as f:
    reader = csv.DictReader(f)
    for r in reader:
        rows.append(r)

# Sort by diff descending
rows.sort(key=lambda x: float(x['diff'] or 0), reverse=True)

wb = openpyxl.Workbook()

H_FILL = PatternFill('solid', start_color='1F4E79')
H_FONT = Font(bold=True, color='FFFFFF', name='Arial', size=10)
T_FONT = Font(name='Arial', size=10)
B_FONT = Font(bold=True, name='Arial', size=10)
G_FILL = PatternFill('solid', start_color='C6EFCE')
R_FILL = PatternFill('solid', start_color='FFC7CE')

# ── Sheet 1: Summary ──
ws1 = wb.active
ws1.title = 'Summary'

ws1['A1'] = 'Sim: PD Zone Fix — Multi-TF Orders'
ws1['A1'].font = Font(bold=True, name='Arial', size=13, color='1F4E79')
ws1.merge_cells('A1:D1')

ws1['A3'] = 'Fix:'
ws1['B3'] = 'เลือก TF เล็กสุดจาก multi-TF (เช่น [M15_M30] -> M15) แทน skip'
ws1['A3'].font = B_FONT
ws1['B3'].font = Font(name='Arial', size=10, italic=True)
ws1.merge_cells('B3:D3')

# Summary stats
stats = [
    ('Orders checked (closed)',   '194'),
    ('PD PASS — entry ถูก zone',  '121'),
    ('PD FAIL — entry ผิด zone',   '73'),
    ('DIFF รวม (USD)',            '+197.52'),
]
for row_i, hdr in enumerate(['Label', 'Value'], 1):
    c = ws1.cell(5, row_i, hdr)
    c.font = H_FONT; c.fill = H_FILL; c.alignment = Alignment(horizontal='center')
for row_i, (label, val) in enumerate(stats, 6):
    ws1.cell(row_i, 1, label).font = B_FONT
    c = ws1.cell(row_i, 2, val)
    c.font = Font(bold=True, name='Arial', size=10,
                  color='00AA00' if str(val).startswith('+') else '000000')
    c.alignment = Alignment(horizontal='center')

# By Strategy
by_sid_data = [
    ('S3',  47, 25, '+110.10'),
    ('S2',  56, 31,  '+38.22'),
    ('S10',  2,  1,  '+28.96'),
    ('S11', 30,  5,  '+16.32'),
    ('S1',  51, 10,   '+4.24'),
    ('S14',  6,  0,   '+0.00'),
    ('S9',   1,  0,   '+0.00'),
    ('S4',   1,  1,   '-0.32'),
]
headers = ['Strategy', 'Orders', 'PD FAIL', 'DIFF (USD)']
row = 11
for col_i, h in enumerate(headers, 1):
    c = ws1.cell(row, col_i, h)
    c.font = H_FONT; c.fill = H_FILL; c.alignment = Alignment(horizontal='center')
for sid, cnt, fail, diff in by_sid_data:
    row += 1
    fill = G_FILL if not diff.startswith('-') else R_FILL
    for col_i, val in enumerate([sid, cnt, fail, diff], 1):
        c = ws1.cell(row, col_i, val)
        c.font = B_FONT if col_i == 1 else T_FONT
        c.fill = fill
        c.alignment = Alignment(horizontal='center' if col_i > 1 else 'left')

# Total row
row += 1
ws1.cell(row, 1, 'TOTAL').font = B_FONT
c_diff = ws1.cell(row, 4, '+197.52')
c_diff.font = Font(bold=True, name='Arial', size=10, color='00AA00')
c_diff.alignment = Alignment(horizontal='center')

ws1.column_dimensions['A'].width = 30
ws1.column_dimensions['B'].width = 18
ws1.column_dimensions['C'].width = 12
ws1.column_dimensions['D'].width = 14

# ── Sheet 2: Detail ──
ws2 = wb.create_sheet('Detail')
COLS = ['ticket','fill_ts','side','raw_tf','tf_used','sid','fill_price',
        'eq','pd_result','actual_pl','sim_pl','diff','close_ts']

for col_i, h in enumerate(COLS, 1):
    c = ws2.cell(1, col_i, h.upper())
    c.font = H_FONT; c.fill = H_FILL
    c.alignment = Alignment(horizontal='center')

for row_i, row_data in enumerate(rows, 2):
    pd_res = row_data.get('pd_result', '')
    fill   = G_FILL if pd_res == 'PASS' else (R_FILL if pd_res == 'FAIL' else PatternFill())
    for col_i, col in enumerate(COLS, 1):
        raw = row_data.get(col, '')
        # แปลงตัวเลข
        try:
            if col in ('fill_price','eq','actual_pl','sim_pl','diff'):
                val = float(raw) if raw != '' else ''
            elif col == 'sid':
                val = int(raw) if raw != '' else ''
            else:
                val = raw
        except (ValueError, TypeError):
            val = raw
        c = ws2.cell(row_i, col_i, val)
        c.font = T_FONT; c.fill = fill
        if col in ('actual_pl','sim_pl','diff') and isinstance(val, float):
            c.number_format = '+#,##0.00;[Red]-#,##0.00'

# Total row
total_row = len(rows) + 2
ws2.cell(total_row, 1, 'TOTAL').font = B_FONT
for col_name in ('actual_pl', 'sim_pl', 'diff'):
    col_i = COLS.index(col_name) + 1
    total = sum(float(r[col_name] or 0) for r in rows)
    c = ws2.cell(total_row, col_i, total)
    c.font = Font(bold=True, name='Arial',
                  color='00AA00' if col_name == 'diff' and total >= 0 else '000000')
    c.number_format = '+#,##0.00;[Red]-#,##0.00'

widths = [14, 20, 6, 14, 10, 6, 12, 10, 10, 11, 9, 9, 20]
for i, w in enumerate(widths, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = 'A2'

out = 'excel_reports/sim_pd_multitf.xlsx'
wb.save(out)
print(f'saved -> {out}')
