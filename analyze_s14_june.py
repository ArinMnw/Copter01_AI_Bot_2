"""
analyze_s14_june.py — S14 June 2026 Checklist
จำแนก 2 มิติ:
  A. Trend alignment: BUY in BULL / SELL in BEAR = "ตามเทรนด์"
                      BUY in BEAR/SIDEWAY / SELL in BULL/SIDEWAY = "สวนเทรนด์ / กลับตัว"
  B. RSI quality:   RSI < 30 (BUY) หรือ > 70 (SELL) = "RSI oversold/overbought (จุดกลับตัว)"
                    RSI 30-50 (BUY) หรือ 50-70 (SELL) = "RSI neutral (Swing)"
"""
import re, os, sys
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

log_files = [
    'logs/old_logs/bot-2026-06.log',
    'logs/bot.log',
]

def fld(line, key):
    m = re.search(rf'{key}=([^|\s]+)', line)
    return m.group(1) if m else ''

# ── 1. Read fills + closes ──────────────────────────────────────
fills = {}; closes = {}; seen = set()
for path in log_files:
    if not os.path.exists(path): continue
    for line in open(path, encoding='utf-8', errors='replace'):
        m = re.match(r'^\[(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2})\]\s+(\S+)', line)
        if not m: continue
        ts, kind = m.group(1), m.group(2)
        tk = fld(line, 'ticket')
        if not tk: continue
        # June 2026 only, sid=14
        if 'sid=14' not in line: continue
        if not ts.startswith('2026-06'): continue

        if kind == 'ENTRY_FILL' and tk not in fills:
            fills[tk] = {
                'fill_ts':   ts,
                'side':      fld(line, 'side'),
                'tf':        fld(line, 'tf'),
                'price':     fld(line, 'price'),
                'sl':        fld(line, 'sl'),
                'tp':        fld(line, 'tp'),
                'trend':     fld(line, 'trend'),
                'rsi':       fld(line, 'rsi'),
                'rsi2':      fld(line, 'rsi2_state'),
                'pd_h':      fld(line, 'pd_h'),
                'pd_l':      fld(line, 'pd_l'),
                'pd_eq':     fld(line, 'pd_eq'),
                'pattern':   fld(line, 'pattern'),
            }
        elif kind == 'POSITION_CLOSED' and tk not in seen and 'XAUUSD' in line:
            seen.add(tk)
            closes[tk] = {
                'close_ts':    ts,
                'profit':      fld(line, 'profit'),
                'reason':      fld(line, 'reason'),
                'close_price': fld(line, 'close_price'),
            }

# deduplicate (same ticket appears in both log files)
seen_tk = set()
order_rows = []
for tk, fi in sorted(fills.items(), key=lambda x: x[1]['fill_ts']):
    if tk in seen_tk: continue
    seen_tk.add(tk)
    cl = closes.get(tk, {})
    side   = fi['side']
    trend  = fi['trend'].lower()
    rsi_v  = float(fi['rsi']) if fi['rsi'] else 50.0
    price  = float(fi['price']) if fi['price'] else 0
    pd_eq  = float(fi['pd_eq']) if fi['pd_eq'] else 0
    profit = float(cl.get('profit', 0) or 0) if cl else None
    reason = cl.get('reason', '') if cl else 'OPEN'

    # ── Trend alignment ──────────────────────────────────────────
    bull = 'bull' in trend
    bear = 'bear' in trend
    sideway = 'sideway' in trend or (not bull and not bear)
    strong = 'strong' in trend

    if side == 'BUY':
        if bull:   trend_align = 'ตามเทรนด์ ✅'
        elif bear: trend_align = 'สวนเทรนด์ ❌'
        else:      trend_align = 'SIDEWAY ⚠️'
    else:  # SELL
        if bear:   trend_align = 'ตามเทรนด์ ✅'
        elif bull: trend_align = 'สวนเทรนด์ ❌'
        else:      trend_align = 'SIDEWAY ⚠️'

    # ── RSI quality ──────────────────────────────────────────────
    if side == 'BUY':
        if rsi_v < 30:   rsi_type = 'Oversold <30 🔥จุดกลับตัว'
        elif rsi_v < 40: rsi_type = 'Low 30-40 🟡'
        elif rsi_v < 50: rsi_type = 'Neutral 40-50 ⚪Swing'
        else:            rsi_type = 'BUY RSI>50 ❌สวน'
    else:
        if rsi_v > 70:   rsi_type = 'Overbought >70 🔥จุดกลับตัว'
        elif rsi_v > 60: rsi_type = 'High 60-70 🟡'
        elif rsi_v > 50: rsi_type = 'Neutral 50-60 ⚪Swing'
        else:            rsi_type = 'SELL RSI<50 ❌สวน'

    # ── PD zone ──────────────────────────────────────────────────
    if pd_eq > 0 and price > 0:
        if side == 'BUY':
            pd_zone = 'Discount ✅' if price < pd_eq else 'Premium ❌'
        else:
            pd_zone = 'Premium ✅' if price > pd_eq else 'Discount ❌'
    else:
        pd_zone = '-'

    # ── RR calculation ───────────────────────────────────────────
    sl_v = float(fi['sl']) if fi['sl'] else 0
    tp_v = float(fi['tp']) if fi['tp'] else 0
    if sl_v > 0 and tp_v > 0 and price > 0:
        risk   = abs(price - sl_v)
        reward = abs(tp_v - price)
        rr     = round(reward / risk, 2) if risk > 0 else 0
    else:
        rr = 0

    # ── Outcome ──────────────────────────────────────────────────
    if not cl:
        outcome = '🔄 OPEN'
    elif '[tp' in reason.lower():
        outcome = '🎯 TP ✅'
    elif '[sl' in reason.lower():
        outcome = '❌ SL'
    elif 'sl guard' in reason.lower():
        outcome = '🛡️ SL Guard'
    elif 'strong-count' in reason.lower():
        outcome = '⚡ S14-cnt'
    elif 'pd zone' in reason.lower():
        outcome = '🚫 PD Zone'
    elif 'fill trend' in reason.lower():
        outcome = '↩️ TrendRecheck'
    else:
        outcome = reason[:15]

    # ── Type classification (Sweep type) ─────────────────────────
    if 'ตามเทรนด์' in trend_align:
        if 'Oversold' in rsi_type or 'Overbought' in rsi_type:
            sweep_type = 'A: Sweep จุดกลับตัว (ตามเทรนด์)'
        else:
            sweep_type = 'B: Sweep Swing (ตามเทรนด์)'
    elif 'สวนเทรนด์' in trend_align:
        if 'Oversold' in rsi_type or 'Overbought' in rsi_type:
            sweep_type = 'C: Sweep กลับตัว (สวนเทรนด์)'
        else:
            sweep_type = 'D: Sweep Counter (สวนเทรนด์+RSI neutral)'
    else:
        sweep_type = 'E: Sweep SIDEWAY'

    order_rows.append({
        'no':         len(order_rows) + 1,
        'ticket':     tk,
        'fill_ts':    fi['fill_ts'],
        'close_ts':   cl.get('close_ts', '') if cl else 'OPEN',
        'side':       side,
        'tf':         fi['tf'],
        'trend':      fi['trend'],
        'trend_align': trend_align,
        'rsi':        round(rsi_v, 1),
        'rsi_type':   rsi_type,
        'rsi2':       fi['rsi2'],
        'pd_zone':    pd_zone,
        'price':      price,
        'sl':         fi['sl'],
        'tp':         fi['tp'],
        'rr':         rr,
        'profit':     profit,
        'outcome':    outcome,
        'sweep_type': sweep_type,
        'strong':     '⚡STRONG' if strong else '',
    })

print(f"S14 June 2026 orders: {len(order_rows)}")

# ── 2. Summary by sweep_type ─────────────────────────────────────
print("\n=== Summary by Sweep Type ===")
by_type = defaultdict(lambda: {'n':0,'wins':0,'pl':0.0,'open':0})
for r in order_rows:
    t = r['sweep_type']
    p = r['profit']
    by_type[t]['n'] += 1
    if p is None:
        by_type[t]['open'] += 1
    else:
        by_type[t]['pl'] += p
        if p > 0: by_type[t]['wins'] += 1

for t, d in sorted(by_type.items()):
    n = d['n'] - d['open']
    wr = round(d['wins']/n*100) if n > 0 else 0
    print(f"  {t}")
    print(f"    Orders={d['n']} | Closed={n} | WR={wr}% | P/L={d['pl']:+.2f}")

print("\n=== Checklist (ทีละ order) ===")
for r in order_rows:
    p_str = f"{r['profit']:+.2f}" if r['profit'] is not None else "OPEN"
    print(f"  #{r['no']:2d} {r['fill_ts']} | {r['side']:4} {r['tf']:4} | {r['trend']:14} | RSI={r['rsi']:5.1f} | {r['outcome']:16} | P/L={p_str:8} | {r['sweep_type']}")

# ── 3. Excel ─────────────────────────────────────────────────────
wb = openpyxl.Workbook()

H_FILL  = PatternFill('solid', start_color='1F4E79')
H_FONT  = Font(bold=True, color='FFFFFF', name='Arial', size=10)
T_FONT  = Font(name='Arial', size=10)
B_FONT  = Font(bold=True, name='Arial', size=10)
G_FILL  = PatternFill('solid', start_color='C6EFCE')  # green
R_FILL  = PatternFill('solid', start_color='FFC7CE')  # red
Y_FILL  = PatternFill('solid', start_color='FFEB9C')  # yellow
O_FILL  = PatternFill('solid', start_color='FFD966')  # orange
LB_FILL = PatternFill('solid', start_color='DDEBF7')  # light blue
LP_FILL = PatternFill('solid', start_color='E2EFDA')  # light green

TYPE_FILLS = {
    'A': G_FILL,   # ดีสุด
    'B': LP_FILL,
    'C': Y_FILL,
    'D': R_FILL,
    'E': O_FILL,
}

# ── Sheet 1: Checklist ───────────────────────────────────────────
ws1 = wb.active
ws1.title = 'Checklist'

ws1['A1'] = 'S14 June 2026 — Sweep Type Analysis'
ws1['A1'].font = Font(bold=True, name='Arial', size=13, color='1F4E79')
ws1.merge_cells('A1:T1')

COLS = ['#','Ticket','Fill Time','Close Time','Side','TF','Trend','Trend Align',
        'RSI','RSI Type','RSI2 State','PD Zone','Entry','SL','TP','R:R',
        'Profit','Outcome','Sweep Type','Strong']
for ci, h in enumerate(COLS, 1):
    c = ws1.cell(3, ci, h)
    c.font = H_FONT; c.fill = H_FILL
    c.alignment = Alignment(horizontal='center', wrap_text=True)

for ri, row in enumerate(order_rows, 4):
    t_key = row['sweep_type'][0]  # A/B/C/D/E
    base_fill = TYPE_FILLS.get(t_key, PatternFill())

    vals = [row['no'], row['ticket'], row['fill_ts'], row['close_ts'],
            row['side'], row['tf'], row['trend'], row['trend_align'],
            row['rsi'], row['rsi_type'], row['rsi2'], row['pd_zone'],
            row['price'], row['sl'], row['tp'], row['rr'],
            row['profit'], row['outcome'], row['sweep_type'], row['strong']]
    for ci, val in enumerate(vals, 1):
        c = ws1.cell(ri, ci, val)
        c.font = T_FONT; c.fill = base_fill
        if ci == 17 and isinstance(val, float):  # profit
            c.number_format = '+#,##0.00;[Red]-#,##0.00'
        c.alignment = Alignment(horizontal='center' if ci not in (3,4,9,18,19) else 'left')

# total row
tr = len(order_rows) + 4
ws1.cell(tr, 1, 'TOTAL').font = B_FONT
total_pl = sum(r['profit'] for r in order_rows if r['profit'] is not None)
c = ws1.cell(tr, 17, total_pl)
c.font = Font(bold=True, name='Arial', size=10, color='00AA00' if total_pl >= 0 else 'CC0000')
c.number_format = '+#,##0.00;[Red]-#,##0.00'; c.fill = Y_FILL

# widths
ws1.freeze_panes = 'A4'
ws1.row_dimensions[3].height = 30
widths = [4,14,20,20,6,6,16,20,7,26,12,14,10,10,10,7,10,18,34,10]
for i,w in enumerate(widths,1):
    ws1.column_dimensions[get_column_letter(i)].width = w

# ── Sheet 2: Summary by Type ─────────────────────────────────────
ws2 = wb.create_sheet('Summary')

ws2['A1'] = 'S14 June — สรุป Sweep Type'
ws2['A1'].font = Font(bold=True, name='Arial', size=13, color='1F4E79')
ws2.merge_cells('A1:G1')

# Type legend
legend = [
    ('A','Sweep จุดกลับตัว (ตามเทรนด์)','BUY in BULL + RSI<30 / SELL in BEAR + RSI>70'),
    ('B','Sweep Swing (ตามเทรนด์)','BUY in BULL RSI 30-70 / SELL in BEAR RSI 30-70'),
    ('C','Sweep กลับตัว (สวนเทรนด์)','BUY in BEAR + RSI<30 / SELL in BULL + RSI>70'),
    ('D','Sweep Counter (สวนเทรนด์+RSI neutral)','BUY in BEAR RSI 30-70 / SELL in BULL RSI 30-70'),
    ('E','Sweep SIDEWAY','TF อยู่ใน SIDEWAY trend'),
]
for ri, (lbl, name, desc) in enumerate(legend, 3):
    ws2.cell(ri, 1, lbl).font = B_FONT
    ws2.cell(ri, 2, name).font = T_FONT
    ws2.cell(ri, 3, desc).font = Font(name='Arial', size=9, italic=True, color='595959')
    for ci in [1,2,3]:
        ws2.cell(ri, ci).fill = TYPE_FILLS.get(lbl, PatternFill())

# Summary stats table
headers_s = ['Type','Name','Orders','Wins','Win%','Closed P/L','Avg per trade']
for ci, h in enumerate(headers_s, 1):
    c = ws2.cell(9, ci, h)
    c.font = H_FONT; c.fill = H_FILL; c.alignment = Alignment(horizontal='center')

ri = 10
for t_key, d in sorted(by_type.items()):
    n_closed = d['n'] - d['open']
    wr = round(d['wins']/n_closed*100, 1) if n_closed > 0 else 0
    avg = round(d['pl']/n_closed, 2) if n_closed > 0 else 0
    type_letter = t_key[0]
    row_vals = [type_letter, t_key[3:], d['n'], d['wins'],
                f"{wr}%", round(d['pl'],2), avg]
    fill = TYPE_FILLS.get(type_letter, PatternFill())
    for ci, val in enumerate(row_vals, 1):
        c = ws2.cell(ri, ci, val)
        c.font = B_FONT if ci == 1 else T_FONT
        c.fill = fill
        c.alignment = Alignment(horizontal='center')
        if ci in [6,7] and isinstance(val, float):
            c.number_format = '+#,##0.00;[Red]-#,##0.00'
    ri += 1

ws2.column_dimensions['A'].width = 8
ws2.column_dimensions['B'].width = 35
ws2.column_dimensions['C'].width = 8
ws2.column_dimensions['D'].width = 8
ws2.column_dimensions['E'].width = 8
ws2.column_dimensions['F'].width = 12
ws2.column_dimensions['G'].width = 14

out = 'excel_reports/s14_june_checklist.xlsx'
wb.save(out)
print(f"\nSaved -> {out}")
