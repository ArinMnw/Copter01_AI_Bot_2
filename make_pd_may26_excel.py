"""สร้าง Excel report สำหรับ sim_pd_may26"""
import csv, openpyxl, os
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import sys
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

rows = []
with open('excel_reports/sim_pd_may26.csv', encoding='utf-8-sig') as f:
    reader = csv.DictReader(f)
    for r in reader:
        rows.append(r)

rows.sort(key=lambda x: float(x['diff'] or 0), reverse=True)

wb = openpyxl.Workbook()

H_FILL = PatternFill('solid', start_color='1F4E79')
H_FONT = Font(bold=True, color='FFFFFF', name='Arial', size=10)
T_FONT = Font(name='Arial', size=10)
B_FONT = Font(bold=True, name='Arial', size=10)
G_FILL = PatternFill('solid', start_color='C6EFCE')
R_FILL = PatternFill('solid', start_color='FFC7CE')
Y_FILL = PatternFill('solid', start_color='FFEB9C')

# ── Sheet 1: Summary ──
ws1 = wb.active
ws1.title = 'Summary'

ws1['A1'] = 'Sim: PD Zone Fix — orders ตั้งแต่ 2026-05-26'
ws1['A1'].font = Font(bold=True, name='Arial', size=13, color='1F4E79')
ws1.merge_cells('A1:E1')

ws1['A3'] = 'Fix:'
ws1['B3'] = 'เลือก TF เล็กสุดจาก multi-TF (เช่น [M15_M30] → M15) แทน skip_no_data forever'
ws1['A3'].font = B_FONT
ws1['B3'].font = Font(name='Arial', size=10, italic=True)
ws1.merge_cells('B3:E3')

# Overall stats
stats = [
    ('Orders analyzed (closed)',   '193'),
    ('PD PASS — entry ถูก zone',   '121'),
    ('PD FAIL — entry ผิด zone',    '72'),
    ('Old P/L รวม (USD)',          '-446.91'),
    ('New P/L รวม (USD)',          '-247.19'),
    ('DIFF รวม (USD)',             '+199.72'),
]
for col_i, hdr in enumerate(['Label', 'Value'], 1):
    c = ws1.cell(5, col_i, hdr)
    c.font = H_FONT; c.fill = H_FILL; c.alignment = Alignment(horizontal='center')

for row_i, (label, val) in enumerate(stats, 6):
    ws1.cell(row_i, 1, label).font = B_FONT
    c = ws1.cell(row_i, 2, val)
    is_pos = str(val).startswith('+')
    is_neg = str(val).startswith('-')
    clr = '00AA00' if is_pos else ('CC0000' if is_neg else '000000')
    c.font = Font(bold=True, name='Arial', size=10, color=clr)
    if row_i == 11:  # DIFF row
        c.fill = G_FILL
    c.alignment = Alignment(horizontal='center')

# By Strategy table
by_sid_data = [
    ('S3',   47, 25, '-143.48',  '-33.38', '+110.10'),
    ('S2',   55, 30, '-121.84',  '-81.42',  '+40.42'),
    ('S10',   2,  1,  '-36.96',   '-8.00',  '+28.96'),
    ('S11',  30,  5,  '+80.13',  '+96.45',  '+16.32'),
    ('S1',   51, 10, '-254.62', '-250.38',   '+4.24'),
    ('S9',    1,  0,   '+5.68',   '+5.68',   '+0.00'),
    ('S14',   6,  0,  '+24.06',  '+24.06',   '+0.00'),
    ('S4',    1,  1,   '+0.12',   '-0.20',   '-0.32'),
]
headers = ['Strategy', 'Orders', 'PD FAIL', 'Old P/L', 'New P/L', 'DIFF']
row = 13
for col_i, h in enumerate(headers, 1):
    c = ws1.cell(row, col_i, h)
    c.font = H_FONT; c.fill = H_FILL; c.alignment = Alignment(horizontal='center')

for sid, cnt, fail, old, new, diff in by_sid_data:
    row += 1
    fill = G_FILL if not diff.startswith('-') else R_FILL
    for col_i, val in enumerate([sid, cnt, fail, old, new, diff], 1):
        c = ws1.cell(row, col_i, val)
        c.font = B_FONT if col_i == 1 else T_FONT
        c.fill = fill
        c.alignment = Alignment(horizontal='center' if col_i > 1 else 'left')

# Total row
row += 1
ws1.cell(row, 1, 'TOTAL').font = B_FONT
for col_i, (val, clr) in enumerate(
    [('193', '000000'), ('72', '000000'),
     ('-446.91', 'CC0000'), ('-247.19', 'CC0000'), ('+199.72', '00AA00')], 2):
    c = ws1.cell(row, col_i, val)
    c.font = Font(bold=True, name='Arial', size=10, color=clr)
    c.alignment = Alignment(horizontal='center')
    c.fill = Y_FILL

ws1.column_dimensions['A'].width = 28
ws1.column_dimensions['B'].width = 12
ws1.column_dimensions['C'].width = 12
ws1.column_dimensions['D'].width = 13
ws1.column_dimensions['E'].width = 13
ws1.column_dimensions['F'].width = 13

# ── Sheet 2: Detail ──
ws2 = wb.create_sheet('Detail')
COLS = ['ticket','fill_ts','side','raw_tf','tf_used','sid','is_multitf',
        'fill_price','eq','pd_result','actual_pl','sim_pl','diff','close_ts']

for col_i, h in enumerate(COLS, 1):
    c = ws2.cell(1, col_i, h.upper())
    c.font = H_FONT; c.fill = H_FILL
    c.alignment = Alignment(horizontal='center')

for row_i, row_data in enumerate(rows, 2):
    pd_res = row_data.get('pd_result', '')
    fill   = G_FILL if pd_res == 'PASS' else (R_FILL if pd_res == 'FAIL' else PatternFill())
    for col_i, col in enumerate(COLS, 1):
        raw = row_data.get(col, '')
        try:
            if col in ('fill_price','eq','actual_pl','sim_pl','diff'):
                val = float(raw) if raw != '' else ''
            elif col == 'sid':
                val = int(raw) if raw != '' else ''
            elif col == 'is_multitf':
                val = raw  # True/False string
            else:
                val = raw
        except (ValueError, TypeError):
            val = raw
        c = ws2.cell(row_i, col_i, val)
        c.font = T_FONT; c.fill = fill
        if col in ('actual_pl','sim_pl','diff') and isinstance(val, float):
            c.number_format = '+#,##0.00;[Red]-#,##0.00'

# Total row
total_row = len(rows) + 2
ws2.cell(total_row, 1, 'TOTAL').font = B_FONT
for col_name in ('actual_pl', 'sim_pl', 'diff'):
    col_i = COLS.index(col_name) + 1
    total = sum(float(r[col_name] or 0) for r in rows)
    c = ws2.cell(total_row, col_i, total)
    c.font = Font(bold=True, name='Arial', size=10,
                  color='00AA00' if total >= 0 else 'CC0000')
    c.number_format = '+#,##0.00;[Red]-#,##0.00'
    c.fill = Y_FILL

widths = [14,20,6,14,10,6,10,12,10,10,11,9,9,20]
for i, w in enumerate(widths, 1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = 'A2'

out = 'excel_reports/sim_pd_may26.xlsx'
wb.save(out)
print(f'saved -> {out}')
