"""
export_s14_compare_xlsx.py
รัน S14 Old vs New backtest แล้ว export เป็น Excel
"""
import sys, os
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
sys.stdout.reconfigure(encoding='utf-8')

import compare_s14_sweep_old_new as cmp
import MetaTrader5 as mt5
from datetime import datetime, timezone, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT_FILE = r"D:\Project\Copter01_AI_Bot_2\excel_reports\s14_compare_old_new.xlsx"

C_HEADER_OLD = "1F4E79"
C_HEADER_NEW = "375623"
C_HEADER_SUM = "404040"
C_HEADER_CMP = "7030A0"
C_TP         = "E2EFDA"
C_SL         = "FCE4D6"
C_FLIP       = "FFF2CC"
C_OPEN       = "EDEDED"
C_WHITE      = "FFFFFF"
C_ONLY_OLD   = "D6E4F0"  # trade มีแค่ OLD
C_ONLY_NEW   = "D9EAD3"  # trade มีแค่ NEW

def _font(bold=False, color="000000", size=10):
    return Font(name="Arial", bold=bold, color=color, size=size)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _border():
    s = Side(border_style="thin", color="BFBFBF")
    return Border(left=s, right=s, top=s, bottom=s)

def _center():
    return Alignment(horizontal="center", vertical="center")

def _right():
    return Alignment(horizontal="right", vertical="center")

def _pnl_color(pnl):
    if pnl is None: return "666666"
    return "375623" if float(pnl) > 0 else ("C00000" if float(pnl) < 0 else "000000")

def _row_bg(close_type):
    ct = str(close_type or "").upper()
    if ct == "TP":   return C_TP
    if ct == "SL":   return C_SL
    if ct in ("FLIP","S14_FLIP"): return C_FLIP
    return C_OPEN

def _header_row(ws, row, values, bg, fg="FFFFFF"):
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.font = _font(bold=True, color=fg)
        c.fill = _fill(bg)
        c.alignment = _center()
        c.border = _border()


# ── fetch real order tickets from MT5 ─────────────────────────────────
def build_ticket_map(since_dt):
    """
    คืน dict: key=(tf, signal, bkk_minute) → order_ticket
    bkk_minute = datetime truncated to minute (BKK time displayed as UTC)
    """
    UTC = timezone.utc
    SRV_TZ = getattr(__import__('config'), 'MT5_SERVER_TZ', 1)
    TZ_OFF  = getattr(__import__('config'), 'TZ_OFFSET',    7)
    offset  = timedelta(hours=TZ_OFF - SRV_TZ)   # +6h

    deals = mt5.history_deals_get(since_dt, datetime.now(UTC))
    if not deals:
        return {}

    tmap = {}
    for d in deals:
        if 'S14' not in (d.comment or ''):
            continue
        if d.entry != 0:          # เอาเฉพาะ entry deal
            continue
        comment = d.comment or ""
        # parse TF from comment: "M5_S14_BRS" → "M5"
        parts = comment.split("_")
        tf = parts[0] if parts else ""
        sig = "BUY" if d.type == 0 else "SELL"
        # convert to BKK displayed as UTC
        bkk = (datetime.fromtimestamp(d.time, UTC) + offset).replace(second=0, microsecond=0)
        key = (tf, sig, bkk)
        tmap[key] = d.order   # order ticket
    return tmap


def lookup_ticket(tmap, tf, signal, entry_time):
    """
    หา ticket จาก tmap โดยเช็คใน window ±2 นาที
    """
    if entry_time is None:
        return None
    et = entry_time.replace(second=0, microsecond=0)
    for delta in (0, 1, -1, 2, -2):
        key = (tf, signal, et + timedelta(minutes=delta))
        ticket = tmap.get(key)
        if ticket:
            return ticket
    return None


# ── match OLD↔NEW trades ───────────────────────────────────────────────
def match_trades(all_old, all_new):
    """
    คืน list ของ dict:
      {tf, signal, sub_pattern, time_label,
       old_trade (or None), new_trade (or None)}
    """
    # non-sweep: match by (tf, signal, entry_time_raw, sub_pattern)
    # sweep:     match by (tf, signal, sweep_bar_time)
    rows = []
    used_old = set()
    used_new = set()

    # index new trades
    new_by_nonsweep = {}   # (tf, signal, entry_time_raw, sub_pattern) -> idx
    new_by_sweep    = {}   # (tf, signal, sweep_bar_time) -> idx

    for idx, t in enumerate(all_new):
        sp = t.get("sub_pattern", "")
        tf = t.get("tf", "")
        sig = t.get("signal", "")
        if sp == "sweep":
            k = (tf, sig, t.get("sweep_bar_time"))
            new_by_sweep[k] = idx
        else:
            k = (tf, sig, t.get("entry_time_raw"), sp)
            new_by_nonsweep[k] = idx

    for o_idx, ot in enumerate(all_old):
        sp  = ot.get("sub_pattern", "")
        tf  = ot.get("tf", "")
        sig = ot.get("signal", "")
        et  = ot.get("entry_time")
        matched_new = None

        if sp == "sweep":
            k = (tf, sig, ot.get("sweep_bar_time"))
            n_idx = new_by_sweep.get(k)
            if n_idx is not None and n_idx not in used_new:
                matched_new = all_new[n_idx]
                used_new.add(n_idx)
        else:
            k = (tf, sig, ot.get("entry_time_raw"), sp)
            n_idx = new_by_nonsweep.get(k)
            if n_idx is not None and n_idx not in used_new:
                matched_new = all_new[n_idx]
                used_new.add(n_idx)

        used_old.add(o_idx)
        rows.append({
            "tf": tf, "signal": sig, "sub_pattern": sp,
            "sort_key": ot.get("entry_time") or cmp.to_bkk(0),
            "old": ot, "new": matched_new,
        })

    # new trades ที่ไม่มี pair
    for n_idx, nt in enumerate(all_new):
        if n_idx not in used_new:
            rows.append({
                "tf": nt.get("tf",""), "signal": nt.get("signal",""),
                "sub_pattern": nt.get("sub_pattern",""),
                "sort_key": nt.get("entry_time") or cmp.to_bkk(0),
                "old": None, "new": nt,
            })

    rows.sort(key=lambda r: (r["sort_key"], r["tf"]))
    return rows


def write_compare_sheet(wb, matched_rows, tmap):
    ws = wb.create_sheet("Compare")

    # header row 1 — group labels
    ws.merge_cells("A1:F1")
    ws["A1"].value = "Trade Info"
    ws["A1"].font  = _font(bold=True, color="FFFFFF")
    ws["A1"].fill  = _fill(C_HEADER_SUM)
    ws["A1"].alignment = _center()

    ws.merge_cells("G1:K1")
    ws["G1"].value = "OLD (sweep bar = rates[-1], no confirm)"
    ws["G1"].font  = _font(bold=True, color="FFFFFF")
    ws["G1"].fill  = _fill(C_HEADER_OLD)
    ws["G1"].alignment = _center()

    ws.merge_cells("L1:P1")
    ws["L1"].value = "NEW (sweep bar = rates[-2], confirm bar required)"
    ws["L1"].font  = _font(bold=True, color="FFFFFF")
    ws["L1"].fill  = _fill(C_HEADER_NEW)
    ws["L1"].alignment = _center()

    # header row 2 (Q = DIFF)
    headers = [
        "TF","Signal","Pattern","Time OLD","Time NEW","Order#",
        "Entry","SL","TP","Close Type","P&L",
        "Entry","SL","TP","Close Type","P&L",
        "DIFF P&L",
    ]
    _header_row(ws, 2, headers, C_HEADER_SUM)
    # override Q col (17) header colour
    c17 = ws.cell(row=1, column=17, value="DIFF P&L")
    c17.font = _font(bold=True, color="FFFFFF")
    c17.fill = _fill(C_HEADER_CMP)
    c17.alignment = _center()
    c17.border = _border()
    ws.cell(row=2, column=17).fill = _fill(C_HEADER_CMP)
    ws.cell(row=2, column=17).font = _font(bold=True, color="FFFFFF")

    for r_idx, row in enumerate(matched_rows, 3):
        ot = row["old"]
        nt = row["new"]
        tf  = row["tf"]
        sig = row["signal"]
        sp  = row["sub_pattern"] or ""

        bg = _row_bg(ot.get("close_type")) if ot and nt else (C_ONLY_OLD if ot else C_ONLY_NEW)

        def _et(t):
            et = t.get("entry_time") if t else None
            return et.strftime("%d-%m %H:%M") if et else "-"

        # lookup ticket: try OLD entry_time first, then NEW
        ref_t = ot or nt
        ticket = lookup_ticket(tmap, tf, sig, ref_t.get("entry_time") if ref_t else None) if ref_t else None

        old_pnl = float(ot.get("pnl", 0)) if ot else None
        new_pnl = float(nt.get("pnl", 0)) if nt else None
        diff = round((new_pnl or 0) - (old_pnl or 0), 2)

        def _v(t, key, default="-"):
            return float(t[key]) if t and t.get(key) is not None else default

        vals = [
            tf, sig, sp[:20], _et(ot), _et(nt), ticket or "-",
            _v(ot,"entry"), _v(ot,"sl"), _v(ot,"tp"), ot.get("close_type","-") if ot else "-", old_pnl,
            _v(nt,"entry"), _v(nt,"sl"), _v(nt,"tp"), nt.get("close_type","-") if nt else "-", new_pnl,
            diff,
        ]

        for col, val in enumerate(vals, 1):
            c = ws.cell(row=r_idx, column=col, value=val)
            c.fill   = _fill(bg)
            c.border = _border()
            c.font   = _font(color="000000")

            if col in (7,8,9,12,13,14):
                if isinstance(val, float):
                    c.number_format = '#,##0.00'; c.alignment = _right()
                else:
                    c.alignment = _center()
            elif col in (11, 16):   # P&L
                if isinstance(val, (int, float)):
                    c.number_format = '+#,##0.00;-#,##0.00;"-"'
                    c.font = _font(color=_pnl_color(val))
                    c.alignment = _right()
                else:
                    c.alignment = _center()
            elif col == 17:          # DIFF
                if isinstance(val, (int, float)):
                    c.number_format = '+#,##0.00;-#,##0.00;"-"'
                    c.fill  = _fill(C_TP if val > 0 else (C_SL if val < 0 else C_WHITE))
                    c.font  = _font(bold=True, color=_pnl_color(val))
                    c.alignment = _right()
            elif col == 6:           # Order#
                c.alignment = _center()
            else:
                c.alignment = _center()

    # total row
    nr = len(matched_rows) + 3
    ws.cell(row=nr, column=1, value="TOTAL").font = _font(bold=True)
    for col, ltr in [(11,"K"), (16,"P"), (17,"Q")]:
        c = ws.cell(row=nr, column=col)
        c.value = f"=SUM({ltr}3:{ltr}{nr-1})"
        c.number_format = '+#,##0.00;-#,##0.00;"-"'
        c.font = _font(bold=True); c.fill = _fill("D9D9D9")
        c.border = _border(); c.alignment = _right()

    ws.cell(row=nr+2, column=1,
            value="สี: น้ำเงินอ่อน = มีแต่ OLD | เขียวอ่อน = มีแต่ NEW | Order# = ticket จาก MT5 history").font = _font(color="666666", size=9)

    widths = [5,7,18,14,14,12, 9,9,9,10,10, 9,9,9,10,10, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 16
    ws.freeze_panes = "A3"
    return ws


def write_trades_sheet(wb, sheet_name, trades, header_color, tmap=None):
    ws = wb.create_sheet(sheet_name)
    headers = ["TF","Version","Entry Time","Signal","Pattern","Order#",
               "Entry","SL","TP","Close Price","Close Type","P&L (USD)"]
    _header_row(ws, 1, headers, header_color)

    for r, t in enumerate(trades, 2):
        ct   = str(t.get("close_type",""))
        bg   = _row_bg(ct)
        et   = t.get("entry_time")
        et_s = et.strftime("%d-%m %H:%M") if et else ""
        pnl  = float(t.get("pnl", 0.0) or 0.0)
        ticket = lookup_ticket(tmap, t.get("tf",""), t.get("signal",""), et) if tmap else None
        vals = [
            t.get("tf",""), t.get("_version",""), et_s,
            t.get("signal",""),
            t.get("sub_pattern", t.get("pattern",""))[:30],
            ticket or "-",
            float(t.get("entry",0)), float(t.get("sl",0)),
            float(t.get("tp",0)), float(t.get("close_price",0)),
            ct, pnl,
        ]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.fill  = _fill(bg)
            c.font  = _font(color="000000")
            c.border = _border()
            if col in (7,8,9,10):
                c.number_format = '#,##0.00'; c.alignment = _right()
            elif col == 12:
                c.number_format = '+#,##0.00;-#,##0.00;"-"'
                c.alignment = _right()
                c.font = _font(color=_pnl_color(pnl))
            else:
                c.alignment = _center()

    col_widths = [6,8,14,7,20,12,10,10,10,12,10,12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    n = len(trades) + 2
    ws.cell(row=n, column=1, value="TOTAL").font = _font(bold=True)
    c = ws.cell(row=n, column=12)
    c.value = f"=SUM(L2:L{n-1})"
    c.font  = _font(bold=True); c.fill = _fill("D9D9D9")
    c.border = _border()
    c.number_format = '+#,##0.00;-#,##0.00;"-"'
    c.alignment = _right()

    ws.freeze_panes = "A2"
    return ws


def write_summary_sheet(wb, all_old, all_new):
    ws = wb.create_sheet("Summary", 0)
    TF_ORDER = ["M1","M5","M15","M30","H1","H4","D1"]

    ws.merge_cells("A1:L1")
    t = ws["A1"]
    t.value = "S14 Sweep  OLD vs NEW  |  06-06-2026 ถึงปัจจุบัน  |  0.01 lot"
    t.font  = _font(bold=True, size=12, color="FFFFFF")
    t.fill  = _fill(C_HEADER_SUM)
    t.alignment = _center()

    ws.merge_cells("B2:G2")
    ws["B2"].value = "OLD"
    ws["B2"].font  = _font(bold=True, color="FFFFFF")
    ws["B2"].fill  = _fill(C_HEADER_OLD)
    ws["B2"].alignment = _center()

    ws.merge_cells("H2:L2")
    ws["H2"].value = "NEW"
    ws["H2"].font  = _font(bold=True, color="FFFFFF")
    ws["H2"].fill  = _fill(C_HEADER_NEW)
    ws["H2"].alignment = _center()

    _header_row(ws, 3,
        ["TF","Trades","TP","SL","WR%","P&L","Trades","TP","SL","WR%","P&L","DIFF P&L"],
        C_HEADER_SUM)

    for r_idx, tf in enumerate(TF_ORDER, 4):
        to = [t for t in all_old if t.get("tf") == tf]
        tn = [t for t in all_new if t.get("tf") == tf]
        tp_o = sum(1 for t in to if t.get("close_type") == "TP")
        sl_o = sum(1 for t in to if t.get("close_type") == "SL")
        wr_o = tp_o / (tp_o + sl_o) if (tp_o + sl_o) > 0 else 0
        po   = round(sum(float(t.get("pnl",0)) for t in to), 2)
        tp_n = sum(1 for t in tn if t.get("close_type") == "TP")
        sl_n = sum(1 for t in tn if t.get("close_type") == "SL")
        wr_n = tp_n / (tp_n + sl_n) if (tp_n + sl_n) > 0 else 0
        pn   = round(sum(float(t.get("pnl",0)) for t in tn), 2)
        row  = [tf, len(to), tp_o, sl_o, wr_o, po,
                    len(tn), tp_n, sl_n, wr_n, pn, round(pn-po,2)]
        for col, val in enumerate(row, 1):
            c = ws.cell(row=r_idx, column=col, value=val)
            c.border = _border()
            c.font   = _font(bold=(col==1))
            if col in (5,10):
                c.number_format = "0%"; c.alignment = _right()
            elif col in (6,11):
                c.number_format = '+#,##0.00;-#,##0.00;"-"'
                c.alignment = _right()
                c.font = _font(color=_pnl_color(val))
            elif col == 12:
                c.number_format = '+#,##0.00;-#,##0.00;"-"'
                c.alignment = _right()
                c.fill = _fill(C_TP if isinstance(val,(int,float)) and val>0 else (C_SL if isinstance(val,(int,float)) and val<0 else C_WHITE))
                c.font = _font(bold=True, color=_pnl_color(val))
            else:
                c.alignment = _center()

    tr = len(TF_ORDER) + 4
    ws.cell(row=tr, column=1, value="TOTAL").font = _font(bold=True)
    for col in range(2, 13):
        ltr = get_column_letter(col)
        c = ws.cell(row=tr, column=col)
        if col not in (5, 10):
            c.value = f"=SUM({ltr}4:{ltr}{tr-1})"
            c.number_format = '+#,##0.00;-#,##0.00;"-"' if col in (6,11,12) else "0"
        c.font = _font(bold=True); c.fill = _fill("D9D9D9")
        c.border = _border(); c.alignment = _right()

    widths = [6,8,5,5,6,10,8,5,5,6,10,12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "B4"
    return ws


def main():
    if not mt5.initialize():
        print("MT5 init failed:", mt5.last_error())
        return

    print("กำลัง run backtest...")
    all_old, all_new = [], []
    for tf_name, tf_val in cmp.TF_MAP.items():
        to, tn = cmp.backtest_tf(tf_name, tf_val)
        for t in to: t["tf"] = tf_name; t["_version"] = "OLD"
        for t in tn: t["tf"] = tf_name; t["_version"] = "NEW"
        all_old.extend(to); all_new.extend(tn)
        print(f"  {tf_name}: OLD={len(to)}  NEW={len(tn)}")

    print("จับคู่ OLD↔NEW...")
    matched = match_trades(all_old, all_new)
    paired   = sum(1 for r in matched if r["old"] and r["new"])
    only_old = sum(1 for r in matched if r["old"] and not r["new"])
    only_new = sum(1 for r in matched if not r["old"] and r["new"])
    print(f"  Paired={paired}  Only-OLD={only_old}  Only-NEW={only_new}")

    print("ดึง order tickets จาก MT5...")
    tmap = build_ticket_map(cmp.SINCE)
    print(f"  พบ S14 entry deals: {len(tmap)}")

    print("กำลัง export Excel...")
    wb = Workbook()
    wb.remove(wb.active)

    write_summary_sheet(wb, all_old, all_new)
    write_compare_sheet(wb, matched, tmap)

    write_trades_sheet(wb, "OLD Trades", sorted(all_old, key=lambda t: t.get("entry_time") or cmp.to_bkk(0)), C_HEADER_OLD, tmap)
    write_trades_sheet(wb, "NEW Trades", sorted(all_new, key=lambda t: t.get("entry_time") or cmp.to_bkk(0)), C_HEADER_NEW, tmap)

    os.makedirs(os.path.dirname(OUT_FILE), exist_ok=True)
    wb.save(OUT_FILE)
    print(f"Saved: {OUT_FILE}")

    scripts_dir = r"C:\Users\Copter\AppData\Roaming\Claude\local-agent-mode-sessions\skills-plugin\86fbf773-5886-4002-ad3a-415400fdad12\7c35f092-618e-4354-809e-8409f54548fa\skills\xlsx\scripts"
    recalc = os.path.join(scripts_dir, "recalc.py")
    if os.path.exists(recalc):
        import subprocess
        r = subprocess.run(["python", recalc, OUT_FILE], capture_output=True, text=True, encoding="utf-8")
        print(r.stdout.strip())

    mt5.shutdown()
    total_old = sum(float(t.get("pnl",0)) for t in all_old)
    total_new = sum(float(t.get("pnl",0)) for t in all_new)
    print(f"\nOLD: {total_old:+.2f} USD  NEW: {total_new:+.2f} USD  DIFF: {total_new-total_old:+.2f} USD")

if __name__ == "__main__":
    main()
