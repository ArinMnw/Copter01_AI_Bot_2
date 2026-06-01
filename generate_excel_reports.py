#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""แปลง CSV -> excel_reports/*.xlsx
   - คอลัมน์ท้าย: ปัญหา / ระดับ / STATUS
   - 🔴 ไม่ผ่าน (แดง) = ขาดทุน + ช่องโหว่ "ยังไม่มีกลไก" (weak counter-trend บนท่า bypass / lot)
   - 🟠 เปิด flag ได้ (ส้ม) = strong counter-trend บนท่า bypass → STRONG_TREND_BLOCK กันได้ (รอเปิด)
   - 🟡 ผ่าน (แก้แล้ว) = bug แก้แล้ว (SL guard / trend recheck)
   - ✅ ผ่าน
"""
import os, csv, sys, io
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ROOT = os.path.dirname(os.path.abspath(__file__))
OUTDIR = os.path.join(ROOT, "excel_reports")
os.makedirs(OUTDIR, exist_ok=True)

FONT = "Arial"
RED_FILL    = PatternFill("solid", fgColor="FFC7CE")   # แดง = ไม่ผ่าน (ยังไม่มีกลไก)
ORANGE_FILL = PatternFill("solid", fgColor="FCE4D6")   # ส้ม = มีกลไก รอเปิด flag
FIXED_FILL  = PatternFill("solid", fgColor="FFF2CC")   # เหลือง = bug แก้แล้ว
PASS_FILL   = PatternFill("solid", fgColor="E2EFDA")   # เขียว = ผ่าน
HDR_FILL    = PatternFill("solid", fgColor="305496")
HDR_FONT    = Font(name=FONT, bold=True, color="FFFFFF", size=10)
CELL_FONT   = Font(name=FONT, size=10)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
NUMERIC = {"profit","lot","entry","sl","close","OLD_profit","NEW_profit","diff"}

BYPASS = {"9","10","11","13","14"}        # ท่าที่ข้าม trend filter
RECHECK_SIDS = {"1","2","3","4","5","6","8","12"}  # ท่าที่ได้ trend recheck (หลังแก้)

def f2(v):
    try: return float(v)
    except: return None

def severity(p):
    if p is None or p >= 0: return "-"
    a = abs(p)
    if a >= 30: return "รุนแรง"
    if a >= 10: return "ปานกลาง"
    return "เบา"

def strength_of(tf):
    if "strong" in tf: return "strong"
    if "weak" in tf: return "weak"
    return "sideway"

def health_status(row):
    """คืน (ปัญหา, ระดับ, status, fill)"""
    trend   = row.get("trend","")
    recheck = row.get("recheck","")
    guard   = row.get("sl_guard","")
    lot_ok  = row.get("lot_ok","")
    sid     = row.get("sid","")
    profit  = f2(row.get("profit")) or 0.0
    sev = severity(profit)

    issues = []
    if trend == "COUNTER":      issues.append("สวนเทรนด์")
    if guard == "SHOULD-BLOCK": issues.append("guard ควรบล็อก")
    if recheck == "BLIND":      issues.append("recheck ตาบอด")
    if lot_ok == "N":           issues.append("lot ผิด")

    if not issues:
        return "-", "-", "ผ่าน", PASS_FILL

    # covered โดย bug fix ที่ deploy แล้ว
    fixed_by_deploy = (guard == "SHOULD-BLOCK") or (recheck in ("BLIND","closed","ran"))
    recheck_applies = sid in RECHECK_SIDS   # ท่าได้ recheck หลังแก้ → counter จะถูกจับ

    if lot_ok == "N":
        return " + ".join(issues), sev, "ไม่ผ่าน", RED_FILL

    if "สวนเทรนด์" in issues and not fixed_by_deploy and not recheck_applies and sid in BYPASS:
        # ท่า bypass ที่ recheck/guard ไม่จับ
        if profit < 0:
            if strength_of(row.get("trend_filter","")) == "strong":
                return " + ".join(issues), sev, "เปิด STRONG_TREND_BLOCK", ORANGE_FILL
            return " + ".join(issues), sev, "ไม่ผ่าน", RED_FILL
        return " + ".join(issues), sev, "ผ่าน", PASS_FILL

    # ที่เหลือ: มีปัญหาแต่ fix แล้ว (guard/recheck)
    return " + ".join(issues), sev, "ผ่าน (แก้แล้ว)", FIXED_FILL

def cmp_status(row):
    st = row.get("status","")
    p = f2(row.get("OLD_profit")) or 0.0
    if st in ("AVOIDED","CLOSE_EARLY"):
        return st, severity(p), "ผ่าน (แก้แล้ว)", FIXED_FILL
    return st, severity(p), "ผ่าน", PASS_FILL

def write_sheet(ws, headers, rows_data, status_fn, extra_cols):
    allh = headers + extra_cols
    for c, h in enumerate(allh, 1):
        cell = ws.cell(1, c, h)
        cell.font = HDR_FONT; cell.fill = HDR_FILL; cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.freeze_panes = "A2"
    for ri, row in enumerate(rows_data, 2):
        vals = [row.get(h,"") for h in headers]
        *extras, status, fill = status_fn(row)   # extras = ทุกค่าก่อน status; แต่ status_fn คืน (issues,sev,status,fill)
        # status_fn คืน 4 ค่า: issues, sev, status, fill
        issues, sev = extras[0], extras[1]
        vals += [issues, sev, status]
        for ci, v in enumerate(vals, 1):
            hdr = headers[ci-1] if ci-1 < len(headers) else None
            num = f2(v)
            use_num = (num is not None and hdr in NUMERIC)
            cell = ws.cell(ri, ci, num if use_num else v)
            if use_num and hdr in ("profit","OLD_profit","NEW_profit","diff"):
                cell.number_format = "#,##0.00;(#,##0.00);-"
            cell.font = CELL_FONT; cell.border = BORDER
            cell.alignment = Alignment(horizontal="center" if ci>1 else "left")
        if fill in (RED_FILL, ORANGE_FILL):
            for ci in range(1, len(vals)+1):
                ws.cell(ri, ci).fill = fill
        sc = ws.cell(ri, len(vals))
        col = {"ไม่ผ่าน":"9C0006","เปิด STRONG_TREND_BLOCK":"BF6000"}.get(status)
        if col: sc.font = Font(name=FONT, size=10, bold=True, color=col)
        elif status.startswith("ผ่าน (แก้"): sc.font = Font(name=FONT, size=10, color="9C6500")
        else: sc.font = Font(name=FONT, size=10, color="375623")

def autosize(ws, widths):
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

def safe_save(wb, path):
    try:
        wb.save(path); return path
    except PermissionError:
        alt = path.replace(".xlsx", "_new.xlsx")
        wb.save(alt)
        print(f"  [!] {os.path.basename(path)} เปิดค้างอยู่ -> บันทึกเป็น {os.path.basename(alt)} แทน")
        return alt

def build_health():
    rows = list(csv.DictReader(open(os.path.join(ROOT,"orders_health.csv"), encoding="utf-8-sig")))
    headers = ["close_ts","ticket","side","tf","sid","lot","lot_ok","trend_filter",
               "trend","recheck","pd","tso","sl_guard","profit","reason"]
    wb = Workbook(); ws = wb.active; ws.title = "Health"
    write_sheet(ws, headers, rows, health_status, ["ปัญหา","ระดับ","STATUS"])
    autosize(ws, {"A":19,"B":11,"C":5,"D":6,"E":5,"F":6,"G":7,"H":11,"I":9,"J":9,
                  "K":8,"L":8,"M":13,"N":9,"O":16,"P":22,"Q":9,"R":22})
    # Legend
    lg = wb.create_sheet("คำอธิบาย")
    legend = [
        ("รายงาน Health-Check ราย Order — XAUUSD 26 พ.ค.+", ""),
        ("", ""),
        ("STATUS", "ความหมาย / การจัดการ"),
        ("ไม่ผ่าน (แดง)", "ขาดทุน + ช่องโหว่ที่ยังไม่มีกลไก (weak counter-trend บนท่า bypass / lot ผิด)"),
        ("เปิด STRONG_TREND_BLOCK (ส้ม)", "strong counter-trend บนท่า bypass → กลไกใหม่กันได้ ตั้ง STRONG_TREND_BLOCK_ENABLED=True"),
        ("ผ่าน (แก้แล้ว) (เหลือง)", "มีปัญหาแต่ bug แก้แล้ว → SL Guard / Trend Recheck"),
        ("ผ่าน (เขียว)", "order ปกติ"),
        ("", ""),
        ("ระดับความรุนแรง (เฉพาะที่ขาดทุน)", ""),
        ("รุนแรง", "ขาดทุน ≥ $30"),
        ("ปานกลาง", "ขาดทุน $10–30"),
        ("เบา", "ขาดทุน < $10"),
        ("", ""),
        ("กลไกที่แก้/เพิ่มแล้ว", ""),
        ("SL Guard swing_ref=0", "guard ควรบล็อกแต่ไม่ทำ — แก้ trailing.py (per-TF/Combined/Group)"),
        ("Trend Recheck _swing_data.clear", "recheck ตาบอด (BLIND) — แก้แล้ว"),
        ("STRONG_TREND_BLOCK (ใหม่)", "กัน counter-strong-trend ท่า bypass — net +314 (default OFF)"),
    ]
    for r,(a,b) in enumerate(legend,1):
        lg.cell(r,1,a).font = Font(name=FONT, bold=(r==1), size=11 if r==1 else 10)
        lg.cell(r,2,b).font = Font(name=FONT, size=10); lg.cell(r,2).alignment = Alignment(wrap_text=True)
    lg.column_dimensions["A"].width = 34; lg.column_dimensions["B"].width = 74
    lg["A4"].fill = RED_FILL; lg["A5"].fill = ORANGE_FILL; lg["A6"].fill = FIXED_FILL; lg["A7"].fill = PASS_FILL

    # Summary
    cnt = {}; pnl_total=0.0; pnl_red=0.0; pnl_orange=0.0
    for row in rows:
        _,_,st,_ = health_status(row); cnt[st]=cnt.get(st,0)+1
        p=f2(row.get("profit")) or 0.0; pnl_total+=p
        if st=="ไม่ผ่าน": pnl_red+=p
        if st=="เปิด STRONG_TREND_BLOCK": pnl_orange+=p
    sm = wb.create_sheet("สรุป")
    sm_rows = [
        ("หัวข้อ","ค่า"),
        ("จำนวน order ทั้งหมด", len(rows)),
        ("✅ ผ่าน", cnt.get("ผ่าน",0)),
        ("🟡 ผ่าน (แก้แล้ว)", cnt.get("ผ่าน (แก้แล้ว)",0)),
        ("🟠 เปิด STRONG_TREND_BLOCK", cnt.get("เปิด STRONG_TREND_BLOCK",0)),
        ("🔴 ไม่ผ่าน (ยังไม่มีกลไก)", cnt.get("ไม่ผ่าน",0)),
        ("รวม P/L (USD)", round(pnl_total,2)),
        ("P/L ส้ม (กันได้ถ้าเปิด flag)", round(pnl_orange,2)),
        ("P/L แดง (ยังไม่มีกลไก)", round(pnl_red,2)),
    ]
    for r,(a,b) in enumerate(sm_rows,1):
        ca=sm.cell(r,1,a); cb=sm.cell(r,2,b)
        ca.font=Font(name=FONT,bold=(r==1),size=10); cb.font=Font(name=FONT,size=10)
        if isinstance(b,(int,float)) and r>=7: cb.number_format="#,##0.00;(#,##0.00);-"
        if r==1: ca.fill=HDR_FILL; cb.fill=HDR_FILL; ca.font=HDR_FONT; cb.font=HDR_FONT
    sm.cell(3,1).fill=PASS_FILL; sm.cell(4,1).fill=FIXED_FILL; sm.cell(5,1).fill=ORANGE_FILL; sm.cell(6,1).fill=RED_FILL
    sm.column_dimensions["A"].width=30; sm.column_dimensions["B"].width=20
    safe_save(wb, os.path.join(OUTDIR,"orders_health.xlsx"))
    return cnt, pnl_total, pnl_red, pnl_orange

def build_cmp():
    rows = list(csv.DictReader(open(os.path.join(ROOT,"orders_old_vs_new.csv"), encoding="utf-8-sig")))
    headers = ["create_ts","close_ts","ticket","side","tf","sid","entry","sl","close",
               "OLD_profit","NEW_profit","diff","trend_filter"]
    wb = Workbook(); ws = wb.active; ws.title = "OLD_vs_NEW"
    write_sheet(ws, headers, rows, cmp_status, ["สถานะ guard","ระดับ","STATUS"])
    autosize(ws, {"A":19,"B":19,"C":11,"D":5,"E":5,"F":5,"G":9,"H":9,"I":9,
                  "J":11,"K":11,"L":9,"M":11,"N":14,"O":9,"P":15})
    safe_save(wb, os.path.join(OUTDIR,"orders_old_vs_new.xlsx"))

if __name__ == "__main__":
    cnt, tot, red, orange = build_health()
    build_cmp()
    print(f"สร้างเสร็จ -> {OUTDIR}")
    print(f"  ผ่าน={cnt.get('ผ่าน',0)} | แก้แล้ว={cnt.get('ผ่าน (แก้แล้ว)',0)} | "
          f"ส้ม(เปิด flag)={cnt.get('เปิด STRONG_TREND_BLOCK',0)} | แดง={cnt.get('ไม่ผ่าน',0)}")
    print(f"  P/L รวม={tot:.2f} | ส้ม={orange:.2f} | แดง={red:.2f}")
