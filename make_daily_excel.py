"""make_daily_excel.py — Daily orders + P/L Excel report"""
import re, os, sys, csv
from datetime import datetime
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
sys.stdout.reconfigure(encoding='utf-8', errors='replace')

# ── 1. Read logs ─────────────────────────────────────────────────
from log_sources import bot_log_files
log_files = bot_log_files()
fills = {}; closes = {}; seen_close = set()

def fld(line, key):
    m = re.search(rf'{key}=([^|\s]+)', line)
    return m.group(1) if m else ''

for path in log_files:
    if not os.path.exists(path): continue
    for line in open(path, encoding='utf-8', errors='replace'):
        m = re.match(r'^\[(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2})\]\s+(\S+)', line)
        if not m: continue
        ts, kind = m.group(1), m.group(2)
        tk = fld(line, 'ticket')
        if not tk: continue
        if kind == 'ENTRY_FILL' and tk not in fills:
            fills[tk] = {
                'fill_ts': ts, 'date': ts[:10],
                'side': fld(line,'side'), 'tf': fld(line,'tf'),
                'sid': fld(line,'sid'), 'price': fld(line,'price'),
                'sl': fld(line,'sl'), 'tp': fld(line,'tp'),
                'trend': fld(line,'trend'),
            }
        elif kind == 'POSITION_CLOSED' and tk not in seen_close and 'XAUUSD' in line:
            seen_close.add(tk)
            closes[tk] = {
                'profit': fld(line,'profit'), 'reason': fld(line,'reason'),
                'close_ts': ts, 'close_price': fld(line,'close_price'),
            }

# ── 2. Build order rows ──────────────────────────────────────────
order_rows = []
by_date_sid = defaultdict(lambda: defaultdict(float))
by_date = defaultdict(float)

for tk, fi in fills.items():
    cl = closes.get(tk, {})
    profit = float(cl.get('profit','0') or 0)
    if cl:
        sid = fi.get('sid','?')
        date = fi['date']
        by_date_sid[date][f'S{sid}'] += profit
        by_date[date] += profit
    side = fi.get('side','')
    trend = fi.get('trend','').lower()
    is_counter = (side=='BUY' and 'bear' in trend) or (side=='SELL' and 'bull' in trend)
    order_rows.append({
        'ticket': tk,
        'fill_ts': fi.get('fill_ts',''),
        'close_ts': cl.get('close_ts',''),
        'side': side,
        'tf': fi.get('tf',''),
        'sid': fi.get('sid',''),
        'fill_price': fi.get('price',''),
        'close_price': cl.get('close_price',''),
        'sl': fi.get('sl',''),
        'tp': fi.get('tp',''),
        'profit': profit if cl else '',
        'reason': cl.get('reason','') if cl else 'OPEN',
        'trend': fi.get('trend',''),
        'is_counter': '🚩' if is_counter else '',
        'is_sl': '❌SL' if 'sl' in cl.get('reason','').lower() else '',
        'status': 'CLOSED' if cl else 'OPEN',
    })

order_rows.sort(key=lambda x: x['fill_ts'], reverse=True)

# ── 3. Styles ────────────────────────────────────────────────────
H_FILL = PatternFill('solid', start_color='1F4E79')
H_FONT = Font(bold=True, color='FFFFFF', name='Arial', size=10)
T_FONT = Font(name='Arial', size=10)
B_FONT = Font(bold=True, name='Arial', size=10)
G_FILL = PatternFill('solid', start_color='C6EFCE')
R_FILL = PatternFill('solid', start_color='FFC7CE')
Y_FILL = PatternFill('solid', start_color='FFEB9C')
O_FILL = PatternFill('solid', start_color='FFD966')

wb = openpyxl.Workbook()

# ── Sheet 1: Daily P/L ───────────────────────────────────────────
ws1 = wb.active
ws1.title = 'Daily PL'
ws1['A1'] = 'Daily P/L Summary — Copter01 AI Bot'
ws1['A1'].font = Font(bold=True, name='Arial', size=13, color='1F4E79')
ws1.merge_cells('A1:L1')

# collect all sids
all_sids = sorted(set(sid for dd in by_date_sid.values() for sid in dd.keys()),
                  key=lambda x: int(x[1:]) if x[1:].isdigit() else 99)
headers = ['Date', 'Total'] + all_sids
for ci, h in enumerate(headers, 1):
    c = ws1.cell(3, ci, h)
    c.font = H_FONT; c.fill = H_FILL; c.alignment = Alignment(horizontal='center')

for ri, date in enumerate(sorted(by_date.keys()), 4):
    total = by_date[date]
    fill = G_FILL if total >= 0 else R_FILL
    ws1.cell(ri, 1, date).font = B_FONT
    c = ws1.cell(ri, 2, round(total,2))
    c.font = Font(bold=True, name='Arial', size=10, color='00AA00' if total>=0 else 'CC0000')
    c.number_format = '+#,##0.00;[Red]-#,##0.00'
    c.fill = fill
    for ci, sid in enumerate(all_sids, 3):
        v = by_date_sid[date].get(sid, 0)
        if abs(v) > 0.01:
            cell = ws1.cell(ri, ci, round(v,2))
            cell.font = T_FONT
            cell.number_format = '+#,##0.00;[Red]-#,##0.00'
            cell.fill = G_FILL if v >= 0 else R_FILL

ws1.column_dimensions['A'].width = 13
ws1.column_dimensions['B'].width = 12
for i in range(3, 3+len(all_sids)):
    ws1.column_dimensions[get_column_letter(i)].width = 11

# ── Sheet 2: Today's Orders ──────────────────────────────────────
ws2 = wb.create_sheet("Today 04-06-2026")
today_rows = [r for r in order_rows if r['fill_ts'].startswith('2026-06-04')]
COLS2 = ['ticket','fill_ts','close_ts','side','tf','sid','fill_price','close_price',
         'sl','tp','profit','reason','trend','is_counter','is_sl']
for ci, h in enumerate(COLS2, 1):
    c = ws2.cell(1, ci, h.upper())
    c.font = H_FONT; c.fill = H_FILL; c.alignment = Alignment(horizontal='center')

for ri, row in enumerate(today_rows, 2):
    profit_val = row.get('profit','')
    is_pos = isinstance(profit_val, (int,float)) and profit_val >= 0
    is_neg = isinstance(profit_val, (int,float)) and profit_val < 0
    base_fill = G_FILL if is_pos else (R_FILL if is_neg else PatternFill())
    if row.get('is_counter') == '🚩':
        base_fill = O_FILL  # orange for counter-trend bugs
    for ci, col in enumerate(COLS2, 1):
        raw = row.get(col, '')
        try:
            if col == 'profit' and raw != '':
                val = float(raw)
            elif col == 'sid' and raw != '':
                val = int(raw)
            else:
                val = raw
        except: val = raw
        c = ws2.cell(ri, ci, val)
        c.font = T_FONT; c.fill = base_fill
        if col == 'profit' and isinstance(val, float):
            c.number_format = '+#,##0.00;[Red]-#,##0.00'

total_today = sum(float(r['profit']) for r in today_rows if r['status']=='CLOSED' and r['profit']!='')
tr = len(today_rows) + 2
ws2.cell(tr, 1, 'TOTAL').font = B_FONT
c = ws2.cell(tr, 11, round(total_today,2))
c.font = Font(bold=True, name='Arial', size=10, color='00AA00' if total_today>=0 else 'CC0000')
c.number_format = '+#,##0.00;[Red]-#,##0.00'
c.fill = Y_FILL

widths2 = [14,20,20,6,10,5,12,12,10,10,10,22,16,10,8]
for i,w in enumerate(widths2,1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = 'A2'

# ── Sheet 3: All Orders ──────────────────────────────────────────
ws3 = wb.create_sheet("All Orders")
COLS3 = ['ticket','fill_ts','close_ts','side','tf','sid','fill_price','close_price',
         'sl','tp','profit','reason','trend','status','is_counter']
for ci, h in enumerate(COLS3, 1):
    c = ws3.cell(1, ci, h.upper())
    c.font = H_FONT; c.fill = H_FILL; c.alignment = Alignment(horizontal='center')

for ri, row in enumerate(order_rows, 2):
    profit_val = row.get('profit','')
    is_pos = isinstance(profit_val, (int,float)) and profit_val >= 0
    is_neg = isinstance(profit_val, (int,float)) and profit_val < 0
    base_fill = G_FILL if is_pos else (R_FILL if is_neg else PatternFill())
    if row.get('is_counter') == '🚩': base_fill = O_FILL
    for ci, col in enumerate(COLS3, 1):
        raw = row.get(col, '')
        try:
            if col == 'profit' and raw != '': val = float(raw)
            elif col == 'sid' and raw != '': val = int(raw)
            else: val = raw
        except: val = raw
        c = ws3.cell(ri, ci, val)
        c.font = T_FONT; c.fill = base_fill
        if col == 'profit' and isinstance(val, float):
            c.number_format = '+#,##0.00;[Red]-#,##0.00'

total_all = sum(float(r['profit']) for r in order_rows if r['status']=='CLOSED' and r['profit']!='')
tr3 = len(order_rows) + 2
ws3.cell(tr3, 1, 'TOTAL').font = B_FONT
c = ws3.cell(tr3, 11, round(total_all,2))
c.font = Font(bold=True, name='Arial', size=10, color='00AA00' if total_all>=0 else 'CC0000')
c.number_format = '+#,##0.00;[Red]-#,##0.00'; c.fill = Y_FILL

widths3 = [14,20,20,6,12,5,12,12,10,10,10,22,16,8,10]
for i,w in enumerate(widths3,1):
    ws3.column_dimensions[get_column_letter(i)].width = w
ws3.freeze_panes = 'A2'

out = 'excel_reports/daily_orders_report.xlsx'
wb.save(out)
print(f'Saved -> {out}')
print(f'Sheets: Daily P/L | Today 04-06 ({len(today_rows)} rows) | All Orders ({len(order_rows)} rows)')
print(f'Total all-time P/L: {total_all:+.2f}')
print(f'Today P/L: {total_today:+.2f}')
