import os
import sys
import argparse
import re
import pandas as pd
from datetime import datetime, timedelta, timezone
import MetaTrader5 as mt5

# Adjust path to import config if needed
sys.path.append(os.path.join(os.path.dirname(__file__), "..", "..", ".."))
import config

BKK = timezone(timedelta(hours=7))


def _tf_from_comment(comment: str) -> str:
    comment = str(comment or "")
    m = re.match(r"(\[[\w-]+\]|M\d+|H\d+|D\d+)(?:_S[\w.]+)?", comment)
    if not m:
        return ""
    return m.group(1).strip("[]").replace("-", "_")


def _deal_delta(deal) -> float:
    try:
        return float(deal.profit) + float(deal.commission) + float(deal.swap)
    except Exception:
        return 0.0


def _balance_deal_delta(deal) -> float:
    balance_types = [
        getattr(mt5, "DEAL_TYPE_BALANCE", None),
        getattr(mt5, "DEAL_TYPE_CREDIT", None),
        getattr(mt5, "DEAL_TYPE_CHARGE", None),
        getattr(mt5, "DEAL_TYPE_CORRECTION", None),
        getattr(mt5, "DEAL_TYPE_BONUS", None),
        getattr(mt5, "DEAL_TYPE_COMMISSION", None),
        getattr(mt5, "DEAL_TYPE_COMMISSION_DAILY", None),
        getattr(mt5, "DEAL_TYPE_COMMISSION_MONTHLY", None),
        getattr(mt5, "DEAL_TYPE_INTEREST", None),
    ]
    balance_types = {t for t in balance_types if t is not None}
    if getattr(deal, "type", None) in balance_types:
        return _deal_delta(deal)
    if getattr(deal, "entry", None) == mt5.DEAL_ENTRY_OUT:
        return _deal_delta(deal)
    return 0.0


def _build_balance_after_by_deal(deals, current_balance: float) -> dict[int, float]:
    """Reconstruct account balance after each balance-affecting deal in history."""
    balance_deals = [d for d in deals if abs(_balance_deal_delta(d)) > 1e-9]
    if not balance_deals:
        return {}
    total_delta = sum(_balance_deal_delta(d) for d in balance_deals)
    running = float(current_balance) - total_delta
    result = {}
    for d in sorted(balance_deals, key=lambda x: (x.time, getattr(x, "ticket", 0))):
        running += _balance_deal_delta(d)
        result[int(getattr(d, "ticket", 0))] = running
    return result


def _fmt_money(value):
    try:
        if value is None or pd.isna(value):
            return None
        return f"{float(value):.2f}"
    except Exception:
        return None

def parse_args():
    parser = argparse.ArgumentParser(description="Compare S20.12 SIM vs MT5 orders")
    parser.add_argument("--start", type=str, default=None,
                         help="เวลาเริ่มต้นแบบเดียวกับที่ใช้รัน backtest_S20_12_runner_mt5.py "
                              "(dd-MM-yyyy HH:mm, BKK) — ใช้กรอง order กำพร้าไม่ให้โผล่นอกช่วงที่ขอจริง "
                              "ถ้าไม่ระบุ จะ fallback ไปใช้เวลาของไม้แรกใน SIM แทน (อาจกว้างกว่าที่ตั้งใจ)")
    parser.add_argument("--end", type=str, default=None,
                         help="เวลาสิ้นสุดแบบเดียวกับที่ใช้รัน backtest_S20_12_runner_mt5.py "
                              "(dd-MM-yyyy HH:mm, BKK) — ถ้าไม่ระบุ จะ fallback ไปใช้เวลาปิดของไม้สุดท้าย"
                              "ใน SIM แทน (อาจแคบกว่าที่ตั้งใจ ตัด order ท้ายช่วงที่ควรอยู่ในขอบเขตออกไป)")
    return parser.parse_args()

def main():
    args = parse_args()
    sim_csv = os.path.join(os.path.dirname(__file__), "..", "excel", "s20_12_sim_trades.csv")
    if not os.path.exists(sim_csv):
        print(f"❌ ไม่พบไฟล์ {sim_csv}\nกรุณารัน backtest_S20_12_runner_mt5.py ก่อนครับ")
        return

    print("📊 กำลังโหลดข้อมูลออเดอร์จำลอง (SIM)...")
    df_sim = pd.read_csv(sim_csv)
    if df_sim.empty:
        print("❌ ไฟล์ SIM ไม่มีข้อมูลออเดอร์")
        return

    # Parse BKK times to UTC for MT5
    df_sim["Time (BKK)"] = pd.to_datetime(df_sim["Time (BKK)"])
    df_sim["Close Time"] = pd.to_datetime(df_sim["Close Time"])

    # ช่วง SIM จริง (ไม่ padding) — ใช้กรองว่า order ไหน "อยู่ในคำขอจริง" ตอนแสดงผล
    # ถ้าระบุ --start มา ใช้ค่านั้นตรงๆ (ตรงกับที่ผู้ใช้ตั้งใจ) ไม่งั้น fallback ไปใช้เวลาไม้แรกใน
    # SIM (ซึ่งอาจช้ากว่าค่า --start จริงที่เคยสั่งไปหลายนาที เพราะเป็นเวลาที่ signal แรกถูกตรวจเจอ
    # ไม่ใช่เวลาที่ขอ — ทำให้ order ก่อนหน้านั้นเล็กน้อยยังหลุดผ่านมาโชว์ได้)
    if args.start:
        _sim_start_exact = pd.Timestamp(datetime.strptime(args.start, "%d-%m-%Y %H:%M"))
    else:
        _sim_start_exact = df_sim["Time (BKK)"].min()
    if args.end:
        _sim_end_exact = pd.Timestamp(datetime.strptime(args.end, "%d-%m-%Y %H:%M"))
    else:
        _sim_end_exact = df_sim["Close Time"].max()

    # ช่วง query MT5 — padding ±1h กันพลาดขอบเขต (คนละจุดประสงค์กับตัวกรองแสดงผลด้านบน)
    start_bkk = _sim_start_exact - timedelta(hours=1)
    end_bkk = _sim_end_exact + timedelta(hours=1)

    print(f"🕒 ช่วงเวลา SIM: {start_bkk} ถึง {end_bkk}")

    if not config.mt5_initialize(mt5):
        print("❌ MT5 Initialize Failed")
        return

    resolved_symbol = config.SYMBOL

    # ใช้ naive UTC+7 แปลง BKK->UTC เพื่อกำหนดขอบเขต query เท่านั้น (MT5_SERVER_TZ ต่างจาก
    # BKK ไม่คงที่ + ห่างได้หลายชม.) แล้วบวก padding กว้างๆ กันเคส server tz ดันช่วงเวลาออกนอกขอบ
    start_utc = start_bkk.tz_localize(BKK).astimezone(timezone.utc) - timedelta(hours=6)
    end_utc = end_bkk.tz_localize(BKK).astimezone(timezone.utc) + timedelta(hours=6)

    print("📥 กำลังดึงประวัติการเทรดจาก MT5...")
    deals = mt5.history_deals_get(start_utc, end_utc)
    account = mt5.account_info()
    current_balance = float(account.balance) if account else 0.0
    balance_end_utc = datetime.now(timezone.utc) + timedelta(hours=6)
    balance_deals = mt5.history_deals_get(start_utc, balance_end_utc) or deals or []
    mt5_balance_after = _build_balance_after_by_deal(balance_deals, current_balance)
    if deals is None or len(deals) == 0:
        print("⚠️ ไม่พบประวัติการเทรดใน MT5 ในช่วงเวลานี้")
        mt5.shutdown()
        return

    # Filter MT5 deals
    act_trades = []
    for d in deals:
        if d.symbol != resolved_symbol:
            continue
        # Only closed deals (OUT) to compare with sim closures
        if d.entry != mt5.DEAL_ENTRY_OUT:
            continue
        
        dt_bkk = config.mt5_ts_to_bkk(d.time)
        typ = "BUY" if d.type == mt5.DEAL_TYPE_SELL else "SELL" # DEAL_ENTRY_OUT for BUY is a SELL deal

        # S20.12 uses SID 20.12, so the original IN deal comment should contain "20.12"
        in_deals = mt5.history_deals_get(position=d.position_id)
        is_target_strat = False
        open_dt_bkk = None
        mt5_open_price = None
        mt5_tf = ""
        if in_deals:
            for ind in in_deals:
                if ind.entry == mt5.DEAL_ENTRY_IN and ind.comment and ("20.12" in str(ind.comment)):
                    is_target_strat = True
                    open_dt_bkk = config.mt5_ts_to_bkk(ind.time)
                    mt5_open_price = ind.price
                    mt5_tf = _tf_from_comment(ind.comment)
                    break
        if not is_target_strat:
             continue

        # ดึง SL/TP จาก opening order ของ position นี้
        mt5_sl, mt5_tp = None, None
        orders = mt5.history_orders_get(position=d.position_id)
        if orders:
            for o in orders:
                if o.sl or o.tp:
                    mt5_sl = o.sl if o.sl else None
                    mt5_tp = o.tp if o.tp else None
                    break

        act_trades.append({
            "MT5_Open_Time": open_dt_bkk,
            "MT5_Close_Time": dt_bkk,
            "MT5_TF": mt5_tf,
            "MT5_Type": typ,
            "MT5_Entry": mt5_open_price,
            "MT5_Close_Price": d.price,
            "MT5_Volume": d.volume,
            "MT5_SL": mt5_sl,
            "MT5_TP": mt5_tp,
            "MT5_P&L": d.profit + d.commission + d.swap,
            "MT5_Balance": _fmt_money(mt5_balance_after.get(int(getattr(d, "ticket", 0)))),
            "MT5_Comment": d.comment,
            "MT5_Position_ID": d.position_id
        })

    df_act = pd.DataFrame(act_trades)
    print(f"✅ พบประวัติการปิดออเดอร์ใน MT5: {len(df_act)} รายการ")

    # Merge logic — open ต้องตรงระดับ ชม:นาที, close ต้องต่างกันไม่เกิน 10 วินาที
    # (SIM บันทึก close time เป็น bar open time ที่ตรวจเจอ SL/TP ส่วน MT5 execute จริงอาจช้ากว่า
    # เล็กน้อย ใช้ 10s tolerance กัน mismatch จากความล่าช้าเล็กน้อยหลังบาร์เปิด)
    # ถ้าไม่มี MT5 record ไหนตรงเป๊ะ ให้ปล่อยว่าง (ไม่ force-pair กับแถวที่ใกล้สุดแต่ไม่ใช่คู่จริง
    # เพราะจะหลอกตาว่า "จับคู่ได้" ทั้งที่เป็นออเดอร์คนละตัว)
    def _hhmm_match(a, b) -> bool:
        if a is None or b is None or pd.isna(a) or pd.isna(b):
            return False
        return (a.hour, a.minute) == (b.hour, b.minute)

    def _close_match(a, b) -> bool:
        if a is None or b is None or pd.isna(a) or pd.isna(b):
            return False
        try:
            return abs((a - b).total_seconds()) <= 65
        except Exception:
            return _hhmm_match(a, b)

    def _price_match(a, b, tol: float = 0.08) -> bool:
        try:
            if a is None or b is None or pd.isna(a) or pd.isna(b):
                return False
            return abs(float(a) - float(b)) <= tol
        except Exception:
            return False

    def _sl_tp_match(sim, cand) -> bool:
        return (
            _price_match(sim.get("SL"), cand.get("MT5_SL"))
            and _price_match(sim.get("TP"), cand.get("MT5_TP"))
        )

    results = []
    matched_indices = set()
    for _, sim in df_sim.iterrows():
        sim_open = sim["Time (BKK)"]
        sim_close = sim["Close Time"]
        sim_type = sim["Type"]
        sim_tf = str(sim["TF"])

        match = None
        if not df_act.empty:
            subset = df_act[
                (df_act["MT5_Type"] == sim_type)
                & (df_act["MT5_TF"].astype(str) == sim_tf)
                & (~df_act.index.isin(matched_indices))
            ].copy()
            for idx, cand in subset.iterrows():
                cand_open = cand["MT5_Open_Time"].tz_localize(None) if cand["MT5_Open_Time"] is not None else None
                cand_close = cand["MT5_Close_Time"].tz_localize(None) if cand["MT5_Close_Time"] is not None else None
                if (
                    _hhmm_match(sim_open, cand_open)
                    and _close_match(sim_close, cand_close)
                    and _sl_tp_match(sim, cand)
                ):
                    match = cand
                    matched_indices.add(idx)
                    break

        mt5_open = match["MT5_Open_Time"].tz_localize(None) if match is not None and match["MT5_Open_Time"] is not None else None
        mt5_close = match["MT5_Close_Time"].tz_localize(None) if match is not None else None

        res = {
            "SIM_Open_Time":  sim["Time (BKK)"],
            "MT5_Open_Time":  mt5_open.strftime('%Y-%m-%d %H:%M:%S') if mt5_open is not None else None,
            "SIM_Close_Time": sim["Close Time"],
            "MT5_Close_Time": mt5_close.strftime('%Y-%m-%d %H:%M:%S') if mt5_close is not None else None,
            "SIM_TF":         sim["TF"],
            "MT5_TF":         match["MT5_TF"] if match is not None else None,
            "SIM_Type":       sim["Type"],
            "MT5_Type":       match["MT5_Type"] if match is not None else None,
            "SIM_Entry":      sim["Entry"],
            "MT5_Entry":      match["MT5_Entry"] if match is not None else None,
            "MT5_Close_Price": match["MT5_Close_Price"] if match is not None else None,
            "SIM_SL":         sim["SL"] if "SL" in sim else None,
            "MT5_SL":         match["MT5_SL"] if match is not None else None,
            "SIM_TP":         sim["TP"] if "TP" in sim else None,
            "MT5_TP":         match["MT5_TP"] if match is not None else None,
            "SIM_Lot":        sim["Lot"] if "Lot" in sim else None,
            "MT5_Volume":     match["MT5_Volume"] if match is not None else None,
            "SIM_P&L":        sim["P&L"],
            "MT5_P&L":        match["MT5_P&L"] if match is not None else None,
            "SIM_Balance":    sim["Balance"] if "Balance" in sim else None,
            "MT5_Balance":    match["MT5_Balance"] if match is not None else None,
            "MT5_Comment":    match["MT5_Comment"] if match is not None else None,
            "MT5_Position_ID": match["MT5_Position_ID"] if match is not None else None,
            "Matched":        match is not None,
            "SIM_Reason":     sim["Reason"],
        }
        results.append(res)

    # ── ออเดอร์ MT5 จริงที่ไม่มี SIM คู่เลย (เช่น backtest ไม่เจอ pattern เดียวกัน) ──
    # loop ด้านบนวนตาม SIM เป็นหลัก ถ้าไม่ทำส่วนนี้ ออเดอร์จริงที่ไม่ match จะหายไปจากไฟล์
    # ทั้งที่มีอยู่จริงใน MT5 — เพิ่มมาแสดงแยกเพื่อให้ตรวจสอบได้ว่ามีไม้ไหนที่ backtest มองไม่เห็น
    # หมายเหตุ: query MT5 กว้างกว่าช่วง SIM จริงมาก (±1h ก่อน query + ±6h กัน server-tz) — ถ้าไม่
    # กรองตรงนี้ ออเดอร์นอกช่วงที่ขอจริงๆ จะโผล่มาปนด้วยจนรก ต้องกรองด้วยช่วง SIM แบบไม่ padding
    # (_sim_start_exact/_sim_end_exact) ก่อนนำมาแสดงเป็นแถวกำพร้า ไม่ใช่ start_bkk/end_bkk ที่มี ±1h
    _sim_start_naive = _sim_start_exact.replace(tzinfo=None) if _sim_start_exact.tzinfo else _sim_start_exact
    _sim_end_naive = _sim_end_exact.replace(tzinfo=None) if _sim_end_exact.tzinfo else _sim_end_exact
    if not df_act.empty:
        unmatched_act = df_act[~df_act.index.isin(matched_indices)]
        for _, act in unmatched_act.iterrows():
            act_open = act["MT5_Open_Time"].tz_localize(None) if act["MT5_Open_Time"] is not None else None
            act_close = act["MT5_Close_Time"].tz_localize(None) if act["MT5_Close_Time"] is not None else None
            if act_open is None or not (_sim_start_naive <= act_open <= _sim_end_naive):
                continue  # นอกช่วงเวลา SIM ที่ขอจริง — ไม่เอามาแสดง กันรก
            results.append({
                "SIM_Open_Time":  None,
                "MT5_Open_Time":  act_open.strftime('%Y-%m-%d %H:%M:%S') if act_open is not None else None,
                "SIM_Close_Time": None,
                "MT5_Close_Time": act_close.strftime('%Y-%m-%d %H:%M:%S') if act_close is not None else None,
                "SIM_TF":         None,
                "MT5_TF":         act["MT5_TF"],
                "SIM_Type":       None,
                "MT5_Type":       act["MT5_Type"],
                "SIM_Entry":      None,
                "MT5_Entry":      act["MT5_Entry"],
                "MT5_Close_Price": act["MT5_Close_Price"],
                "SIM_SL":         None,
                "MT5_SL":         act["MT5_SL"],
                "SIM_TP":         None,
                "MT5_TP":         act["MT5_TP"],
                "SIM_Lot":        None,
                "MT5_Volume":     act["MT5_Volume"],
                "SIM_P&L":        None,
                "MT5_P&L":        act["MT5_P&L"],
                "SIM_Balance":    None,
                "MT5_Balance":    act["MT5_Balance"],
                "MT5_Comment":    act["MT5_Comment"],
                "MT5_Position_ID": act["MT5_Position_ID"],
                "Matched":        False,
                "SIM_Reason":     "ไม่มี SIM คู่ — backtest ไม่เจอ pattern นี้",
            })

    df_res = pd.DataFrame(results)
    if not df_res.empty:
        # เรียงตาม Open Time — ใช้ SIM_Open_Time ถ้ามี ไม่งั้น fallback ไป MT5_Open_Time
        # (แถวที่เป็นออเดอร์ MT5 กำพร้า ไม่มี SIM คู่ จะได้เรียงตามเวลาที่เกิดขึ้นจริงแทนที่จะตกท้ายไฟล์)
        sort_key = pd.to_datetime(df_res["SIM_Open_Time"]).fillna(pd.to_datetime(df_res["MT5_Open_Time"]))
        df_res = df_res.assign(_sort_key=sort_key).sort_values("_sort_key").drop(columns="_sort_key").reset_index(drop=True)
    
    out_dir = os.path.join(os.path.dirname(__file__), "..", "excel")
    os.makedirs(out_dir, exist_ok=True)
    out_csv  = os.path.join(out_dir, "compare_s20_12.csv")
    out_xlsx = os.path.join(out_dir, "compare_s20_12.xlsx")

    df_res.to_csv(out_csv, index=False)

    # ── Excel พร้อมสี SIM (ฟ้า) vs MT5 (ส้ม) ─────────────────────────────
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter

        SIM_HEADER = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
        MT5_HEADER = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")
        OTH_HEADER = PatternFill(start_color="595959", end_color="595959", fill_type="solid")
        SIM_ROW    = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
        MT5_ROW    = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        WHITE_FONT = Font(color="FFFFFF", bold=True)

        wb = Workbook()
        ws = wb.active
        ws.title = "Compare S20.12"

        cols = list(df_res.columns)
        for ci, col in enumerate(cols, start=1):
            cell = ws.cell(row=1, column=ci, value=col)
            cell.font = WHITE_FONT
            cell.alignment = Alignment(horizontal="center")
            if col.startswith("SIM_"):
                cell.fill = SIM_HEADER
            elif col.startswith("MT5_"):
                cell.fill = MT5_HEADER
            else:
                cell.fill = OTH_HEADER

        for ri, row in enumerate(df_res.itertuples(index=False), start=2):
            for ci, col in enumerate(cols, start=1):
                val = getattr(row, col.replace("&", "_").replace(" ", "_"), None)
                # pandas มักเปลี่ยน & เป็น _ ใน namedtuple — fallback ดึงจาก dict
                if val is None:
                    val = df_res.iloc[ri - 2][col]
                ws.cell(row=ri, column=ci, value=val)
                if col.startswith("SIM_"):
                    ws.cell(row=ri, column=ci).fill = SIM_ROW
                elif col.startswith("MT5_"):
                    ws.cell(row=ri, column=ci).fill = MT5_ROW

        # auto width
        for ci, col in enumerate(cols, start=1):
            max_len = max(
                len(str(col)),
                *(len(str(df_res.iloc[r][col]) if df_res.iloc[r][col] is not None else "") for r in range(len(df_res)))
            )
            ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 2, 40)

        wb.save(out_xlsx)
        print(f"🎉 สร้างไฟล์เปรียบเทียบเสร็จสมบูรณ์!\n📍 CSV : {out_csv}\n📍 XLSX: {out_xlsx}")
    except ImportError:
        print(f"🎉 สร้างไฟล์เปรียบเทียบเสร็จสมบูรณ์!\n📍 CSV : {out_csv}\n⚠️  ไม่พบ openpyxl — ข้าม XLSX (pip install openpyxl)")
    except Exception as _e:
        print(f"📍 CSV : {out_csv}\n⚠️  สร้าง XLSX ไม่สำเร็จ: {_e}")
    
    # Summary
    matched = df_res[df_res["Matched"] == True]
    print("-" * 40)
    print(f"สรุปการจับคู่:")
    print(f"ออเดอร์ SIM ทั้งหมด: {len(df_res)}")
    print(f"จับคู่กับ MT5 ได้: {len(matched)} ({len(matched)/len(df_res)*100:.2f}%)")
    print("-" * 40)

    mt5.shutdown()

if __name__ == "__main__":
    main()
