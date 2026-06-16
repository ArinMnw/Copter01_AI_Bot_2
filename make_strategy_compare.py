"""
make_strategy_compare.py
─────────────────────────
สร้างไฟล์ Excel เปรียบเทียบ OLD (ผลจริง) vs NEW (เพิ่ม swing direction check)
สำหรับทุก strategy ตามช่วงเวลาที่กำหนด — แยก sheet ตามแต่ละท่า + sheet รวมทุกท่า

โครงสร้างเหมือน s14_compare_old_new / compare_all_jun10_12

วิธีรัน:
    python make_strategy_compare.py --start 2026-06-15
    python make_strategy_compare.py --start 2026-06-10 --end 2026-06-12
    python make_strategy_compare.py --start 2026-06-15 --out excel_reports/my_report.xlsx

ค่า default:
    --end   = วันปัจจุบัน (รวมทั้งวัน)
    --out   = excel_reports/compare_all_<start>_<end>.xlsx
    --logs  = logs/old_logs/bot-*.log + logs/bot.log (auto-glob)
"""
import re
import sys
import glob
import argparse
import datetime as _dt
from pathlib import Path

# console-safe output (Windows cp874 encode ตัวอักษรพิเศษไม่ได้)
try:
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')
except Exception:
    pass

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Constants ───────────────────────────────────────────────────────────────
ROOT = Path(__file__).resolve().parent

SID_NAME = {
    '1': 'S1 Zone/Engulf', '2': 'S2 FVG', '3': 'S3 EMA Bounce',
    '4': 'S4 Supply/Demand', '5': 'S5 BB Squeeze', '6': 'S6 RSI OB/OS',
    '7': 'S7 Candle', '8': 'S8 SR', '9': 'S9 RSI Div',
    '10': 'S10 CRT', '11': 'S11 Fibo', '12': 'S12 LiqSweep',
    '13': 'S13 OB', '14': 'S14 SweepRSI', '15': 'S15 VP Rev',
    '16': 'S16 Sideway', '17': 'S17 SweepSniper',
    '18': 'S18 TJR/ICT', '19': 'S19 ICT Adv',
}

FONT_NAME = 'Arial'
BG_TITLE  = '2F4F8F'; BG_OLD = 'D6E4F7'; BG_NEW = 'D6F5D6'; BG_DIFF = 'FFF2CC'
BG_HEAD   = 'E8ECEF'; BG_TOTAL = 'F0F0F0'; BG_TP = 'C6EFCE'; BG_SL = 'FFC7CE'
BG_BOT    = 'FFEB9C'; BG_BOT_SW = 'FFD966'

thin = Side(style='thin', color='BFBFBF')
med  = Side(style='medium', color='404040')
THIN_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
MED_BOTTOM  = Border(left=thin, right=thin, top=thin, bottom=med)

COMPARE_COLS = [
    ('TF', 6), ('Signal', 7), ('Pattern', 18), ('Fill Time', 14), ('Ticket', 12),
    ('Entry', 8), ('SL', 8), ('TP', 8), ('Close Price', 10), ('Close Type', 11),
    ('P&L', 8), ('New Close Type', 13), ('New P&L', 8), ('DIFF P&L', 9),
]
SUM_COLS = ['Strategy', 'Trades', 'TP', 'SL', 'Bot Close', 'WR%', 'P&L (USD)',
            'NEW Trades', 'NEW TP', 'NEW SL', 'NEW Bot', 'NEW WR%', 'NEW P&L', 'DIFF P&L']


# ─── Log parsing ─────────────────────────────────────────────────────────────
def parse_logs(log_paths, start_date, end_date):
    """อ่าน log → คืน list ของ closed orders (merge ORDER_CREATED + ENTRY_FILL + POSITION_CLOSED)"""
    closed = []
    order_meta = {}

    for logpath in log_paths:
        try:
            f = open(logpath, encoding='utf-8', errors='replace')
        except OSError:
            continue
        with f:
            for line in f:
                ts_m = re.match(r'\[(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2})\]', line)
                if not ts_m:
                    continue
                ts = ts_m.group(1)
                day = ts[:10]
                if day < start_date or day > end_date:
                    continue

                if 'ORDER_CREATED' in line:
                    t   = re.search(r'ticket=(\d+)', line)
                    if not t:
                        continue
                    lbl = re.search(r'hhll_last_label=(\w+)', line)
                    tf  = re.search(r'\btf=(\w+)', line)
                    sid = re.search(r'\bsid=(\d+)', line)
                    sig = re.search(r'signal=(\w+)', line)
                    e   = re.search(r'entry=([\d.]+)', line)
                    sl  = re.search(r'\bsl=([\d.]+)', line)
                    tp  = re.search(r'\btp=([\d.]+)', line)
                    trf = re.search(r'trend_filter=(\S+)', line)
                    order_meta[t.group(1)] = {
                        'last_label':   lbl.group(1) if lbl else '',
                        'trend_filter': trf.group(1) if trf else '',
                        'tf':  tf.group(1)  if tf  else '',
                        'sid': sid.group(1) if sid else '',
                        'side': sig.group(1) if sig else '',
                        'entry': e.group(1) if e else '',
                        'sl':  sl.group(1)  if sl  else '',
                        'tp':  tp.group(1)  if tp  else '',
                    }

                elif 'ENTRY_FILL' in line:
                    t = re.search(r'ticket=(\d+)', line)
                    if not t:
                        continue
                    tk = t.group(1)
                    trend = re.search(r'trend=([^|]+)', line)
                    price = re.search(r'\bprice=([\d.]+)', line)
                    sl    = re.search(r'\bsl=([\d.]+)', line)
                    tp    = re.search(r'\btp=([\d.]+)', line)
                    meta = order_meta.setdefault(tk, {})
                    meta['fill_ts'] = ts
                    meta['trend']   = trend.group(1).strip() if trend else meta.get('trend_filter', '')
                    if price and not meta.get('entry'):
                        meta['entry'] = price.group(1)
                    if sl and not meta.get('sl'):
                        meta['sl'] = sl.group(1)
                    if tp and not meta.get('tp'):
                        meta['tp'] = tp.group(1)

                elif 'POSITION_CLOSED' in line:
                    t = re.search(r'ticket=(\d+)', line)
                    if not t:
                        continue
                    sid    = re.search(r'\bsid=(\d+)', line)
                    sig    = re.search(r'\bside=(\w+)', line)
                    tf     = re.search(r'\btf=(\w+)', line)
                    op     = re.search(r'open_price=([\d.]+)', line)
                    cp     = re.search(r'close_price=([\d.]+)', line)
                    sl     = re.search(r'\bsl=([\d.]+)', line)
                    tp_m   = re.search(r'\btp=([\d.]+)', line)
                    pnl    = re.search(r'profit=([-\d.]+)', line)
                    reason = re.search(r'reason=([^|]+)', line)
                    pat    = re.search(r'pattern=([^|]+)', line)
                    trf    = re.search(r'trend_filter=(\S+)', line)
                    closed.append({
                        'ticket':       t.group(1),
                        'sid':          sid.group(1)    if sid    else '',
                        'side':         sig.group(1)    if sig    else '',
                        'tf':           tf.group(1)     if tf     else '',
                        'entry':        op.group(1)     if op     else '',
                        'close_price':  cp.group(1)     if cp     else '',
                        'sl':           sl.group(1)     if sl     else '',
                        'tp':           tp_m.group(1)   if tp_m   else '',
                        'profit':       pnl.group(1)    if pnl    else '',
                        'reason':       reason.group(1).strip() if reason else '',
                        'pattern':      pat.group(1).strip()    if pat    else '',
                        'trend_filter': trf.group(1)    if trf    else '',
                        'close_ts':     ts,
                    })

    # dedup (log rotate ทำให้บางบรรทัดซ้ำข้ามไฟล์ .bak) — key = ticket + close_ts
    seen = set()
    deduped = []
    for row in closed:
        key = (row['ticket'], row['close_ts'])
        if key in seen:
            continue
        seen.add(key)
        deduped.append(row)
    closed = deduped

    # merge metadata
    for row in closed:
        meta = order_meta.get(row['ticket'], {})
        row['fill_ts']    = meta.get('fill_ts', row['close_ts'])
        row['last_label'] = meta.get('last_label', '')
        row['trend']      = meta.get('trend', row.get('trend_filter', ''))
        for k in ('entry', 'sl', 'tp'):
            if not row[k]:
                row[k] = meta.get(k, '')
    return closed


# ─── Verdict logic ───────────────────────────────────────────────────────────
def close_type(row):
    r = row.get('reason', '')
    if '[tp ' in r:                       return 'TP'
    if '[sl ' in r:                       return 'SL'
    if 'Trend Reche' in r:                return 'Bot-Trend'
    if 'PD Zone' in r:                    return 'Bot-PD'
    if 'SL Guard' in r:                   return 'Bot-SLGuard'
    if 'sweep exit' in r or 'engulf exit' in r: return 'Bot-Exit'
    if r in ('-', ''):                    return '-'
    return 'Bot-Other'


def new_verdict(row):
    """NEW = OLD + swing direction filter. คืน close type ที่จะเกิดถ้ามี swing dir check."""
    ct = close_type(row)
    if ct.startswith('Bot-'):
        return ct  # already blocked by existing filter
    side = row.get('side', '')
    lbl  = (row.get('last_label', '') or '').lower()
    if not lbl:
        return '?'
    if side == 'BUY':
        if lbl in ('ll', 'lh'): return 'Bot-SwingDir'
        if lbl == 'hh':         return ct
        return '?(HL?)'         # hl — ไม่มี prev เทียบ
    if side == 'SELL':
        if lbl in ('hh', 'hl'): return 'Bot-SwingDir'
        if lbl == 'll':         return ct
        return '?(LH?)'         # lh — ไม่มี prev เทียบ
    return ct


def new_pnl(row):
    nct = new_verdict(row)
    pnl = safe_float(row.get('profit'))
    if nct == 'Bot-SwingDir':
        return 0.0            # ถูก block → ปิดที่ entry (ประมาณ 0)
    return pnl


# ─── Helpers ─────────────────────────────────────────────────────────────────
def safe_float(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def fmt_ts(ts):
    return ts[5:16].replace('-', '/') if ts else '-'


def fmt_pattern(p):
    m = re.search(r'ท่าที่ \d+ (.+?) \[', p) or re.search(r'ท่าที่ \d+ (.+)', p)
    if m:
        return m.group(1).strip()[:25]
    return (p[:25] if p else '-')


def pnl_bg(v):
    if v is None or v == 0: return None
    return BG_TP if v > 0 else BG_SL


def ctype_bg(ct):
    if ct == 'TP':            return BG_TP
    if ct == 'SL':            return BG_SL
    if ct == 'Bot-SwingDir':  return BG_BOT_SW
    if ct.startswith('Bot-'): return BG_BOT
    return None


def cell(ws, r, c, value, bold=False, bg=None, color=None, align='center', num_fmt=None, border=None):
    cl = ws.cell(row=r, column=c, value=value)
    cl.font = Font(name=FONT_NAME, bold=bold, color=color or '000000')
    cl.alignment = Alignment(horizontal=align, vertical='center')
    if bg:      cl.fill = PatternFill('solid', start_color=bg)
    if num_fmt: cl.number_format = num_fmt
    if border:  cl.border = border
    return cl


# ─── Sheet builders ──────────────────────────────────────────────────────────
def build_compare_sheet(ws, rows, title):
    cell(ws, 1, 1, title, bold=True, bg=BG_TITLE, color='FFFFFF', align='left')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COMPARE_COLS))

    cell(ws, 2, 1, 'Trade Info', bold=True, bg=BG_HEAD)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=5)
    cell(ws, 2, 6, 'ACTUAL (OLD code)', bold=True, bg=BG_OLD)
    ws.merge_cells(start_row=2, start_column=6, end_row=2, end_column=11)
    cell(ws, 2, 12, 'NEW (swing dir check)', bold=True, bg=BG_NEW)
    ws.merge_cells(start_row=2, start_column=12, end_row=2, end_column=13)
    cell(ws, 2, 14, 'DIFF', bold=True, bg=BG_DIFF)

    for ci, (hdr, w) in enumerate(COMPARE_COLS, 1):
        bg = BG_OLD if 6 <= ci <= 11 else BG_NEW if 12 <= ci <= 13 else BG_DIFF if ci == 14 else BG_HEAD
        cell(ws, 3, ci, hdr, bold=True, bg=bg, border=MED_BOTTOM)
        ws.column_dimensions[get_column_letter(ci)].width = w

    ridx = 4
    for r in sorted(rows, key=lambda x: (x.get('sid', ''), x.get('tf', ''), x.get('fill_ts', ''))):
        ct   = close_type(r)
        nct  = new_verdict(r)
        pnl  = safe_float(r.get('profit'))
        npnl = new_pnl(r)
        diff = (npnl - pnl) if (pnl is not None and npnl is not None) else None
        vals = [r.get('tf', '-'), r.get('side', '-'), fmt_pattern(r.get('pattern', '-')),
                fmt_ts(r.get('fill_ts', '')), r.get('ticket', '-'),
                safe_float(r.get('entry')), safe_float(r.get('sl')), safe_float(r.get('tp')),
                safe_float(r.get('close_price')), ct, pnl, nct, npnl, diff]
        bgs = [None]*5 + [None, None, None, None, ctype_bg(ct), pnl_bg(pnl),
                          ctype_bg(nct), pnl_bg(npnl),
                          (None if diff is None else (BG_TP if diff > 0 else BG_SL if diff < 0 else None))]
        for ci, (v, bg) in enumerate(zip(vals, bgs), 1):
            nf = '#,##0.00' if ci in (6, 7, 8, 9, 11, 13, 14) else None
            cell(ws, ridx, ci, v, bg=bg, num_fmt=nf, border=THIN_BORDER)
        ridx += 1

    if ridx > 4:
        cell(ws, ridx, 1, 'TOTAL', bold=True, bg=BG_TOTAL, border=THIN_BORDER)
        ws.merge_cells(start_row=ridx, start_column=1, end_row=ridx, end_column=10)
        cell(ws, ridx, 11, f'=SUM(K4:K{ridx-1})', bold=True, bg=BG_TOTAL, num_fmt='#,##0.00', border=THIN_BORDER)
        cell(ws, ridx, 12, '-', bg=BG_TOTAL, border=THIN_BORDER)
        cell(ws, ridx, 13, f'=SUM(M4:M{ridx-1})', bold=True, bg=BG_TOTAL, num_fmt='#,##0.00', border=THIN_BORDER)
        cell(ws, ridx, 14, f'=SUM(N4:N{ridx-1})', bold=True, bg=BG_DIFF, num_fmt='#,##0.00', border=THIN_BORDER)
    ws.freeze_panes = 'A4'


def sid_stats(rows):
    tp  = sum(1 for r in rows if close_type(r) == 'TP')
    sl  = sum(1 for r in rows if close_type(r) == 'SL')
    bot = sum(1 for r in rows if close_type(r).startswith('Bot-'))
    pnl = sum(safe_float(r.get('profit')) or 0 for r in rows)
    wr  = tp / (tp + sl) if (tp + sl) else 0
    return len(rows), tp, sl, bot, wr, round(pnl, 2)


def sid_new_stats(rows):
    tp = sl = bot = 0
    total = 0.0
    for r in rows:
        nct = new_verdict(r)
        p   = safe_float(r.get('profit')) or 0
        if nct == 'Bot-SwingDir':
            bot += 1
        elif nct == 'TP':
            tp += 1; total += p
        elif nct == 'SL':
            sl += 1; total += p
        elif nct.startswith('Bot-'):
            bot += 1; total += p
        else:
            total += p
            if   p > 0: tp += 1
            elif p < 0: sl += 1
    wr = tp / (tp + sl) if (tp + sl) else 0
    return len(rows), tp, sl, bot, wr, round(total, 2)


def build_summary_sheet(ws, sid_rows, sids_present, start_date, end_date):
    cell(ws, 1, 1, f'All Strategies Compare  |  {start_date} → {end_date}  |  OLD vs NEW swing dir check',
         bold=True, bg=BG_TITLE, color='FFFFFF', align='left')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=14)
    cell(ws, 2, 1, '', bg=BG_HEAD)
    cell(ws, 2, 2, 'ACTUAL (OLD)', bold=True, bg=BG_OLD)
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=7)
    cell(ws, 2, 8, 'NEW (swing dir filter)', bold=True, bg=BG_NEW)
    ws.merge_cells(start_row=2, start_column=8, end_row=2, end_column=13)
    cell(ws, 2, 14, 'DIFF', bold=True, bg=BG_DIFF)

    ws.column_dimensions['A'].width = 22
    for c in 'BCDEFGHIJKLMN':
        ws.column_dimensions[c].width = 10
    for ci, h in enumerate(SUM_COLS, 1):
        bg = BG_OLD if 2 <= ci <= 7 else BG_NEW if 8 <= ci <= 13 else BG_DIFF if ci == 14 else BG_HEAD
        cell(ws, 3, ci, h, bold=True, bg=bg, border=MED_BOTTOM)

    sr = 4
    for sid in sids_present:
        rows = sid_rows[sid]
        n, tp, sl, bot, wr, pnl       = sid_stats(rows)
        nn, ntp, nsl, nbot, nwr, npnl = sid_new_stats(rows)
        diff = round(npnl - pnl, 2)
        vals = [SID_NAME.get(sid, f'S{sid}'), n, tp, sl, bot, wr, pnl,
                nn, ntp, nsl, nbot, nwr, npnl, diff]
        bgs  = [None]*6 + [pnl_bg(pnl)] + [None]*5 + [pnl_bg(npnl), pnl_bg(diff)]
        nfs  = [None, None, None, None, None, '0.0%', '#,##0.00',
                None, None, None, None, '0.0%', '#,##0.00', '#,##0.00']
        for ci, (v, bg, nf) in enumerate(zip(vals, bgs, nfs), 1):
            cell(ws, sr, ci, v, bg=bg, align='left' if ci == 1 else 'center', num_fmt=nf, border=THIN_BORDER)
        sr += 1

    cell(ws, sr, 1, 'TOTAL', bold=True, bg=BG_TOTAL, border=THIN_BORDER)
    for ci in range(2, 15):
        col = get_column_letter(ci)
        if ci in (6, 12):
            cell(ws, sr, ci, '-', bold=True, bg=BG_TOTAL, border=THIN_BORDER)
        elif ci == 14:
            cell(ws, sr, ci, f'=SUM({col}4:{col}{sr-1})', bold=True, bg=BG_DIFF, num_fmt='#,##0.00', border=THIN_BORDER)
        elif ci in (7, 13):
            cell(ws, sr, ci, f'=SUM({col}4:{col}{sr-1})', bold=True, bg=BG_TOTAL, num_fmt='#,##0.00', border=THIN_BORDER)
        else:
            cell(ws, sr, ci, f'=SUM({col}4:{col}{sr-1})', bold=True, bg=BG_TOTAL, border=THIN_BORDER)
    ws.freeze_panes = 'A4'


# ─── Main ────────────────────────────────────────────────────────────────────
def build_workbook(closed, start_date, end_date, out_path):
    sids_present = sorted(set(r['sid'] for r in closed if r['sid']), key=lambda x: int(x))
    sid_rows = {s: [] for s in sids_present}
    for row in closed:
        if row['sid']:
            sid_rows[row['sid']].append(row)

    wb = Workbook()
    wb.remove(wb.active)

    build_summary_sheet(wb.create_sheet('Summary'), sid_rows, sids_present, start_date, end_date)
    build_compare_sheet(wb.create_sheet('All_Compare'), closed,
                         f'All Strategies — {start_date} → {end_date} | OLD vs NEW swing dir check')
    for sid in sids_present:
        build_compare_sheet(wb.create_sheet(f'S{sid}_Compare'), sid_rows[sid],
                            f'{SID_NAME.get(sid, "S"+sid)} — {start_date} → {end_date} | OLD vs NEW swing dir check')
    wb.save(out_path)
    return wb.sheetnames, len(closed), sids_present


def main():
    today = _dt.date.today().isoformat()
    ap = argparse.ArgumentParser(description='สร้างไฟล์ Excel เปรียบเทียบ strategy ตามช่วงเวลา')
    ap.add_argument('--start', required=True, help='วันเริ่ม YYYY-MM-DD')
    ap.add_argument('--end', default=today, help='วันสิ้นสุด YYYY-MM-DD (default = วันนี้)')
    ap.add_argument('--out', default=None, help='path ไฟล์ output (.xlsx)')
    ap.add_argument('--logs', default=None, help='glob pattern ของ log (default = logs/old_logs/bot-*.log + logs/bot.log)')
    args = ap.parse_args()

    if args.logs:
        log_paths = sorted(glob.glob(args.logs))
    else:
        # ครอบทั้ง bot-YYYY-MM.log และไฟล์ rotate .log.bak-* (log rotate ทำให้ข้อมูลกระจาย)
        log_paths = sorted(set(
            glob.glob(str(ROOT / 'logs' / 'old_logs' / 'bot-*.log*'))
        ))
        live = ROOT / 'logs' / 'bot.log'
        if live.exists():
            log_paths.append(str(live))

    out = args.out or str(ROOT / 'excel_reports' / f'compare_all_{args.start}_{args.end}.xlsx')
    Path(out).parent.mkdir(parents=True, exist_ok=True)

    closed = parse_logs(log_paths, args.start, args.end)
    if not closed:
        print(f'No closed orders in range {args.start} -> {args.end} (logs: {len(log_paths)} files)')
        return

    sheets, n, sids = build_workbook(closed, args.start, args.end, out)
    print(f'OK saved: {out}')
    print(f'  range  : {args.start} -> {args.end}')
    print(f'  orders : {n}')
    print(f'  sids   : {", ".join("S"+s for s in sids)}')
    print(f'  sheets : {len(sheets)}')


if __name__ == '__main__':
    main()
