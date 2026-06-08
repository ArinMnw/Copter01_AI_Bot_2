"""สร้าง Excel sim_sweep_fix_june.xlsx จากผล sweep_filter fix simulation"""
import io, sys, os
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
os.chdir(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("pip install openpyxl"); sys.exit(1)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Sweep Fix Sim Jun2026"

# ── Colors ────────────────────────────────────────────────────────────
RED    = PatternFill("solid", fgColor="FFCCCC")
GREEN  = PatternFill("solid", fgColor="CCFFCC")
YELLOW = PatternFill("solid", fgColor="FFFACC")
GREY   = PatternFill("solid", fgColor="EEEEEE")
BLUE   = PatternFill("solid", fgColor="CCE5FF")
HEADER = PatternFill("solid", fgColor="2F4F8F")

bold = Font(bold=True)
white_bold = Font(bold=True, color="FFFFFF")
center = Alignment(horizontal="center", vertical="center")
right  = Alignment(horizontal="right")

def hdr(ws, row, col, val, fill=HEADER):
    c = ws.cell(row=row, column=col, value=val)
    c.font = white_bold; c.fill = fill; c.alignment = center
    return c

def cell(ws, row, col, val, fill=None, bold_=False, align=center):
    c = ws.cell(row=row, column=col, value=val)
    if fill: c.fill = fill
    if bold_: c.font = Font(bold=True)
    c.alignment = align
    return c

# ── Header ────────────────────────────────────────────────────────────
ws.merge_cells('A1:J1')
t = ws['A1']; t.value = "Sweep Filter Fix — Jun 2026 Simulation"
t.font = Font(bold=True, size=13, color="FFFFFF"); t.fill = HEADER; t.alignment = center
ws.row_dimensions[1].height = 25

headers = ['Ticket','TF','Signal','Why','Create','Status','P/L ก่อน fix','Fix?','P/L หลัง fix','ผลต่าง']
for col, h in enumerate(headers, 1):
    hdr(ws, 2, col, h)

# ── Data ──────────────────────────────────────────────────────────────
data = [
    ('537988219','M5','BUY','sweep_low','07:40 05-Jun','CLOSED',   -5.80, 'BLOCK', 0.00),
    ('537986275','M5','BUY','sweep_low','07:35 05-Jun','CLOSED',   -5.67, 'BLOCK', 0.00),
    ('537959048','M1','SELL','sweep_high','05:30 05-Jun','CLOSED',  2.24, 'BLOCK', 0.00),
    ('537959508','M1','SELL','sweep_high','05:35 05-Jun','CANCELED',0.00, 'BLOCK', 0.00),
    ('537958101','M1','BUY','sweep_low','05:23 05-Jun','CANCELED',  0.00, 'BLOCK', 0.00),
    ('538059433','M5','BUY','sweep_low','09:50 05-Jun','PENDING',   0.00, 'BLOCK', 0.00),
    ('538059434','M5','BUY','sweep_low','09:50 05-Jun','PENDING',   0.00, 'BLOCK', 0.00),
]

for r, (tk, tf, sig, why, create, status, pnl_before, fix, pnl_after) in enumerate(data, 3):
    diff = pnl_after - pnl_before
    row_fill = RED if pnl_before < 0 else (GREEN if pnl_before > 0 else GREY)

    cell(ws, r, 1, tk,         row_fill, align=right)
    cell(ws, r, 2, tf,         row_fill)
    sig_fill = BLUE if sig=='BUY' else RED
    cell(ws, r, 3, sig,        sig_fill, bold_=True)
    cell(ws, r, 4, why,        row_fill)
    cell(ws, r, 5, create,     row_fill)
    cell(ws, r, 6, status,     row_fill)
    c7 = cell(ws, r, 7, pnl_before, row_fill, bold_=(pnl_before!=0))
    c7.number_format = '+#,##0.00;-#,##0.00;0.00'
    fix_fill = GREEN if fix=='PASS' else PatternFill("solid", fgColor="FF9999")
    cell(ws, r, 8, fix, fix_fill, bold_=True)
    c9 = cell(ws, r, 9, pnl_after, GREY)
    c9.number_format = '+#,##0.00;-#,##0.00;0.00'
    diff_fill = GREEN if diff > 0 else (RED if diff < 0 else GREY)
    cd = cell(ws, r, 10, diff, diff_fill, bold_=(diff!=0))
    cd.number_format = '+#,##0.00;-#,##0.00;0.00'

# ── Summary ───────────────────────────────────────────────────────────
sr = len(data) + 3
ws.merge_cells(f'A{sr}:F{sr}')
c = ws[f'A{sr}']; c.value='TOTAL'; c.font=white_bold; c.fill=HEADER; c.alignment=center

total_before = sum(d[6] for d in data)
total_after  = sum(d[8] for d in data)
diff_total   = total_after - total_before

ct = cell(ws, sr, 7, total_before, HEADER, bold_=True)
ct.font=white_bold; ct.number_format='+#,##0.00;-#,##0.00;0.00'
cell(ws, sr, 8, '', HEADER)
ct2 = cell(ws, sr, 9, total_after, HEADER, bold_=True)
ct2.font=white_bold; ct2.number_format='+#,##0.00;-#,##0.00;0.00'
dt = cell(ws, sr, 10, diff_total, GREEN if diff_total>0 else RED, bold_=True)
dt.font=Font(bold=True); dt.number_format='+#,##0.00;-#,##0.00;0.00'

# สรุปล่าง
sr2 = sr + 2
ws.merge_cells(f'A{sr2}:J{sr2}')
note = ws[f'A{sr2}']
note.value = (f"Bug: sweep_filter ไม่เช็ค open > ref_price → orders ถูก unblock ผิด | "
              f"Fix: bo>ref (SWEEP_LOW) / bo<ref (SWEEP_HIGH) | "
              f"ผลต่าง: {diff_total:+.2f} USD ({len(data)} orders)")
note.font = Font(italic=True, size=9)
note.alignment = Alignment(wrap_text=True)
ws.row_dimensions[sr2].height = 30

# ── Column widths ─────────────────────────────────────────────────────
widths = [13,5,7,13,15,10,13,10,13,10]
for col, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(col)].width = w

out = os.path.join('excel_reports', 'sim_sweep_fix_june.xlsx')
wb.save(out)
print(f'Saved: {out}')
print(f'P/L ก่อน fix : {total_before:+.2f}')
print(f'P/L หลัง fix  : {total_after:+.2f}')
print(f'ผลต่าง        : {diff_total:+.2f}')
